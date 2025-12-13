import traceback
import csv
from datetime import datetime
from decimal import Decimal
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.cache import cache_page
from django.core.cache import cache
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q, Sum, Count, F, Case, When, Value, DecimalField
from django.db.models.functions import Coalesce, TruncMonth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from almacenes.models import MovimientoAlmacen, Almacen, DetalleMovimientoAlmacen
from beneficiarios.models import MovimientoCliente, Cliente, DetalleMovimientoCliente
from productos.models import Producto
from reportes.models import ReporteStock, ReporteEntregas, ReporteMovimiento, ReporteStockReal

# L√≠mite para exportaciones (seguridad y rendimiento)
MAX_EXPORT_ROWS = 25000  # Reducido de 50k a 25k para mejor rendimiento
#  HELPER: C√ÅLCULO MASIVO DE STOCK EST√ÅNDAR (OPTIMIZADO CON CACHE)
# ==============================================================================
def get_stock_bulk(almacen_id, producto_id=None):
    """
    Obtiene el stock f√≠sico del almac√©n desde la tabla StockCache pre-calculada.
    üöÄ OPTIMIZACI√ìN EXTREMA: Lectura instant√°nea desde cache pre-calculado
    """
    cache_key = f'stock_cache_bulk_{almacen_id}_{producto_id or "all"}'
    cached = cache.get(cache_key)
    if cached:
        return cached
    
    from stock_cache.models import StockCache

    # Query ultra-optimizada: solo campos necesarios, sin select_related para velocidad m√°xima
    queryset = StockCache.objects.filter(almacen_id=almacen_id)

    if producto_id:
        queryset = queryset.filter(producto_id=producto_id)

    # Ejecutar query y construir resultado
    result = {}
    for stock_entry in queryset.values('producto_id', 'stock_bueno', 'stock_danado', 'stock_total'):
        result[stock_entry['producto_id']] = {
            'stock_bueno': stock_entry['stock_bueno'],
            'stock_danado': stock_entry['stock_danado'],
            'stock_total': stock_entry['stock_total'],
            'data': {}  # Mantener compatibilidad con c√≥digo existente
        }

    # Cache por 1 hora para m√°xima velocidad
    cache.set(cache_key, result, 3600)
    return result

# ==============================================================================
#  HELPER: C√ÅLCULO MASIVO DE STOCK REAL (OPTIMIZACI√ìN EXTREMA)
# ==============================================================================
def get_stock_real_bulk(almacen_id, producto_id=None):
    """
    Calcula el stock real del almac√©n usando StockCache + ajustes de cliente.
    üöÄ OPTIMIZACI√ìN EXTREMA: Todo desde cache pre-calculado
    """
    cache_key = f'stock_real_cache_bulk_{almacen_id}_{producto_id or "all"}'
    cached = cache.get(cache_key)
    if cached:
        return cached

    from stock_cache.models import StockCache
    from django.db import connection

    # Calcular componentes detallados de almac√©n (con cache separado)
    cache_key_alm = f'stock_real_alm_{almacen_id}_{producto_id or "all"}'
    componentes_almacen = cache.get(cache_key_alm)

    if componentes_almacen is None:
        sql_alm = """
        SELECT
            d.producto_id,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id = %s THEN d.cantidad ELSE 0 END
            ), 0) AS ent_alm_b,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id = %s THEN d.cantidad_danada ELSE 0 END
            ), 0) AS ent_alm_d,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id = %s THEN d.cantidad ELSE 0 END
            ), 0) AS sal_alm_b,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id = %s THEN d.cantidad_danada ELSE 0 END
            ), 0) AS sal_alm_d,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id = %s THEN d.cantidad ELSE 0 END
            ), 0) AS tras_rec_b,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id = %s THEN d.cantidad_danada ELSE 0 END
            ), 0) AS tras_rec_d,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id = %s THEN d.cantidad ELSE 0 END
            ), 0) AS tras_env_b,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id = %s THEN d.cantidad_danada ELSE 0 END
            ), 0) AS tras_env_d
        FROM almacenes_detallemovimientoalmacen d
        JOIN almacenes_movimientoalmacen m ON d.movimiento_id = m.id
        WHERE m.almacen_origen_id = %s OR m.almacen_destino_id = %s
        """
        params_alm = [almacen_id] * 10

        if producto_id:
            sql_alm += " AND d.producto_id = %s"
            params_alm.append(producto_id)

        sql_alm += " GROUP BY d.producto_id"

        with connection.cursor() as cursor:
            cursor.execute(sql_alm, params_alm)
            rows_alm = cursor.fetchall()

        componentes_almacen = {row[0]: {
            'ent_alm_b': row[1], 'ent_alm_d': row[2],
            'sal_alm_b': row[3], 'sal_alm_d': row[4],
            'tras_rec_b': row[5], 'tras_rec_d': row[6],
            'tras_env_b': row[7], 'tras_env_d': row[8]
        } for row in rows_alm}
        cache.set(cache_key_alm, componentes_almacen, 3600)  # Cache 1 hora

    # Calcular ajustes de movimientos de cliente (con cache separado)
    cache_key_cli = f'stock_real_cli_{almacen_id}_{producto_id or "all"}'
    ajustes_cliente = cache.get(cache_key_cli)

    if ajustes_cliente is None:
        sql_cli = """
        SELECT
            d.producto_id,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'ENTRADA' AND m.almacen_origen_id = %s THEN d.cantidad ELSE 0 END
            ), 0) AS ent_cli_b,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'ENTRADA' AND m.almacen_origen_id = %s THEN d.cantidad_danada ELSE 0 END
            ), 0) AS ent_cli_d,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'SALIDA' AND m.almacen_destino_id = %s THEN d.cantidad ELSE 0 END
            ), 0) AS sal_cli_b,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo = 'SALIDA' AND m.almacen_destino_id = %s THEN d.cantidad_danada ELSE 0 END
            ), 0) AS sal_cli_d
        FROM beneficiarios_detallemovimientocliente d
        JOIN beneficiarios_movimientocliente m ON d.movimiento_id = m.id
        WHERE (m.almacen_origen_id = %s OR m.almacen_destino_id = %s) AND m.tipo != 'TRASLADO'
        """
        params_cli = [almacen_id] * 6

        if producto_id:
            sql_cli += " AND d.producto_id = %s"
            params_cli.append(producto_id)

        sql_cli += " GROUP BY d.producto_id"

        with connection.cursor() as cursor:
            cursor.execute(sql_cli, params_cli)
            rows_cli = cursor.fetchall()

        ajustes_cliente = {row[0]: {
            'ent_cli_b': row[1], 'ent_cli_d': row[2],
            'sal_cli_b': row[3], 'sal_cli_d': row[4]
        } for row in rows_cli}
        cache.set(cache_key_cli, ajustes_cliente, 3600)  # Cache 1 hora

    # Combinar resultados de forma eficiente
    result = {}

    # Procesar todos los productos que tienen componentes
    all_productos = set(componentes_almacen.keys()) | set(ajustes_cliente.keys())
    
    for pid in all_productos:
        vals_alm = componentes_almacen.get(pid, {
            'ent_alm_b': Decimal(0), 'ent_alm_d': Decimal(0),
            'sal_alm_b': Decimal(0), 'sal_alm_d': Decimal(0),
            'tras_rec_b': Decimal(0), 'tras_rec_d': Decimal(0),
            'tras_env_b': Decimal(0), 'tras_env_d': Decimal(0)
        })
        vals_cli = ajustes_cliente.get(pid, {
            'ent_cli_b': Decimal(0), 'ent_cli_d': Decimal(0),
            'sal_cli_b': Decimal(0), 'sal_cli_d': Decimal(0)
        })

        # Calcular stock total desde componentes
        stock_bueno = (vals_alm['ent_alm_b'] + vals_alm['tras_rec_b'] + vals_cli['ent_cli_b'] - 
                      vals_alm['sal_alm_b'] - vals_alm['tras_env_b'] - vals_cli['sal_cli_b'])
        stock_danado = (vals_alm['ent_alm_d'] + vals_alm['tras_rec_d'] + vals_cli['ent_cli_d'] - 
                       vals_alm['sal_alm_d'] - vals_alm['tras_env_d'] - vals_cli['sal_cli_d'])

        result[pid] = {
            'stock_bueno': stock_bueno,
            'stock_danado': stock_danado,
            'stock_total': stock_bueno + stock_danado,
            'data': {
                **vals_alm,
                **vals_cli
            }
        }

    # Cache del resultado final
    cache.set(cache_key, result, 3600)
    return result


@staff_member_required
def obtener_datos_graficos_movimientos(request):
    """
    Retorna datos agregados por mes para graficar Entradas vs Salidas.
    ‚úÖ CORREGIDO: Usa los nombres correctos de campos
    """
    try:
        # Cache key basado en filtros
        cache_key = f"graficos_mov_{request.GET.get('fecha_inicio')}_{request.GET.get('fecha_fin')}_{request.GET.get('almacen')}_{request.GET.get('proveedor')}_{request.GET.get('recepcionista')}"
        cached = cache.get(cache_key)
        if cached:
            return JsonResponse(cached)
        
        # 1. Obtener filtros de la solicitud
        fecha_inicio_str = request.GET.get('fecha_inicio')
        fecha_fin_str = request.GET.get('fecha_fin')
        almacen_id = request.GET.get('almacen')
        proveedor_id = request.GET.get('proveedor')
        recepcionista_id = request.GET.get('recepcionista')

        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else None
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else None

        from reportes.models import ReporteMovimiento
        from django.db.models.functions import Coalesce
        
        # 2. Obtener movimientos de ALMAC√âN agregados por mes
        movimientos_almacen = ReporteMovimiento.obtener_movimientos_almacen(
            fecha_inicio=fecha_inicio, 
            fecha_fin=fecha_fin, 
            almacen=almacen_id,
            proveedor=proveedor_id, 
            recepcionista=recepcionista_id
        ).annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            # ‚úÖ CORRECCI√ìN: Usar 'cantidad' en lugar de 'cantidad_buena'
            entradas_almacen=Coalesce(
                Sum('detalles__cantidad', filter=Q(tipo='ENTRADA')),
                Decimal('0')
            ) + Coalesce(
                Sum('detalles__cantidad_danada', filter=Q(tipo='ENTRADA')),
                Decimal('0')
            ),
            salidas_almacen=Coalesce(
                Sum('detalles__cantidad', filter=Q(tipo='SALIDA')),
                Decimal('0')
            ) + Coalesce(
                Sum('detalles__cantidad_danada', filter=Q(tipo='SALIDA')),
                Decimal('0')
            ),
            traslados_enviados=Coalesce(
                Sum('detalles__cantidad', filter=Q(tipo='TRASLADO')),
                Decimal('0')
            ) + Coalesce(
                Sum('detalles__cantidad_danada', filter=Q(tipo='TRASLADO')),
                Decimal('0')
            )
        ).order_by('mes')
        
        # 3. Procesar datos para Chart.js
        fechas = []
        entradas = []
        salidas = []
        traslados = []
        
        for item in movimientos_almacen:
            mes_str = item['mes'].strftime('%Y-%m')
            fechas.append(mes_str)
            entradas.append(float(item['entradas_almacen'] or 0))
            salidas.append(float(item['salidas_almacen'] or 0))
            traslados.append(float(item['traslados_enviados'] or 0))
            
        # 4. Estructura de datos para Chart.js
        datos = {
            'labels': fechas,
            'datasets': [
                {
                    'label': 'Entradas',
                    'data': entradas,
                    'backgroundColor': 'rgba(46, 204, 113, 0.2)',
                    'borderColor': 'rgba(46, 204, 113, 1)',
                    'borderWidth': 2,
                    'fill': True,
                    'tension': 0.4,
                },
                {
                    'label': 'Salidas',
                    'data': salidas,
                    'backgroundColor': 'rgba(231, 76, 60, 0.2)',
                    'borderColor': 'rgba(231, 76, 60, 1)',
                    'borderWidth': 2,
                    'fill': True,
                    'tension': 0.4,
                },
                {
                    'label': 'Traslados',
                    'data': traslados,
                    'backgroundColor': 'rgba(52, 152, 219, 0.2)',
                    'borderColor': 'rgba(52, 152, 219, 1)',
                    'borderWidth': 2,
                    'fill': True,
                    'tension': 0.4,
                }
            ]
        }

        cache.set(cache_key, datos, 300)  # Cache 5 minutos
        return JsonResponse(datos)

    except Exception as e:
        print(f"‚ùå ERROR en obtener_datos_graficos_movimientos: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': str(e), 
            'traceback': traceback.format_exc()
        }, status=500)

def exportar_movimientos_excel(request):
    """Exporta los movimientos filtrados a Excel"""
    
# Obtener filtros
    tipo_reporte = request.GET.get('tipo_reporte', 'almacen')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    almacen_id = request.GET.get('almacen')
    cliente_id = request.GET.get('cliente')
    proveedor_id = request.GET.get('proveedor')
    recepcionista_id = request.GET.get('recepcionista')
    producto_id = request.GET.get('producto')
    numero_movimiento = request.GET.get('numero_movimiento', '').strip() # ‚≠ê CORRECCI√ìN: Agregar filtro de n√∫mero
    
    # Convertir fechas
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Construir queryset seg√∫n tipo de reporte
    if tipo_reporte == 'almacen':
        movimientos = MovimientoAlmacen.objects.all()
        
        if fecha_inicio_obj:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos = movimientos.filter(fecha__lte=fecha_fin_obj)
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento)
        if almacen_id:
            movimientos = movimientos.filter(
                Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
            )
        if proveedor_id:
            movimientos = movimientos.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos = movimientos.filter(recepcionista_id=recepcionista_id)
        if producto_id:
            movimientos = movimientos.filter(detalles__producto_id=producto_id).distinct()
        
# ‚≠ê CORRECCI√ìN: Aplicar filtro de n√∫mero de movimiento
        if numero_movimiento:
            movimientos = movimientos.filter(numero_movimiento__icontains=numero_movimiento)
            
        movimientos = movimientos.select_related(
            'almacen_origen', 'almacen_destino', 'proveedor', 'recepcionista'
        ).prefetch_related('detalles__producto__unidad_medida').annotate(
            total_detalles=Count('detalles')
        ).order_by('-fecha')[:MAX_EXPORT_ROWS]
        
    else:  # cliente
        movimientos = MovimientoCliente.objects.all()
        
        if fecha_inicio_obj:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos = movimientos.filter(fecha__lte=fecha_fin_obj)
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento)
        if cliente_id:
            movimientos = movimientos.filter(cliente_id=cliente_id)
        if almacen_id:
            movimientos = movimientos.filter(
                Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
            )
        if proveedor_id:
            movimientos = movimientos.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos = movimientos.filter(recepcionista_id=recepcionista_id)
        if producto_id:
            movimientos = movimientos.filter(detalles__producto_id=producto_id).distinct()
        
 # ‚≠ê CORRECCI√ìN: Aplicar filtro de n√∫mero de movimiento
        if numero_movimiento:
            movimientos = movimientos.filter(numero_movimiento__icontains=numero_movimiento)
        
        movimientos = movimientos.select_related(
            'cliente', 'almacen_origen', 'almacen_destino', 'proveedor', 'recepcionista'
        ).prefetch_related('detalles__producto__unidad_medida').annotate(
            total_detalles=Count('detalles')
        ).order_by('-fecha')[:MAX_EXPORT_ROWS]
    
    # Crear workbook
    wb = Workbook(write_only=True)
    ws1 = wb.active
    ws1.title = "Resumen Movimientos"
    
    # Estilos
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados hoja 1
    if tipo_reporte == 'almacen':
        headers = ['N¬∞ Movimiento', 'Tipo', 'Fecha', 'Almac√©n Origen', 'Almac√©n Destino', 
                   'Proveedor', 'Recepcionista', 'Total Productos']
    else:
        headers = ['N¬∞ Movimiento', 'Tipo', 'Fecha', 'Cliente', 'Almac√©n Origen', 
                   'Almac√©n Destino', 'Proveedor', 'Recepcionista', 'Total Productos']
    
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos hoja 1
    for row_idx, movimiento in enumerate(movimientos, start=2):
        if tipo_reporte == 'almacen':
            row_data = [
                movimiento.numero_movimiento,
                movimiento.get_tipo_display(),
                movimiento.fecha.strftime('%d/%m/%Y'),
                str(movimiento.almacen_origen) if movimiento.almacen_origen else '-',
                str(movimiento.almacen_destino) if movimiento.almacen_destino else '-',
                str(movimiento.proveedor) if movimiento.proveedor else '-',
                str(movimiento.recepcionista) if movimiento.recepcionista else '-',
                movimiento.total_detalles
            ]
        else:
            row_data = [
                movimiento.numero_movimiento,
                movimiento.get_tipo_display(),
                movimiento.fecha.strftime('%d/%m/%Y'),
                str(movimiento.cliente) if movimiento.cliente else '-',
                str(movimiento.almacen_origen) if movimiento.almacen_origen else '-',
                str(movimiento.almacen_destino) if movimiento.almacen_destino else '-',
                str(movimiento.proveedor) if movimiento.proveedor else '-',
                str(movimiento.recepcionista) if movimiento.recepcionista else '-',
                movimiento.total_detalles
            ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Hoja 2: Detalle de productos
    ws2 = wb.create_sheet(title="Detalle Productos")
    
    detail_headers = ['N¬∞ Movimiento', 'Fecha', 'Tipo', 'C√≥digo', 'Producto', 
                      'Cant. Buena', 'Cant. Da√±ada', 'Total', 'Unidad']
    
    for col, header in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    detail_row = 2
    for movimiento in movimientos:
        for detalle in movimiento.detalles.all():
            total = (detalle.cantidad or 0) + (detalle.cantidad_danada or 0)
            unidad = detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND'
            
            row_data = [
                movimiento.numero_movimiento,
                movimiento.fecha.strftime('%d/%m/%Y'),
                movimiento.get_tipo_display(),
                detalle.producto.codigo,
                detalle.producto.nombre,
                detalle.cantidad or 0,
                detalle.cantidad_danada or 0,
                total,
                unidad
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=detail_row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            detail_row += 1
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'movimientos_{tipo_reporte}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def exportar_movimientos_csv(request):
    """Exporta los movimientos filtrados a CSV"""
    
# Obtener filtros
    tipo_reporte = request.GET.get('tipo_reporte', 'almacen')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    almacen_id = request.GET.get('almacen')
    cliente_id = request.GET.get('cliente')
    proveedor_id = request.GET.get('proveedor')
    recepcionista_id = request.GET.get('recepcionista')
    producto_id = request.GET.get('producto')
    numero_movimiento = request.GET.get('numero_movimiento', '').strip() # ‚≠ê CORRECCI√ìN: Agregar filtro de n√∫mero
    
    # Convertir fechas
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Construir queryset
    if tipo_reporte == 'almacen':
        movimientos = MovimientoAlmacen.objects.all()
        
        if fecha_inicio_obj:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos = movimientos.filter(fecha__lte=fecha_fin_obj)
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento)
        if almacen_id:
            movimientos = movimientos.filter(
                Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
            )
        if proveedor_id:
            movimientos = movimientos.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos = movimientos.filter(recepcionista_id=recepcionista_id)
        if producto_id:
            movimientos = movimientos.filter(detalles__producto_id=producto_id).distinct()
        
# ‚≠ê CORRECCI√ìN: Aplicar filtro de n√∫mero de movimiento
        if numero_movimiento:
            movimientos = movimientos.filter(numero_movimiento__icontains=numero_movimiento)

        movimientos = movimientos.select_related(
            'almacen_origen', 'almacen_destino', 'proveedor', 'recepcionista'
        ).prefetch_related('detalles__producto__unidad_medida').annotate(
            total_detalles=Count('detalles')
        ).order_by('-fecha')[:MAX_EXPORT_ROWS]
        
    else:  # cliente
        movimientos = MovimientoCliente.objects.all()
        
        if fecha_inicio_obj:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos = movimientos.filter(fecha__lte=fecha_fin_obj)
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento)
        if cliente_id:
            movimientos = movimientos.filter(cliente_id=cliente_id)
        if almacen_id:
            movimientos = movimientos.filter(
                Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
            )
        if proveedor_id:
            movimientos = movimientos.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos = movimientos.filter(recepcionista_id=recepcionista_id)
        if producto_id:
            movimientos = movimientos.filter(detalles__producto_id=producto_id).distinct()
        
 # ‚≠ê CORRECCI√ìN: Aplicar filtro de n√∫mero de movimiento
        if numero_movimiento:
            movimientos = movimientos.filter(numero_movimiento__icontains=numero_movimiento)

        movimientos = movimientos.select_related(
            'cliente', 'almacen_origen', 'almacen_destino', 'proveedor', 'recepcionista'
        ).prefetch_related('detalles__producto__unidad_medida').annotate(
            total_detalles=Count('detalles')
        ).order_by('-fecha')[:MAX_EXPORT_ROWS]
    
    # Crear respuesta CSV con streaming
    def csv_generator():
        import io
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # BOM para Excel
        yield '\ufeff'
        
        # Encabezados
        if tipo_reporte == 'almacen':
            writer.writerow(['N¬∞ Movimiento', 'Tipo', 'Fecha', 'Almac√©n Origen', 'Almac√©n Destino',
                            'Proveedor', 'Recepcionista', 'C√≥digo', 'Producto',
                            'Cant. Buena', 'Cant. Da√±ada', 'Total', 'Unidad'])
        else:
            writer.writerow(['N¬∞ Movimiento', 'Tipo', 'Fecha', 'Cliente', 'Almac√©n Origen',
                            'Almac√©n Destino', 'Proveedor', 'Recepcionista', 'C√≥digo',
                            'Producto', 'Cant. Buena', 'Cant. Da√±ada', 'Total', 'Unidad'])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        # Datos
        for movimiento in movimientos:
            for detalle in movimiento.detalles.all():
                total = (detalle.cantidad or 0) + (detalle.cantidad_danada or 0)
                unidad = detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND'
                
                if tipo_reporte == 'almacen':
                    writer.writerow([
                        movimiento.numero_movimiento,
                        movimiento.get_tipo_display(),
                        movimiento.fecha.strftime('%d/%m/%Y'),
                        str(movimiento.almacen_origen) if movimiento.almacen_origen else '-',
                        str(movimiento.almacen_destino) if movimiento.almacen_destino else '-',
                        str(movimiento.proveedor) if movimiento.proveedor else '-',
                        str(movimiento.recepcionista) if movimiento.recepcionista else '-',
                        detalle.producto.codigo,
                        detalle.producto.nombre,
                        detalle.cantidad or 0,
                        detalle.cantidad_danada or 0,
                        total,
                        unidad
                    ])
                else:
                    writer.writerow([
                        movimiento.numero_movimiento,
                        movimiento.get_tipo_display(),
                        movimiento.fecha.strftime('%d/%m/%Y'),
                        str(movimiento.cliente) if movimiento.cliente else '-',
                        str(movimiento.almacen_origen) if movimiento.almacen_origen else '-',
                        str(movimiento.almacen_destino) if movimiento.almacen_destino else '-',
                        str(movimiento.proveedor) if movimiento.proveedor else '-',
                        str(movimiento.recepcionista) if movimiento.recepcionista else '-',
                        detalle.producto.codigo,
                        detalle.producto.nombre,
                        detalle.cantidad or 0,
                        detalle.cantidad_danada or 0,
                        total,
                        unidad
                    ])
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)
    
    response = StreamingHttpResponse(
        csv_generator(),
        content_type='text/csv; charset=utf-8'
    )
    filename = f'movimientos_{tipo_reporte}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

# ========================================
# NUEVA FUNCI√ìN - AGREGAR AL FINAL
# ========================================
@staff_member_required
def obtener_detalle_stock(request):
    """
    Vista optimizada para detalle de stock de un producto en un almac√©n.
    """
    producto_id = request.GET.get('producto_id')
    almacen_id = request.GET.get('almacen_id')
    
    if not producto_id or not almacen_id:
        return JsonResponse({'success': False, 'error': 'Faltan par√°metros'})
    
    try:
        producto = Producto.objects.get(id=producto_id)
        almacen = Almacen.objects.get(id=almacen_id)
        
        # 1. C√°lculo r√°pido usando el helper
        calc_bulk = get_stock_bulk(almacen_id, producto_id)
        stock_data = calc_bulk.get(int(producto_id))
        
        if not stock_data:
            # Inicializar en cero si no hay movimientos
            stock_data = {
                'stock_bueno': Decimal(0), 'stock_danado': Decimal(0), 'stock_total': Decimal(0),
                'data': {k: Decimal(0) for k in ['ent_b','ent_d','sal_b','sal_d','tras_rec_b','tras_rec_d','tras_env_b','tras_env_d']}
            }
            
        d = stock_data['data']
        
        # 2. Obtener lista de movimientos (Limitada a los √∫ltimos 200 para velocidad)
        movimientos = MovimientoAlmacen.objects.filter(
            Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id),
            detalles__producto_id=producto_id
        ).annotate(
            cant_b=Sum('detalles__cantidad', filter=Q(detalles__producto_id=producto_id)),
            cant_d=Sum('detalles__cantidad_danada', filter=Q(detalles__producto_id=producto_id)),
            origen_nombre=F('almacen_origen__nombre'),
            destino_nombre=F('almacen_destino__nombre'),
            prov_nombre=F('proveedor__nombre'),
            rec_nombre=F('recepcionista__nombre')
        ).exclude(cant_b=None, cant_d=None).order_by('-fecha', '-id')[:200]
        
        movimientos_list = []
        for mov in movimientos:
            cb = mov.cant_b or Decimal(0)
            cd = mov.cant_d or Decimal(0)
            total_mov = cb + cd
            
            # Determinar signo visual para la tabla
            signo = 1
            if mov.tipo == 'SALIDA' or (mov.tipo == 'TRASLADO' and mov.almacen_origen_id == int(almacen_id)):
                signo = -1
                
            estado = 'MIXTO' if cb > 0 and cd > 0 else ('BUENO' if cb > 0 else ('DA√ëADO' if cd > 0 else '-'))
            
            movimientos_list.append({
                'numero_movimiento': mov.numero_movimiento,
                'tipo': mov.get_tipo_display(),
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M'),
                'cantidad': str(total_mov * signo), # Visual con signo
                'cantidad_buena': str(cb),
                'cantidad_danada': str(cd),
                'estado': estado,
                'proveedor': mov.prov_nombre,
                'recepcionista': mov.rec_nombre,
                'almacen_origen': mov.origen_nombre,
                'almacen_destino': mov.destino_nombre,
            })
            
        return JsonResponse({
            'success': True,
            'producto': {
                'nombre': producto.nombre,
                'codigo': producto.codigo,
                'categoria': producto.categoria.nombre if producto.categoria else None,
                'unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND'
            },
            'almacen': {'nombre': almacen.nombre},
            'resumen': {
                'total_entradas': str(d['ent_b'] + d['ent_d']),
                'total_salidas': str(d['sal_b'] + d['sal_d']),
                'traslados_recibidos': str(d['tras_rec_b'] + d['tras_rec_d']),
                'traslados_enviados': str(d['tras_env_b'] + d['tras_env_d']),
                'stock_bueno': str(stock_data['stock_bueno']),
                'stock_danado': str(stock_data['stock_danado']),
                'stock_total': str(stock_data['stock_total'])
            },
            'movimientos': movimientos_list
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'traceback': traceback.format_exc()})

# ============================================================
# CORRECCI√ìN 3: views.py - obtener_detalle_almacen (l√≠nea ~550)
# ============================================================
@staff_member_required
def obtener_detalle_almacen(request):
    """
    Vista optimizada para listar todos los productos de un almac√©n con su stock.
    Si no se especifica almacen_id, devuelve lista vac√≠a (lazy loading).
    """
    almacen_id = request.GET.get('almacen_id')
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 50))
    
    # Si no hay almacen_id, devolver vac√≠o (no error)
    if not almacen_id:
        return JsonResponse({
            'success': True,
            'almacen': None,
            'total_productos': 0,
            'productos': [],
            'page': page,
            'limit': limit,
            'total_pages': 0,
            'message': 'Selecciona un almac√©n para ver el stock'
        })
    
    try:
        almacen = Almacen.objects.get(id=almacen_id)
        
        # 1. Obtener c√°lculo masivo
        bulk_stocks = get_stock_bulk(almacen_id)
        
        if not bulk_stocks:
             return JsonResponse({
                'success': True,
                'almacen': {'id': almacen.id, 'nombre': almacen.nombre},
                'total_productos': 0,
                'productos': [],
                'page': page,
                'limit': limit,
                'total_pages': 0
            })

        # 2. Obtener info de productos en una sola consulta
        productos_ids = list(bulk_stocks.keys())
        productos_info = Producto.objects.filter(id__in=productos_ids).values(
            'id', 'codigo', 'nombre', 'categoria__nombre', 'unidad_medida__abreviatura'
        )
        productos_map = {p['id']: p for p in productos_info}
        
        # 3. Construir lista
        productos_list = []
        for pid, data in bulk_stocks.items():
            if pid not in productos_map: continue
            
            p_info = productos_map[pid]
            d = data['data']
            
            productos_list.append({
                'producto_id': pid,
                'codigo': p_info['codigo'],
                'nombre': p_info['nombre'],
                'categoria': p_info['categoria__nombre'] or '-',
                'unidad': p_info['unidad_medida__abreviatura'] or 'UND',
                'stock_bueno': str(data['stock_bueno']),
                'stock_danado': str(data['stock_danado']),
                'stock_total': str(data['stock_total']),
                'total_entradas': str(d['ent_b'] + d['ent_d']),
                'total_salidas': str(d['sal_b'] + d['sal_d']),
                'total_traslados_recibidos': str(d['tras_rec_b'] + d['tras_rec_d']),
                'total_traslados_enviados': str(d['tras_env_b'] + d['tras_env_d']),
            })
            
        productos_list.sort(key=lambda x: x['codigo'])
        
        # 4. Paginaci√≥n
        total_productos = len(productos_list)
        start = (page - 1) * limit
        end = start + limit
        paginated_products = productos_list[start:end]
        total_pages = (total_productos + limit - 1) // limit
        
        return JsonResponse({
            'success': True,
            'almacen': {'id': almacen.id, 'nombre': almacen.nombre},
            'total_productos': total_productos,
            'productos': paginated_products,
            'page': page,
            'limit': limit,
            'total_pages': total_pages
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ============================================================
# CORRECCI√ìN 4: views.py - obtener_detalle_producto_almacenes (l√≠nea ~700)
# ============================================================
@staff_member_required
def obtener_detalle_producto_almacenes(request):
    """
    Vista optimizada: Muestra el stock de un producto en TODOS los almacenes.
    """
    producto_id = request.GET.get('producto_id')
    if not producto_id:
        return JsonResponse({'success': False, 'error': 'Falta producto_id'})
    
@staff_member_required
def obtener_detalle_producto_almacenes(request):
    """
    Vista optimizada: Muestra el stock de un producto en TODOS los almacenes.
    Si no se especifica producto_id, devuelve lista vac√≠a (lazy loading).
    """
    producto_id = request.GET.get('producto_id')
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 50))
    
    # Si no hay producto_id, devolver vac√≠o (no error)
    if not producto_id:
        return JsonResponse({
            'success': True,
            'producto': None,
            'total_almacenes': 0,
            'almacenes': [],
            'page': page,
            'limit': limit,
            'total_pages': 0,
            'totales': {'stock_bueno': '0', 'stock_danado': '0', 'stock_total': '0'},
            'message': 'Selecciona un producto para ver su stock en almacenes'
        })
    
    try:
        producto = Producto.objects.get(id=producto_id)
        almacenes = Almacen.objects.filter(activo=True)
        
        almacenes_list = []
        total_sb = Decimal(0)
        total_sd = Decimal(0)
        total_st = Decimal(0)
        
        for almacen in almacenes:
            # Usamos el helper filtrado por producto (muy r√°pido)
            calc = get_stock_bulk(almacen.id, producto_id)
            data = calc.get(int(producto_id))
            
            if data and data['stock_total'] != 0:
                d = data['data']
                almacenes_list.append({
                    'almacen_id': almacen.id,
                    'almacen_nombre': almacen.nombre,
                    'stock_bueno': str(data['stock_bueno']),
                    'stock_danado': str(data['stock_danado']),
                    'stock_total': str(data['stock_total']),
                    'total_entradas': str(d['ent_b'] + d['ent_d']),
                    'total_salidas': str(d['sal_b'] + d['sal_d']),
                    'total_traslados_recibidos': str(d['tras_rec_b'] + d['tras_rec_d']),
                    'total_traslados_enviados': str(d['tras_env_b'] + d['tras_env_d']),
                })
                total_sb += data['stock_bueno']
                total_sd += data['stock_danado']
                total_st += data['stock_total']
        
        # Paginaci√≥n
        total_almacenes = len(almacenes_list)
        start = (page - 1) * limit
        end = start + limit
        paginated_almacenes = almacenes_list[start:end]
        total_pages = (total_almacenes + limit - 1) // limit
        
        return JsonResponse({
            'success': True,
            'producto': {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else '-',
                'unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND'
            },
            'total_almacenes': total_almacenes,
            'almacenes': paginated_almacenes,
            'page': page,
            'limit': limit,
            'total_pages': total_pages,
            'totales': {
                'stock_bueno': str(total_sb),
                'stock_danado': str(total_sd),
                'stock_total': str(total_st)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==========================================
# NUEVA FUNCI√ìN: API para n√∫meros de movimiento
# ==========================================
@staff_member_required
def obtener_numeros_movimiento_json(request):
    """
    Retorna una lista de n√∫meros de movimiento √∫nicos basada en filtros.
    """
    tipo_reporte = request.GET.get('tipo_reporte', 'almacen')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    almacen_id = request.GET.get('almacen')
    cliente_id = request.GET.get('cliente')
    proveedor_id = request.GET.get('proveedor')
    recepcionista_id = request.GET.get('recepcionista')

    # Convertir fechas
    fecha_inicio_obj = None
    fecha_fin_obj = None
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if tipo_reporte == 'almacen':
        movimientos_query = MovimientoAlmacen.objects.all()
        
        # Aplicar filtros existentes
        if fecha_inicio_obj:
            movimientos_query = movimientos_query.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos_query = movimientos_query.filter(fecha__lte=fecha_fin_obj)
        if almacen_id:
            movimientos_query = movimientos_query.filter(
                Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
            )
        if tipo_movimiento:
            movimientos_query = movimientos_query.filter(tipo=tipo_movimiento)
        if proveedor_id:
            movimientos_query = movimientos_query.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos_query = movimientos_query.filter(recepcionista_id=recepcionista_id)
            
    else:
        movimientos_query = MovimientoCliente.objects.all()
        
        # Aplicar filtros existentes
        if fecha_inicio_obj:
            movimientos_query = movimientos_query.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos_query = movimientos_query.filter(fecha__lte=fecha_fin_obj)
        if cliente_id:
            movimientos_query = movimientos_query.filter(cliente_id=cliente_id)
        if tipo_movimiento:
            movimientos_query = movimientos_query.filter(tipo=tipo_movimiento)
        if proveedor_id:
            movimientos_query = movimientos_query.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos_query = movimientos_query.filter(recepcionista_id=recepcionista_id)
            
    # Obtener los n√∫meros de movimiento, limitar a 200 y ordenar
    numeros = movimientos_query.values_list('numero_movimiento', flat=True).distinct().order_by('-numero_movimiento')[:200]
    
    return JsonResponse({
        'numeros': list(numeros)
    })

# ... (funci√≥n exportar_movimientos_excel ya est√° implementada) ...

@staff_member_required
def exportar_movimientos_csv(request):
    """Exporta los movimientos filtrados a CSV"""
    
    # Obtener filtros (similar a exportar_movimientos_excel)
    tipo_reporte = request.GET.get('tipo_reporte', 'almacen')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    almacen_id = request.GET.get('almacen')
    cliente_id = request.GET.get('cliente')
    proveedor_id = request.GET.get('proveedor')
    recepcionista_id = request.GET.get('recepcionista')
    producto_id = request.GET.get('producto')
    numero_movimiento = request.GET.get('numero_movimiento', '').strip()

    # Convertir fechas (similar a exportar_movimientos_excel)
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Construir queryset seg√∫n tipo de reporte (similar a exportar_movimientos_excel)
    if tipo_reporte == 'almacen':
        movimientos = MovimientoAlmacen.objects.all()
        
        if fecha_inicio_obj:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos = movimientos.filter(fecha__lte=fecha_fin_obj)
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento)
        if almacen_id:
            movimientos = movimientos.filter(
                Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
            )
        if proveedor_id:
            movimientos = movimientos.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos = movimientos.filter(recepcionista_id=recepcionista_id)
        if producto_id:
            movimientos = movimientos.filter(detalles__producto_id=producto_id).distinct()
        if numero_movimiento:
            movimientos = movimientos.filter(numero_movimiento__icontains=numero_movimiento)
            
        movimientos = movimientos.select_related(
            'almacen_origen', 'almacen_destino', 'proveedor', 'recepcionista'
        ).prefetch_related('detalles__producto__unidad_medida').order_by('-fecha')
        
    else:  # cliente
        movimientos = MovimientoCliente.objects.all()
        
        if fecha_inicio_obj:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos = movimientos.filter(fecha__lte=fecha_fin_obj)
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento)
        if cliente_id:
            movimientos = movimientos.filter(cliente_id=cliente_id)
        if almacen_id:
            movimientos = movimientos.filter(
                Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
            )
        if proveedor_id:
            movimientos = movimientos.filter(proveedor_id=proveedor_id)
        if recepcionista_id:
            movimientos = movimientos.filter(recepcionista_id=recepcionista_id)
        if producto_id:
            movimientos = movimientos.filter(detalles__producto_id=producto_id).distinct()
        if numero_movimiento:
            movimientos = movimientos.filter(numero_movimiento__icontains=numero_movimiento)
        
        movimientos = movimientos.select_related(
            'cliente', 'almacen_origen', 'almacen_destino', 'proveedor', 'recepcionista'
        ).prefetch_related('detalles__producto__unidad_medida').order_by('-fecha')

    # Respuesta HTTP para CSV
    response = HttpResponse(content_type='text/csv')
    filename = f'movimientos_{tipo_reporte}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Encabezados
    if tipo_reporte == 'almacen':
        headers = ['N¬∞ Movimiento', 'Tipo', 'Fecha', 'Almac√©n Origen', 'Almac√©n Destino', 
                   'Proveedor', 'Recepcionista', 'C√≥digo Producto', 'Nombre Producto', 
                   'Cant. Buena', 'Cant. Da√±ada', 'Total', 'Unidad']
    else:
        headers = ['N¬∞ Movimiento', 'Tipo', 'Fecha', 'Cliente', 'Almac√©n Origen', 
                   'Almac√©n Destino', 'Proveedor', 'Recepcionista', 'C√≥digo Producto', 
                   'Nombre Producto', 'Cant. Buena', 'Cant. Da√±ada', 'Total', 'Unidad']
        
    writer.writerow(headers)
    
    # Datos
    for movimiento in movimientos:
        for detalle in movimiento.detalles.all():
            total = (detalle.cantidad or Decimal('0')) + (detalle.cantidad_danada or Decimal('0'))
            unidad = detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND'
            
            common_data = [
                movimiento.numero_movimiento,
                movimiento.get_tipo_display(),
                movimiento.fecha.strftime('%d/%m/%Y'),
                str(movimiento.almacen_origen) if movimiento.almacen_origen else '-',
                str(movimiento.almacen_destino) if movimiento.almacen_destino else '-',
                str(movimiento.proveedor) if movimiento.proveedor else '-',
                str(movimiento.recepcionista) if movimiento.recepcionista else '-',
                detalle.producto.codigo,
                detalle.producto.nombre,
                detalle.cantidad or 0,
                detalle.cantidad_danada or 0,
                total,
                unidad
            ]
            
            if tipo_reporte == 'almacen':
                row_data = common_data
            else:
                cliente_nombre = str(movimiento.cliente) if movimiento.cliente else '-'
                row_data = [
                    movimiento.numero_movimiento,
                    movimiento.get_tipo_display(),
                    movimiento.fecha.strftime('%d/%m/%Y'),
                    cliente_nombre,
                    str(movimiento.almacen_origen) if movimiento.almacen_origen else '-',
                    str(movimiento.almacen_destino) if movimiento.almacen_destino else '-',
                    str(movimiento.proveedor) if movimiento.proveedor else '-',
                    str(movimiento.recepcionista) if movimiento.recepcionista else '-',
                    detalle.producto.codigo,
                    detalle.producto.nombre,
                    detalle.cantidad or 0,
                    detalle.cantidad_danada or 0,
                    total,
                    unidad
                ]
                
            writer.writerow(row_data)
            
    return response

# ==========================================
# NUEVAS FUNCIONES DE EXPORTACI√ìN DE STOCK
# ==========================================

@staff_member_required
def exportar_stock_excel(request):
    """Exportaci√≥n OPTIMIZADA de stock est√°ndar."""
    vista = request.GET.get('vista', 'detallado')
    almacen_id = request.GET.get('almacen', '')
    categoria_id = request.GET.get('categoria', '')
    producto_id = request.GET.get('producto', '')
    stock_minimo = request.GET.get('stock_minimo', '')
    solo_con_stock = request.GET.get('solo_con_stock', '')

    almacenes = Almacen.objects.filter(activo=True)
    if almacen_id: almacenes = almacenes.filter(id=almacen_id)

    productos_qs = Producto.objects.filter(activo=True).select_related('categoria', 'unidad_medida')
    if categoria_id: productos_qs = productos_qs.filter(categoria_id=categoria_id)
    if producto_id: productos_qs = productos_qs.filter(id=producto_id)
    
    # Mapeo para acceso r√°pido
    productos_map = {p.id: p for p in productos_qs}
    target_pids = set(productos_map.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Stock"
    
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    if vista == "detallado":
        headers = ["Almac√©n", "C√≥digo", "Producto", "Categor√≠a", "Unidad", "Entradas", "Salidas", "Traslados Rec", "Traslados Env", "Stock Bueno", "Stock Da√±ado", "Stock Total"]
        ws.append(headers)
        
        for alm in almacenes:
            # C√°lculo masivo por almac√©n
            bulk = get_stock_bulk(alm.id)
            
            for pid, data in bulk.items():
                if pid not in target_pids: continue
                if solo_con_stock and data['stock_total'] == 0: continue
                
                prod = productos_map[pid]
                if stock_minimo and prod.stock_minimo and data['stock_bueno'] > prod.stock_minimo: continue
                
                d = data['data']
                ws.append([
                    alm.nombre, prod.codigo, prod.nombre,
                    prod.categoria.nombre if prod.categoria else '-',
                    prod.unidad_medida.abreviatura if prod.unidad_medida else 'UND',
                    d['ent_b'] + d['ent_d'],
                    d['sal_b'] + d['sal_d'],
                    d['tras_rec_b'] + d['tras_rec_d'],
                    d['tras_env_b'] + d['tras_env_d'],
                    data['stock_bueno'], data['stock_danado'], data['stock_total']
                ])

    elif vista == "por_almacen":
        headers = ["Almac√©n", "Total Productos", "Stock Bueno Total", "Stock Da√±ado Total", "Stock Total"]
        ws.append(headers)
        
        for alm in almacenes:
            bulk = get_stock_bulk(alm.id)
            # Filtrar solo productos seleccionados
            filtered_data = [v for k, v in bulk.items() if k in target_pids]
            
            # Filtro adicional si se requiere (solo con stock)
            if solo_con_stock:
                filtered_data = [v for v in filtered_data if v['stock_total'] != 0]
                
            if not filtered_data and solo_con_stock: continue
            
            ws.append([
                alm.nombre,
                len(filtered_data),
                sum(v['stock_bueno'] for v in filtered_data),
                sum(v['stock_danado'] for v in filtered_data),
                sum(v['stock_total'] for v in filtered_data)
            ])

    else: # por_producto
        headers = ["C√≥digo", "Producto", "Categor√≠a", "Unidad", "Almacenes con Stock", "Stock Bueno Total", "Stock Da√±ado Total", "Stock Total"]
        ws.append(headers)
        
        # Para vista por producto, necesitamos iterar productos y sumar almacenes
        # Estrategia: Calcular todo en memoria primero
        global_map = {pid: {'bueno': 0, 'danado': 0, 'total': 0, 'alms': 0} for pid in target_pids}
        
        for alm in Almacen.objects.filter(activo=True): # Todos los almacenes para sumar
            bulk = get_stock_bulk(alm.id)
            for pid, data in bulk.items():
                if pid in global_map:
                    if data['stock_total'] != 0:
                        global_map[pid]['bueno'] += data['stock_bueno']
                        global_map[pid]['danado'] += data['stock_danado']
                        global_map[pid]['total'] += data['stock_total']
                        global_map[pid]['alms'] += 1
        
        for pid, vals in global_map.items():
            if solo_con_stock and vals['total'] == 0: continue
            prod = productos_map[pid]
            
            ws.append([
                prod.codigo, prod.nombre,
                prod.categoria.nombre if prod.categoria else '-',
                prod.unidad_medida.abreviatura if prod.unidad_medida else 'UND',
                vals['alms'], vals['bueno'], vals['danado'], vals['total']
            ])

    # Estilos
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=stock_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


@staff_member_required
def exportar_stock_csv(request):
    """Exporta el reporte de stock filtrado a CSV."""
    
    vista = request.GET.get('vista', 'detallado')
    almacen_id = request.GET.get('almacen', '')
    categoria_id = request.GET.get('categoria', '')
    producto_id = request.GET.get('producto', '')
    solo_con_stock = request.GET.get('solo_con_stock', '') == 'on'
    
    stocks = []
    
    # Obtener almacenes
    almacenes = Almacen.objects.filter(activo=True)
    if almacen_id:
        almacenes = almacenes.filter(id=almacen_id)
    
    # Obtener productos
    productos = Producto.objects.filter(activo=True)
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    if producto_id:
        productos = productos.filter(id=producto_id)
    
    if vista == 'detallado':
        for almacen in almacenes:
            for producto in productos:
                # Calcular ENTRADAS
                entradas = DetalleMovimientoAlmacen.objects.filter(
                    producto=producto
                ).filter(
                    Q(movimiento__tipo='ENTRADA', movimiento__almacen_destino=almacen) |
                    Q(movimiento__tipo='TRASLADO', movimiento__almacen_destino=almacen)
                ).aggregate(
                    buenas=Sum('cantidad'),
                    danadas=Sum('cantidad_danada')
                )
                
                # Calcular SALIDAS
                salidas = DetalleMovimientoAlmacen.objects.filter(
                    producto=producto
                ).filter(
                    Q(movimiento__tipo='SALIDA', movimiento__almacen_origen=almacen) |
                    Q(movimiento__tipo='TRASLADO', movimiento__almacen_origen=almacen)
                ).aggregate(
                    buenas=Sum('cantidad'),
                    danadas=Sum('cantidad_danada')
                )
                
                stock_bueno = Decimal(str(entradas['buenas'] or 0)) - Decimal(str(salidas['buenas'] or 0))
                stock_danado = Decimal(str(entradas['danadas'] or 0)) - Decimal(str(salidas['danadas'] or 0))
                stock_total = stock_bueno + stock_danado
                
                if solo_con_stock and stock_total == 0:
                    continue
                
                entradas_total = Decimal(str(entradas['buenas'] or 0)) + Decimal(str(entradas['danadas'] or 0))
                salidas_total = Decimal(str(salidas['buenas'] or 0)) + Decimal(str(salidas['danadas'] or 0))
                
                # Traslados netos
                traslados_recibidos = DetalleMovimientoAlmacen.objects.filter(
                    movimiento__tipo='TRASLADO',
                    movimiento__almacen_destino=almacen,
                    producto=producto
                ).aggregate(total=Sum('cantidad'), danada=Sum('cantidad_danada'))
                
                traslados_enviados = DetalleMovimientoAlmacen.objects.filter(
                    movimiento__tipo='TRASLADO',
                    movimiento__almacen_origen=almacen,
                    producto=producto
                ).aggregate(total=Sum('cantidad'), danada=Sum('cantidad_danada'))
                
                traslados_netos = (
                    Decimal(str(traslados_recibidos['total'] or 0)) + 
                    Decimal(str(traslados_recibidos['danada'] or 0)) -
                    Decimal(str(traslados_enviados['total'] or 0)) -
                    Decimal(str(traslados_enviados['danada'] or 0))
                )
                
                stocks.append({
                    'almacen': almacen,
                    'producto': producto,
                    'stock_bueno': float(stock_bueno),
                    'stock_danado': float(stock_danado),
                    'stock_total': float(stock_total),
                    'entradas_total': float(entradas_total),
                    'salidas_total': float(salidas_total),
                    'traslados_netos': float(traslados_netos)
                })
    
    elif vista == 'por_almacen':
        for almacen in almacenes:
            productos_ids = set(
                list(DetalleMovimientoAlmacen.objects.filter(
                    Q(movimiento__almacen_origen=almacen) | Q(movimiento__almacen_destino=almacen)
                ).values_list('producto_id', flat=True).distinct())
            )
            
            total_productos = 0
            stock_buena_total = Decimal('0')
            stock_danada_total = Decimal('0')
            
            for producto_id_item in productos_ids:
                try:
                    producto = Producto.objects.get(id=producto_id_item)
                    
                    entradas = DetalleMovimientoAlmacen.objects.filter(
                        producto=producto
                    ).filter(
                        Q(movimiento__tipo='ENTRADA', movimiento__almacen_destino=almacen) |
                        Q(movimiento__tipo='TRASLADO', movimiento__almacen_destino=almacen)
                    ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
                    
                    salidas = DetalleMovimientoAlmacen.objects.filter(
                        producto=producto
                    ).filter(
                        Q(movimiento__tipo='SALIDA', movimiento__almacen_origen=almacen) |
                        Q(movimiento__tipo='TRASLADO', movimiento__almacen_origen=almacen)
                    ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
                    
                    stock_bueno = Decimal(str(entradas['buenas'] or 0)) - Decimal(str(salidas['buenas'] or 0))
                    stock_danado = Decimal(str(entradas['danadas'] or 0)) - Decimal(str(salidas['danadas'] or 0))
                    stock_total_prod = stock_bueno + stock_danado
                    
                    if stock_total_prod != 0:
                        total_productos += 1
                        stock_buena_total += stock_bueno
                        stock_danada_total += stock_danado
                        
                except Producto.DoesNotExist:
                    continue
            
            stock_total = stock_buena_total + stock_danada_total
            
            if solo_con_stock and stock_total == 0:
                continue
            
            if total_productos > 0 or stock_total != 0:
                stocks.append({
                    'almacen': almacen,
                    'total_productos': total_productos,
                    'stock_buena_total': float(stock_buena_total),
                    'stock_danada_total': float(stock_danada_total),
                    'stock_total': float(stock_total)
                })
    
    else:  # por_producto
        for producto in productos:
            stock_buena_total = Decimal('0')
            stock_danada_total = Decimal('0')
            total_almacenes = 0
            
            for almacen in Almacen.objects.filter(activo=True):
                entradas = DetalleMovimientoAlmacen.objects.filter(
                    producto=producto
                ).filter(
                    Q(movimiento__tipo='ENTRADA', movimiento__almacen_destino=almacen) |
                    Q(movimiento__tipo='TRASLADO', movimiento__almacen_destino=almacen)
                ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
                
                salidas = DetalleMovimientoAlmacen.objects.filter(
                    producto=producto
                ).filter(
                    Q(movimiento__tipo='SALIDA', movimiento__almacen_origen=almacen) |
                    Q(movimiento__tipo='TRASLADO', movimiento__almacen_origen=almacen)
                ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
                
                stock_bueno = Decimal(str(entradas['buenas'] or 0)) - Decimal(str(salidas['buenas'] or 0))
                stock_danado = Decimal(str(entradas['danadas'] or 0)) - Decimal(str(salidas['danadas'] or 0))
                stock_total_alm = stock_bueno + stock_danado
                
                if stock_total_alm != 0:
                    stock_buena_total += stock_bueno
                    stock_danada_total += stock_danado
                    total_almacenes += 1
            
            stock_total_producto = stock_buena_total + stock_danada_total
            
            if solo_con_stock and stock_total_producto == 0:
                continue
            
            if total_almacenes > 0 or stock_total_producto != 0:
                stocks.append({
                    'producto': producto,
                    'total_almacenes': total_almacenes,
                    'stock_buena_total': float(stock_buena_total),
                    'stock_danada_total': float(stock_danada_total),
                    'stock_total': float(stock_total_producto)
                })
    
    # Respuesta HTTP
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f'stock_{vista}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # BOM para Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # Encabezados seg√∫n vista
    if vista == 'detallado':
        headers = ['Almac√©n', 'C√≥digo', 'Producto', 'Categor√≠a', 'Unidad', 
                   'Entradas', 'Salidas', 'Traslados', 'Stock Bueno', 'Stock Da√±ado', 'Stock Total']
    elif vista == 'por_almacen':
        headers = ['Almac√©n', 'Total Productos', 'Stock Bueno', 'Stock Da√±ado', 'Stock Total']
    else:  # por_producto
        headers = ['C√≥digo', 'Producto', 'Categor√≠a', 'Unidad', 'Almacenes', 
                   'Stock Bueno', 'Stock Da√±ado', 'Stock Total']
    
    writer.writerow(headers)
    
    # Datos
    for stock in stocks:
        if vista == 'detallado':
            row_data = [
                stock['almacen'].nombre,
                stock['producto'].codigo,
                stock['producto'].nombre,
                stock['producto'].categoria.nombre if stock['producto'].categoria else '-',
                stock['producto'].unidad_medida.abreviatura if stock['producto'].unidad_medida else 'UND',
                stock['entradas_total'],
                stock['salidas_total'],
                stock['traslados_netos'],
                stock['stock_bueno'],
                stock['stock_danado'],
                stock['stock_total']
            ]
        elif vista == 'por_almacen':
            row_data = [
                stock['almacen'].nombre,
                stock['total_productos'],
                stock['stock_buena_total'],
                stock['stock_danada_total'],
                stock['stock_total']
            ]
        else:  # por_producto
            row_data = [
                stock['producto'].codigo,
                stock['producto'].nombre,
                stock['producto'].categoria.nombre if stock['producto'].categoria else '-',
                stock['producto'].unidad_medida.abreviatura if stock['producto'].unidad_medida else 'UND',
                stock['total_almacenes'],
                stock['stock_buena_total'],
                stock['stock_danada_total'],
                stock['stock_total']
            ]
        
        writer.writerow(row_data)
    
    return response

# ========================================
# NUEVA FUNCI√ìN - AGREGAR AL FINAL
# ========================================
@staff_member_required
def obtener_detalle_stock(request):
    """ 
    Vista para obtener el detalle completo de stock de un producto en un almac√©n espec√≠fico
    CORREGIDO: Ahora devuelve los totales de entradas, salidas y traslados para evitar NaN
    """
    producto_id = request.GET.get('producto_id')
    almacen_id = request.GET.get('almacen_id')
    
    if not producto_id or not almacen_id:
        return JsonResponse({
            'success': False, 
            'error': 'Faltan par√°metros requeridos'
        })
        
    try:
        producto = Producto.objects.get(id=producto_id)
        almacen = Almacen.objects.get(id=almacen_id)

        # 1. Calcular Totales para el Resumen (Para evitar NaN en el frontend)
        # ------------------------------------------------------------------
        
        # Total Entradas
        entradas_agg = DetalleMovimientoAlmacen.objects.filter(
            movimiento__tipo='ENTRADA',
            movimiento__almacen_destino=almacen,
            producto=producto
        ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
        total_entradas = (entradas_agg['buenas'] or Decimal('0')) + (entradas_agg['danadas'] or Decimal('0'))

        # Total Salidas
        salidas_agg = DetalleMovimientoAlmacen.objects.filter(
            movimiento__tipo='SALIDA',
            movimiento__almacen_origen=almacen,
            producto=producto
        ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
        total_salidas = (salidas_agg['buenas'] or Decimal('0')) + (salidas_agg['danadas'] or Decimal('0'))

        # Traslados Recibidos
        tras_rec_agg = DetalleMovimientoAlmacen.objects.filter(
            movimiento__tipo='TRASLADO',
            movimiento__almacen_destino=almacen,
            producto=producto
        ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
        total_traslados_recibidos = (tras_rec_agg['buenas'] or Decimal('0')) + (tras_rec_agg['danadas'] or Decimal('0'))

        # Traslados Enviados
        tras_env_agg = DetalleMovimientoAlmacen.objects.filter(
            movimiento__tipo='TRASLADO',
            movimiento__almacen_origen=almacen,
            producto=producto
        ).aggregate(buenas=Sum('cantidad'), danadas=Sum('cantidad_danada'))
        total_traslados_enviados = (tras_env_agg['buenas'] or Decimal('0')) + (tras_env_agg['danadas'] or Decimal('0'))

        # 2. Calcular Stock Actual (Balance)
        # ------------------------------------------------------------------
        stock_bueno = (
            (entradas_agg['buenas'] or Decimal('0')) + 
            (tras_rec_agg['buenas'] or Decimal('0')) - 
            (salidas_agg['buenas'] or Decimal('0')) - 
            (tras_env_agg['buenas'] or Decimal('0'))
        )
        
        stock_danado = (
            (entradas_agg['danadas'] or Decimal('0')) + 
            (tras_rec_agg['danadas'] or Decimal('0')) - 
            (salidas_agg['danadas'] or Decimal('0')) - 
            (tras_env_agg['danadas'] or Decimal('0'))
        )
        
        stock_total = stock_bueno + stock_danado
        
        # 3. Obtener movimientos detallados
        # ------------------------------------------------------------------
        movimientos_qs = MovimientoAlmacen.objects.filter(
            Q(almacen_origen=almacen) | Q(almacen_destino=almacen)
        ).filter(
            detalles__producto=producto
        ).select_related(
            'proveedor', 'recepcionista', 'almacen_origen', 'almacen_destino'
        ).prefetch_related(
            'detalles'
        ).order_by('-fecha', '-numero_movimiento')
        
        movimientos_list = []
        for mov in movimientos_qs:
            detalle = mov.detalles.filter(producto=producto).first()
            if not detalle:
                continue

            cantidad_buena = detalle.cantidad or Decimal('0')
            cantidad_danada = detalle.cantidad_danada or Decimal('0')
            cantidad_total = cantidad_buena + cantidad_danada
            
            # Ajuste de signos visuales seg√∫n tipo
            if mov.tipo == 'SALIDA' and mov.almacen_origen == almacen:
                cantidad_total = -cantidad_total # Salida resta
            elif mov.tipo == 'TRASLADO' and mov.almacen_origen == almacen:
                cantidad_total = -cantidad_total # Traslado enviado resta
            
            estado = '-'
            if cantidad_buena > 0 and cantidad_danada == 0: estado = 'BUENO'
            elif cantidad_buena == 0 and cantidad_danada > 0: estado = 'DA√ëADO'
            elif cantidad_buena > 0 and cantidad_danada > 0: estado = 'MIXTO'

            movimientos_list.append({
                'numero_movimiento': mov.numero_movimiento,
                'tipo': mov.get_tipo_display(),
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M'),
                'cantidad': str(cantidad_total),
                'cantidad_buena': str(cantidad_buena),
                'cantidad_danada': str(cantidad_danada),
                'estado': estado,
                'proveedor': mov.proveedor.nombre if mov.proveedor else None,
                'recepcionista': str(mov.recepcionista) if mov.recepcionista else None,
                'almacen_origen': mov.almacen_origen.nombre if mov.almacen_origen else None,
                'almacen_destino': mov.almacen_destino.nombre if mov.almacen_destino else None,
            })
            
        return JsonResponse({
            'success': True,
            'producto': {
                'nombre': producto.nombre,
                'codigo': producto.codigo,
                'categoria': producto.categoria.nombre if producto.categoria else None,
                'unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND'
            },
            'almacen': {
                'nombre': almacen.nombre
            },
            'resumen': {
                # Aqu√≠ enviamos los datos que faltaban y causaban el NaN
                'total_entradas': str(total_entradas),
                'total_salidas': str(total_salidas),
                'traslados_recibidos': str(total_traslados_recibidos),
                'traslados_enviados': str(total_traslados_enviados),
                # Datos de stock
                'stock_bueno': str(stock_bueno),
                'stock_danado': str(stock_danado),
                'stock_total': str(stock_total)
            },
            'movimientos': movimientos_list
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'})
    except Almacen.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Almac√©n no encontrado'})
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}', 'traceback': traceback.format_exc()})

# ==========================================
# ‚≠ê NUEVA FUNCI√ìN - AGREGAR AL FINAL DEL ARCHIVO
# ==========================================

@staff_member_required
def obtener_detalle_estadistica(request):
    """
    Vista AJAX optimizada para estad√≠sticas del reporte de stock est√°ndar.
    """
    tipo = request.GET.get('tipo')
    
    try:
        # Calcular mapa global de stock en memoria
        almacenes = Almacen.objects.filter(activo=True)
        global_stock = {} 
        
        # 1. Construir mapa global (Iterar almacenes es r√°pido, son pocos)
        for alm in almacenes:
            alm_stocks = get_stock_bulk(alm.id)
            for pid, data in alm_stocks.items():
                if pid not in global_stock:
                    global_stock[pid] = {'bueno': Decimal(0), 'danado': Decimal(0), 'total': Decimal(0)}
                global_stock[pid]['bueno'] += data['stock_bueno']
                global_stock[pid]['danado'] += data['stock_danado']
                global_stock[pid]['total'] += data['stock_total']

        total_productos_sistema = Producto.objects.filter(activo=True).count()
        
        if tipo == 'total_productos':
            con_stock = 0
            con_stock_neg = 0
            
            for pid, vals in global_stock.items():
                if vals['total'] > 0: con_stock += 1
                elif vals['total'] < 0: con_stock_neg += 1
            
            sin_stock = total_productos_sistema - con_stock - con_stock_neg
            
            # Categor√≠as
            pids_mov = list(global_stock.keys())
            cats = Producto.objects.filter(id__in=pids_mov).values('categoria__nombre').annotate(total=Count('id')).order_by('-total')
            categorias_list = [{'categoria': c['categoria__nombre'] or 'Sin Categor√≠a', 'total': c['total']} for c in cats]
            
            return JsonResponse({
                'success': True,
                'total_productos': total_productos_sistema,
                'productos_con_stock': con_stock,
                'productos_sin_stock': sin_stock,
                'productos_con_stock_negativo': con_stock_neg,
                'por_categoria': categorias_list
            })
            
        elif tipo == 'stock_bueno':
            total_bueno = sum(i['bueno'] for i in global_stock.values())
            prod_pos = sum(1 for i in global_stock.values() if i['bueno'] > 0)
            prod_neg = sum(1 for i in global_stock.values() if i['bueno'] < 0)
            
            por_almacen = []
            for alm in almacenes:
                alm_stocks = get_stock_bulk(alm.id)
                suma = sum(v['stock_bueno'] for v in alm_stocks.values())
                por_almacen.append({
                    'almacen': alm.nombre, 'stock_bueno': float(suma),
                    'es_negativo': suma < 0, 'es_cero': suma == 0
                })
            por_almacen.sort(key=lambda x: -abs(x['stock_bueno']))
            
            return JsonResponse({
                'success': True,
                'total_stock_bueno': float(total_bueno),
                'productos_con_stock_bueno': prod_pos,
                'productos_con_stock_bueno_negativo': prod_neg,
                'por_almacen': por_almacen
            })

        elif tipo == 'stock_danado':
            total_danado = sum(i['danado'] for i in global_stock.values())
            prod_pos = sum(1 for i in global_stock.values() if i['danado'] > 0)
            prod_neg = sum(1 for i in global_stock.values() if i['danado'] < 0)
            
            por_almacen = []
            for alm in almacenes:
                alm_stocks = get_stock_bulk(alm.id)
                suma = sum(v['stock_danado'] for v in alm_stocks.values())
                por_almacen.append({
                    'almacen': alm.nombre, 'stock_danado': float(suma),
                    'es_negativo': suma < 0, 'es_cero': suma == 0
                })
            por_almacen.sort(key=lambda x: -abs(x['stock_danado']))
            
            # Top da√±ados
            top_danados = []
            if total_danado != 0:
                pids = [pid for pid, v in global_stock.items() if v['danado'] != 0]
                p_info = Producto.objects.filter(id__in=pids).in_bulk()
                for pid in pids:
                    if pid in p_info:
                        top_danados.append({
                            'producto': p_info[pid].nombre,
                            'codigo': p_info[pid].codigo,
                            'stock_danado': float(global_stock[pid]['danado']),
                            'es_negativo': global_stock[pid]['danado'] < 0
                        })
                top_danados.sort(key=lambda x: -abs(x['stock_danado']))
            
            return JsonResponse({
                'success': True,
                'total_stock_danado': float(total_danado),
                'productos_con_stock_danado': prod_pos,
                'productos_con_stock_danado_negativo': prod_neg,
                'por_almacen': por_almacen,
                'productos_mas_danados': top_danados[:10]
            })
            
        elif tipo == 'total_almacenes':
            almacenes_list = []
            for alm in almacenes:
                alm_stocks = get_stock_bulk(alm.id)
                activos = [v for v in alm_stocks.values() if v['stock_total'] != 0]
                
                almacenes_list.append({
                    'nombre': alm.nombre,
                    'total_productos': len(activos),
                    'stock_bueno': float(sum(v['stock_bueno'] for v in activos)),
                    'stock_danado': float(sum(v['stock_danado'] for v in activos)),
                    'stock_total': float(sum(v['stock_total'] for v in activos)),
                })
            
            return JsonResponse({'success': True, 'total_almacenes': len(almacenes), 'almacenes': almacenes_list})
            
        elif tipo == 'bajo_minimo':
            # Productos bajo m√≠nimo (comparando stock bueno global vs stock minimo producto)
            # Nota: Si el stock minimo es por almac√©n, habr√≠a que ajustar la l√≥gica.
            # Asumimos stock minimo global del producto.
            productos_criticos = []
            
            # Traer productos con stock minimo definido
            prods_min = Producto.objects.filter(activo=True, stock_minimo__gt=0)
            
            for prod in prods_min:
                stock_actual = global_stock.get(prod.id, {}).get('bueno', Decimal(0))
                if stock_actual < prod.stock_minimo:
                    productos_criticos.append({
                        'producto': prod.nombre,
                        'codigo': prod.codigo,
                        'stock_actual': float(stock_actual),
                        'stock_minimo': float(prod.stock_minimo),
                        'diferencia': float(prod.stock_minimo - stock_actual)
                    })
            
            productos_criticos.sort(key=lambda x: x['diferencia'], reverse=True)
            
            return JsonResponse({
                'success': True, 
                'total_bajo_minimo': len(productos_criticos),
                'productos': productos_criticos[:20]
            })
            
        elif tipo == 'valor_inventario':
            total_items = sum(v['total'] for v in global_stock.values() if v['total'] > 0)
            
            # Por categor√≠a
            cat_totals = {}
            p_cats = Producto.objects.filter(id__in=global_stock.keys()).values('id', 'categoria__nombre')
            p_cat_map = {p['id']: (p['categoria__nombre'] or 'Sin Categor√≠a') for p in p_cats}
            
            for pid, vals in global_stock.items():
                if vals['total'] > 0 and pid in p_cat_map:
                    cat = p_cat_map[pid]
                    cat_totals[cat] = cat_totals.get(cat, Decimal(0)) + vals['total']
            
            cat_list = [{'categoria': k, 'total_items': float(v)} for k, v in cat_totals.items()]
            cat_list.sort(key=lambda x: x['total_items'], reverse=True)
            
            return JsonResponse({
                'success': True,
                'total_productos': total_productos_sistema,
                'total_items': float(total_items),
                'por_categoria': cat_list
            })

        return JsonResponse({'success': False, 'error': 'Tipo desconocido'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'traceback': traceback.format_exc()})

@staff_member_required
def obtener_detalle_entrega_cliente(request):
    """
    Vista para obtener el detalle completo de entregas de un producto a un cliente espec√≠fico
    CORREGIDO: Entrada (+), Salida (-), Traslado (seg√∫n direcci√≥n).
    Se elimin√≥ la resta por defecto en traslados ambiguos.
    """
    cliente_id = request.GET.get('cliente_id')
    producto_id = request.GET.get('producto_id')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    if not cliente_id or not producto_id:
        return JsonResponse({'success': False, 'error': 'Faltan par√°metros requeridos'})
    
    try:
        from beneficiarios.models import Cliente, DetalleMovimientoCliente, MovimientoCliente
        from productos.models import Producto
        
        cliente = Cliente.objects.get(id=cliente_id)
        producto = Producto.objects.get(id=producto_id)
        
        # Convertir fechas
        fecha_inicio_obj = None
        fecha_fin_obj = None
        if fecha_inicio:
            try: fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            except ValueError: pass
        if fecha_fin:
            try: fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError: pass
        
        # Filtrar movimientos: Buscamos movimientos donde el cliente sea el actor principal, origen o destino
        movimientos_qs = MovimientoCliente.objects.filter(
            Q(cliente=cliente) | Q(cliente_origen=cliente) | Q(cliente_destino=cliente)
        ).select_related(
            'proveedor', 'recepcionista', 'almacen_origen', 'almacen_destino'
        ).prefetch_related('detalles')
        
        if fecha_inicio_obj: movimientos_qs = movimientos_qs.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj: movimientos_qs = movimientos_qs.filter(fecha__lte=fecha_fin_obj)
        
        # Filtrar solo movimientos con este producto
        movimientos_qs = movimientos_qs.filter(detalles__producto=producto).distinct()
        
        total_entregas = movimientos_qs.count()
        
        # Calcular Stock Neto (Balance)
        stock_bueno = Decimal('0')
        stock_danado = Decimal('0')
        
        cantidad_entrada = Decimal('0')
        cantidad_salida = Decimal('0')
        cantidad_traslado = Decimal('0')
        
        movimientos_list = []
        
        # Iterar sobre movimientos ordenados por fecha descendente para obtener un balance acumulativo
        for mov in movimientos_qs.order_by('fecha', 'id'): # Cambi√© a ascendente para calcular stock correctamente
            detalle = mov.detalles.filter(producto=producto).first()
            if not detalle: continue
            
            cant_b = detalle.cantidad or Decimal('0')
            cant_d = detalle.cantidad_danada or Decimal('0')
            cant_total = cant_b + cant_d
            
            # --- L√ìGICA DE C√ÅLCULO DE STOCK ---
            signo = 0
            
            if mov.tipo == 'ENTRADA':
                # Entrada SUMA
                stock_bueno += cant_b
                stock_danado += cant_d
                cantidad_entrada += cant_total
                signo = 1
                
            elif mov.tipo == 'SALIDA':
                # Salida RESTA
                stock_bueno -= cant_b
                stock_danado -= cant_d
                cantidad_salida += cant_total
                signo = -1
                
            elif mov.tipo == 'TRASLADO':
                cantidad_traslado += cant_total
                
                # Verificar direcci√≥n para este cliente
                cliente_id_int = int(cliente_id)
                es_origen = mov.cliente_origen_id == cliente_id_int
                es_destino = mov.cliente_destino_id == cliente_id_int
                
                if es_origen:
                    # Sale del cliente -> Resta
                    stock_bueno -= cant_b
                    stock_danado -= cant_d
                    signo = -1
                elif es_destino:
                    # Entra al cliente -> Suma
                    stock_bueno += cant_b
                    stock_danado += cant_d
                    signo = 1
                # ELSE: Si no es ni origen ni destino (aunque est√© en el filtro Q), es neutro para el stock de este cliente.
                # No se aplica ninguna operaci√≥n de suma/resta al stock.
            
            # Preparar datos para la lista (se muestran con stock acumulado)
            
            estado = '-'
            if cant_b > 0 and cant_d > 0: estado = 'MIXTO'
            elif cant_b > 0: estado = 'BUENO'
            elif cant_d > 0: estado = 'DA√ëADO'
            
            # Formatear la cantidad visualmente con el signo correcto
            cant_visual = cant_total * signo
            
            # ‚úÖ Preparar informaci√≥n de clientes para traslados
            cliente_origen_nombre = None
            cliente_destino_nombre = None
            
            if mov.tipo == 'TRASLADO':
                cliente_origen_nombre = mov.cliente_origen.nombre if mov.cliente_origen else None
                cliente_destino_nombre = mov.cliente_destino.nombre if mov.cliente_destino else None
            
            movimientos_list.append({
                'numero_movimiento': mov.numero_movimiento,
                'tipo': mov.get_tipo_display(),
                'fecha_ordenamiento': mov.fecha,
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M'),
                'cantidad': str(cant_visual), # Cantidad con signo visual
                'cantidad_buena': str(cant_b),
                'cantidad_danada': str(cant_d),
                'stock_bueno_actual': str(stock_bueno), # Stock acumulado al momento del movimiento
                'stock_danado_actual': str(stock_danado), # Stock acumulado al momento del movimiento
                'stock_total_actual': str(stock_bueno + stock_danado), # Stock total acumulado
                'estado': estado,
                'proveedor': mov.proveedor.nombre if mov.proveedor else None,
                'recepcionista': str(mov.recepcionista) if mov.recepcionista else None,
                'almacen': mov.almacen_origen.nombre if mov.almacen_origen else (mov.almacen_destino.nombre if mov.almacen_destino else None),
                'cliente_origen': cliente_origen_nombre,  # ‚úÖ NUEVO
                'cliente_destino': cliente_destino_nombre,  # ‚úÖ NUEVO
                'observaciones': mov.observaciones if hasattr(mov, 'observaciones') else None,
            })
        
        stock_total = stock_bueno + stock_danado
        
        # Invertir la lista para que los movimientos m√°s recientes se muestren primero en el modal
        movimientos_list.reverse()
        
        return JsonResponse({
            'success': True,
            'cliente': {
                'nombre': cliente.nombre,
                'codigo': cliente.codigo,
                'direccion': cliente.direccion or '-',
                'telefono': cliente.telefono or '-'
            },
            'producto': {
                'nombre': producto.nombre,
                'codigo': producto.codigo,
                'categoria': producto.categoria.nombre if producto.categoria else '-',
                'unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND'
            },
            'resumen': {
                'total_entregas': total_entregas,
                'cantidad_total': str(stock_total), # Stock Neto final
                'cantidad_buena': str(stock_bueno), # Stock Bueno Neto final
                'cantidad_danada': str(stock_danado), # Stock Da√±ado Neto final
                'cantidad_entrada': str(cantidad_entrada), # Volumen de entradas
                'cantidad_salida': str(cantidad_salida), # Volumen de salidas
                'cantidad_traslado': str(cantidad_traslado), # Volumen de traslados
                'stock_bueno': str(stock_bueno),
                'stock_danado': str(stock_danado),
                'stock_total': str(stock_total)
            },
            'movimientos': movimientos_list
        })
        
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cliente no encontrado'})
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'})
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}', 'traceback': traceback.format_exc()})

@staff_member_required
def obtener_productos_cliente(request):
    """
    Obtiene el resumen de productos (saldo) de un cliente espec√≠fico 
    basado en sus movimientos de ENTRADA, SALIDA y TRASLADO.
    """
    cliente_id = request.GET.get('cliente_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if not cliente_id:
        return JsonResponse({'error': 'Falta el par√°metro cliente_id'}, status=400)

    try:
        # 1. Obtener el cliente y preparar el ID
        cliente = Cliente.objects.get(id=cliente_id)
        cli_id_param = int(cliente_id)

        # 2. Queryset principal: Filtra todos los detalles de movimientos 
        #    donde el cliente est√© involucrado (como cliente principal, origen o destino).
        detalles_qs = DetalleMovimientoCliente.objects.filter(
            # 1. Movimientos ENTRADA y SALIDA (usando el campo 'cliente' principal)
            Q(movimiento__cliente_id=cli_id_param) | 
            
            # 2. Movimientos TRASLADO donde el cliente es ORIGEN (Salida/Resta)
            Q(movimiento__cliente_origen_id=cli_id_param, movimiento__tipo='TRASLADO') |
            
            # 3. Movimientos TRASLADO donde el cliente es DESTINO (Entrada/Suma)
            Q(movimiento__cliente_destino_id=cli_id_param, movimiento__tipo='TRASLADO')
            
        ).select_related(
            'producto__categoria', 
            'producto__unidad_medida', 
            'movimiento'
        ).order_by('producto__codigo', 'movimiento__fecha')

        # 3. Aplicar filtros de fecha si existen
        filtros_movimiento = Q()
        if fecha_inicio:
            # Se usa .date() para comparar solo la parte de la fecha
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            filtros_movimiento &= Q(movimiento__fecha__gte=fecha_inicio_obj)
        if fecha_fin:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            # Asegura que se incluya todo el d√≠a final (hasta las 23:59:59 si se usa un campo datetime)
            filtros_movimiento &= Q(movimiento__fecha__lte=fecha_fin_obj)

        if filtros_movimiento:
             detalles_qs = detalles_qs.filter(filtros_movimiento)
        
        # 4. Agrupar por producto y calcular el saldo (con la l√≥gica de suma/resta)
        productos_dict = {}
        
        for detalle in detalles_qs:
            producto_id = detalle.producto.id
            mov = detalle.movimiento
            
            if producto_id not in productos_dict:
                productos_dict[producto_id] = {
                    'producto_id': producto_id,
                    'codigo': detalle.producto.codigo,
                    'nombre': detalle.producto.nombre,
                    'categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                    'unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                    'movimientos': set(),
                    'cantidad_buena': Decimal('0'),
                    'cantidad_danada': Decimal('0'),
                }
            
            productos_dict[producto_id]['movimientos'].add(mov.id)
            
            cant_b = detalle.cantidad or Decimal('0')
            cant_d = detalle.cantidad_danada or Decimal('0')
            
            # === L√ìGICA DE SALDO (Suma o Resta) ===
            
            # Caso 1: ENTRADA directa al cliente (SUMA)
            if mov.tipo == 'ENTRADA' and mov.cliente_id == cli_id_param:
                productos_dict[producto_id]['cantidad_buena'] += cant_b
                productos_dict[producto_id]['cantidad_danada'] += cant_d
                
            # Caso 2: SALIDA directa del cliente (RESTA)
            elif mov.tipo == 'SALIDA' and mov.cliente_id == cli_id_param:
                productos_dict[producto_id]['cantidad_buena'] -= cant_b
                productos_dict[producto_id]['cantidad_danada'] -= cant_d
                
            # Caso 3: TRASLADO
            elif mov.tipo == 'TRASLADO':
                es_origen = (mov.cliente_origen_id == cli_id_param)
                es_destino = (mov.cliente_destino_id == cli_id_param)
                
                if es_origen:
                    # Cliente env√≠a producto (RESTA)
                    productos_dict[producto_id]['cantidad_buena'] -= cant_b
                    productos_dict[producto_id]['cantidad_danada'] -= cant_d
                elif es_destino:
                    # Cliente recibe producto (SUMA)
                    productos_dict[producto_id]['cantidad_buena'] += cant_b
                    productos_dict[producto_id]['cantidad_danada'] += cant_d

        # 5. Convertir a lista y calcular totales
        productos_list = []
        total_buena = Decimal('0')
        total_danada = Decimal('0')

        for prod_id, data in productos_dict.items():
            data['total_movimientos'] = len(data['movimientos'])
            data['cantidad_total'] = data['cantidad_buena'] + data['cantidad_danada']
            
            # Solo incluir productos con stock distinto de cero o con movimientos en el rango
            if data['cantidad_total'] != Decimal('0') or len(data['movimientos']) > 0:
                productos_list.append({
                    'producto_id': data['producto_id'],
                    'codigo': data['codigo'],
                    'nombre': data['nombre'],
                    'categoria': data['categoria'],
                    'unidad': data['unidad'],
                    'total_movimientos': data['total_movimientos'],
                    # Convertir Decimal a float para JsonResponse (puede haber p√©rdida de precisi√≥n)
                    'cantidad_buena': float(data['cantidad_buena']), 
                    'cantidad_danada': float(data['cantidad_danada']),
                    'cantidad_total': float(data['cantidad_total']),
                })
                
                total_buena += data['cantidad_buena']
                total_danada += data['cantidad_danada']
        
        # Respuesta de √©xito
        return JsonResponse({
            'success': True,  # ‚úÖ CORRECCI√ìN: Cambiar 'status' por 'success'
            'cliente': {
                'id': cliente.id,
                'nombre': cliente.nombre,
                'codigo': cliente.codigo,
                'direccion': cliente.direccion or '-',
                'telefono': cliente.telefono or '-'
            },
            'total_productos': len(productos_list),
            'productos': productos_list,
            'totales': {
                'total_entregas': sum(p['total_movimientos'] for p in productos_list),
                'cantidad_buena': float(total_buena),
                'cantidad_danada': float(total_danada),
                'cantidad_total': float(total_buena + total_danada)
            }
        })
            
    # --- Manejo de errores ---
    except Cliente.DoesNotExist:
        # Error 404
        return JsonResponse({
            'success': False,
            'error': 'Cliente no encontrado'
        }, status=404)
        
    except Exception as e:
        # Error 500
        print("ERROR en obtener_productos_cliente:")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)

@staff_member_required
def obtener_clientes_producto(request):
    """
    Vista para obtener todos los clientes que recibieron un producto espec√≠fico
    """
    try:
        producto_id = request.GET.get('producto_id')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        
        if not producto_id:
            return JsonResponse({
                'success': False,
                'error': 'Falta el par√°metro producto_id'
            })
        
        from beneficiarios.models import Cliente, DetalleMovimientoCliente, MovimientoCliente
        from productos.models import Producto
        
        producto = Producto.objects.get(id=producto_id)
        
        # Convertir fechas
        fecha_inicio_obj = None
        fecha_fin_obj = None
        
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Filtrar movimientos
        movimientos_qs = MovimientoCliente.objects.all()
        
        if fecha_inicio_obj:
            movimientos_qs = movimientos_qs.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos_qs = movimientos_qs.filter(fecha__lte=fecha_fin_obj)
        
        # Obtener detalles de este producto
        detalles_qs = DetalleMovimientoCliente.objects.filter(
            movimiento__in=movimientos_qs,
            producto=producto
        ).select_related('movimiento', 'movimiento__cliente')
        
        # Agrupar por cliente
        clientes_dict = {}
        
        for detalle in detalles_qs:
            mov = detalle.movimiento
            # Determinar cu√°l es el cliente principal de esta fila para el reporte
            # Nota: En traslados, un solo movimiento involucra 2 clientes. 
            # Esta l√≥gica agrupa por el cliente "principal" del movimiento o eval√∫a origen/destino.
            
            # Para simplificar y asegurar que aparezcan AMBOS clientes en un traslado,
            # deber√≠amos procesar origen y destino por separado si es traslado.
            
            clientes_a_procesar = []
            
            cant_b = detalle.cantidad or Decimal('0')
            cant_d = detalle.cantidad_danada or Decimal('0')
            
            if mov.tipo == 'TRASLADO':
                if mov.cliente_origen:
                    clientes_a_procesar.append({
                        'cliente': mov.cliente_origen,
                        'signo': -1 # Resta
                    })
                if mov.cliente_destino:
                    clientes_a_procesar.append({
                        'cliente': mov.cliente_destino,
                        'signo': 1 # Suma
                    })
            elif mov.tipo == 'ENTRADA':
                if mov.cliente:
                    clientes_a_procesar.append({'cliente': mov.cliente, 'signo': 1}) # Suma
            elif mov.tipo == 'SALIDA':
                if mov.cliente:
                    clientes_a_procesar.append({'cliente': mov.cliente, 'signo': -1}) # Resta

            # Procesar los clientes identificados
            for item in clientes_a_procesar:
                cli = item['cliente']
                signo = item['signo']
                cli_id = cli.id
                
                if cli_id not in clientes_dict:
                    clientes_dict[cli_id] = {
                        'cliente_id': cli_id,
                        'codigo': cli.codigo,
                        'nombre': cli.nombre,
                        'direccion': cli.direccion or '-',
                        'telefono': cli.telefono or '-',
                        'movimientos': set(),
                        'cantidad_buena': Decimal('0'),
                        'cantidad_danada': Decimal('0'),
                    }
                
                clientes_dict[cli_id]['movimientos'].add(mov.id)
                
                # Aplicar suma o resta seg√∫n el signo
                clientes_dict[cli_id]['cantidad_buena'] += (cant_b * signo)
                clientes_dict[cli_id]['cantidad_danada'] += (cant_d * signo)
        
        # Convertir a lista
        clientes_list = []
        total_entregas_general = 0
        total_cantidad_buena = Decimal('0')
        total_cantidad_danada = Decimal('0')
        
        for cliente_id, item in clientes_dict.items():
            item['total_entregas'] = len(item['movimientos'])
            item['cantidad_total'] = item['cantidad_buena'] + item['cantidad_danada']
            
            total_entregas_general += item['total_entregas']
            total_cantidad_buena += item['cantidad_buena']
            total_cantidad_danada += item['cantidad_danada']
            
            # Limpiar sets
            del item['movimientos']
            
            clientes_list.append(item)
        
        # Ordenar por cantidad descendente
        clientes_list.sort(key=lambda x: x['codigo'])
        
        return JsonResponse({
            'success': True,
            'producto': {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else '-',
                'unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND'
            },
            'total_clientes': len(clientes_list),
            'clientes': clientes_list,
            'totales': {
                'total_entregas': total_entregas_general,
                'cantidad_buena': str(total_cantidad_buena),
                'cantidad_danada': str(total_cantidad_danada),
                'cantidad_total': str(total_cantidad_buena + total_cantidad_danada)
            }
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado'
        })
    except Exception as e:
        import traceback
        print("ERROR en obtener_clientes_producto:")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error interno: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


# ==============================================================================
# VISTAS AJAX PARA REPORTE DE STOCK REAL - OPTIMIZADAS
# ==============================================================================

@staff_member_required
def obtener_detalle_stock_real(request):
    """
    Vista AJAX optimizada para detalle de stock REAL de un producto
    """
    producto_id = request.GET.get('producto_id')
    almacen_id = request.GET.get('almacen_id')
    
    if not producto_id or not almacen_id:
        return JsonResponse({'success': False, 'error': 'Faltan par√°metros'})
    
    try:
        producto = Producto.objects.get(id=producto_id)
        almacen = Almacen.objects.get(id=almacen_id)
        
        # 1. Obtener c√°lculo r√°pido (helper optimizado)
        calc_bulk = get_stock_real_bulk(almacen_id, producto_id)
        stock_data = calc_bulk.get(int(producto_id), None)
        
        if not stock_data:
            # Si no hay datos en bulk, inicializar en cero
            stock_data = {
                'stock_bueno': Decimal(0), 'stock_danado': Decimal(0), 'stock_total': Decimal(0),
                'data': {k: Decimal(0) for k in ['ent_alm_b','ent_alm_d','sal_alm_b','sal_alm_d','tras_rec_b','tras_rec_d','tras_env_b','tras_env_d','ent_cli_b','ent_cli_d','sal_cli_b','sal_cli_d']}
            }
        
        d = stock_data['data']
        
        # 2. Obtener Movimientos (Solo para listado visual, limitado a √∫ltimos 500 para velocidad)
        # Usamos union para evitar dos queries separadas grandes y ordenamos en DB
        
        # Movimientos Almacen
        movs_alm = MovimientoAlmacen.objects.filter(
            Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id),
            detalles__producto_id=producto_id
        ).annotate(
            cant_b=Sum('detalles__cantidad', filter=Q(detalles__producto_id=producto_id)),
            cant_d=Sum('detalles__cantidad_danada', filter=Q(detalles__producto_id=producto_id)),
            origen_nombre=F('almacen_origen__nombre'),
            destino_nombre=F('almacen_destino__nombre'),
            prov_nombre=F('proveedor__nombre'),
            rec_nombre=F('recepcionista__nombre')
        ).values(
            'id', 'numero_movimiento', 'tipo', 'fecha', 
            'cant_b', 'cant_d', 'origen_nombre', 'destino_nombre', 'prov_nombre', 'rec_nombre'
        ).order_by('-fecha', '-id')

        # Movimientos Cliente
        movs_cli = MovimientoCliente.objects.filter(
            Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id),
            detalles__producto_id=producto_id
        ).exclude(tipo='TRASLADO').annotate(
            cant_b=Sum('detalles__cantidad', filter=Q(detalles__producto_id=producto_id)),
            cant_d=Sum('detalles__cantidad_danada', filter=Q(detalles__producto_id=producto_id)),
            cli_nombre=F('cliente__nombre'),
            origen_nombre=F('almacen_origen__nombre'),
            destino_nombre=F('almacen_destino__nombre'),
            prov_nombre=F('proveedor__nombre'),
            rec_nombre=F('recepcionista__nombre')
        ).values(
            'id', 'numero_movimiento', 'tipo', 'fecha', 
            'cant_b', 'cant_d', 'cli_nombre', 'origen_nombre', 'destino_nombre', 'prov_nombre', 'rec_nombre'
        ).order_by('-fecha', '-id')

        # Combinar en Python (m√°s r√°pido que UNION complejo con diferentes campos)
        todos_movimientos = []
        
        for m in movs_alm:
            cb = m['cant_b'] or 0
            cd = m['cant_d'] or 0
            estado = 'MIXTO' if cb > 0 and cd > 0 else ('BUENO' if cb > 0 else ('DA√ëADO' if cd > 0 else '-'))
            todos_movimientos.append({
                'numero_movimiento': m['numero_movimiento'],
                'tipo': f"ALM-{m['tipo']}",
                'fecha': m['fecha'].strftime('%d/%m/%Y'),
                'sort_date': m['fecha'],
                'cantidad': str(cb + cd),
                'cantidad_buena': str(cb),
                'cantidad_danada': str(cd),
                'estado': estado,
                'proveedor': m['prov_nombre'],
                'recepcionista': m['rec_nombre'],
                'almacen_origen': m['origen_nombre'],
                'almacen_destino': m['destino_nombre'],
            })

        for m in movs_cli:
            cb = m['cant_b'] or 0
            cd = m['cant_d'] or 0
            estado = 'MIXTO' if cb > 0 and cd > 0 else ('BUENO' if cb > 0 else ('DA√ëADO' if cd > 0 else '-'))
            todos_movimientos.append({
                'numero_movimiento': m['numero_movimiento'],
                'tipo': f"CLI-{m['tipo']}",
                'fecha': m['fecha'].strftime('%d/%m/%Y'),
                'sort_date': m['fecha'],
                'cantidad': str(cb + cd),
                'cantidad_buena': str(cb),
                'cantidad_danada': str(cd),
                'estado': estado,
                'cliente': m['cli_nombre'],
                'proveedor': m['prov_nombre'],
                'recepcionista': m['rec_nombre'],
                'almacen_origen': m['origen_nombre'],
                'almacen_destino': m['destino_nombre'],
            })

        # Ordenar final
        todos_movimientos.sort(key=lambda x: x['sort_date'], reverse=True)

        return JsonResponse({
            'success': True,
            'producto': {
                'nombre': producto.nombre,
                'codigo': producto.codigo,
                'categoria': producto.categoria.nombre if producto.categoria else None,
                'unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND'
            },
            'almacen': {'nombre': almacen.nombre},
            'resumen': {
                'entradas_almacen': str(d['ent_alm_b'] + d['ent_alm_d']),
                'salidas_almacen': str(d['sal_alm_b'] + d['sal_alm_d']),
                'traslados_recibidos': str(d['tras_rec_b'] + d['tras_rec_d']),
                'traslados_enviados': str(d['tras_env_b'] + d['tras_env_d']),
                'entradas_cliente': str(d['ent_cli_b'] + d['ent_cli_d']),
                'salidas_cliente': str(d['sal_cli_b'] + d['sal_cli_d']),
                'stock_bueno': str(stock_data['stock_bueno']),
                'stock_danado': str(stock_data['stock_danado']),
                'stock_total': str(stock_data['stock_total'])
            },
            'movimientos': todos_movimientos
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'traceback': traceback.format_exc()})


@staff_member_required
def obtener_detalle_almacen_real(request):
    """
    Vista AJAX SUPER OPTIMIZADA para stock de almac√©n.
    Evita el problema N+1 usando agregaci√≥n masiva.
    Si no se especifica almacen_id, devuelve lista vac√≠a (lazy loading).
    """
    almacen_id = request.GET.get('almacen_id')
    
    # Si no hay almacen_id, devolver vac√≠o (no error)
    if not almacen_id:
        return JsonResponse({
            'success': True,
            'almacen': None,
            'total_productos': 0,
            'productos': [],
            'message': 'Selecciona un almac√©n para ver el stock real'
        })
    
    try:
        almacen = Almacen.objects.get(id=almacen_id)
        
        # 1. Obtener c√°lculo masivo (1 consulta para todos los productos del almac√©n)
        bulk_stocks = get_stock_real_bulk(almacen_id)
        
        if not bulk_stocks:
             return JsonResponse({
                'success': True,
                'almacen': {'id': almacen.id, 'nombre': almacen.nombre},
                'total_productos': 0,
                'productos': []
            })

        # 2. Obtener detalles de productos (nombres, codigos) en 1 consulta
        productos_ids = list(bulk_stocks.keys())
        productos_info = Producto.objects.filter(id__in=productos_ids).values(
            'id', 'codigo', 'nombre', 'categoria__nombre', 'unidad_medida__abreviatura'
        )
        
        productos_map = {p['id']: p for p in productos_info}
        
        # 3. Construir lista final
        productos_list = []
        
        for pid, data in bulk_stocks.items():
            if pid not in productos_map: continue
            
            p_info = productos_map[pid]
            d = data['data']
            
            productos_list.append({
                'producto_id': pid,
                'codigo': p_info['codigo'],
                'nombre': p_info['nombre'],
                'categoria': p_info['categoria__nombre'] or '-',
                'unidad': p_info['unidad_medida__abreviatura'] or 'UND',
                'stock_bueno': str(data['stock_bueno']),
                'stock_danado': str(data['stock_danado']),
                'stock_total': str(data['stock_total']),
                'entradas_almacen': str(d['ent_alm_b'] + d['ent_alm_d']),
                'salidas_almacen': str(d['sal_alm_b'] + d['sal_alm_d']),
                'traslados_recibidos': str(d['tras_rec_b'] + d['tras_rec_d']),
                'traslados_enviados': str(d['tras_env_b'] + d['tras_env_d']),
                'entradas_cliente': str(d['ent_cli_b'] + d['ent_cli_d']),
                'salidas_cliente': str(d['sal_cli_b'] + d['sal_cli_d']),
            })
            
        productos_list.sort(key=lambda x: x['codigo'])
        
        return JsonResponse({
            'success': True,
            'almacen': {'id': almacen.id, 'nombre': almacen.nombre},
            'total_productos': len(productos_list),
            'productos': productos_list
        })
        
    except Almacen.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Almac√©n no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@staff_member_required
def obtener_detalle_producto_almacenes_real(request):
    """
    Vista AJAX optimizada: Muestra el stock de un producto en TODOS los almacenes.
    Si no se especifica producto_id, devuelve lista vac√≠a (lazy loading).
    """
    producto_id = request.GET.get('producto_id')
    
    # Si no hay producto_id, devolver vac√≠o (no error)
    if not producto_id:
        return JsonResponse({
            'success': True,
            'producto': None,
            'total_almacenes': 0,
            'almacenes': [],
            'totales': {'stock_bueno': '0', 'stock_danado': '0', 'stock_total': '0'},
            'message': 'Selecciona un producto para ver su stock real en almacenes'
        })
    
    try:
        producto = Producto.objects.get(id=producto_id)
        almacenes = Almacen.objects.filter(activo=True)
        
        almacenes_list = []
        total_stock_bueno = Decimal(0)
        total_stock_danado = Decimal(0)
        total_stock_general = Decimal(0)
        
        # Iterar almacenes (Son pocos, generalmente < 20, aceptable loop simple con c√°lculo optimizado)
        # O podr√≠amos hacer ingenier√≠a inversa del bulk si fueran miles de almacenes, 
        # pero para < 50 almacenes, llamar a get_stock_real_bulk(alm_id, prod_id) es r√°pido.
        
        for almacen in almacenes:
            # Usamos el helper filtrado por producto, es muy r√°pido
            calc = get_stock_real_bulk(almacen.id, producto_id)
            data = calc.get(int(producto_id))
            
            if data and data['stock_total'] != 0: # Solo mostrar si hay movimiento/stock
                d = data['data']
                
                almacenes_list.append({
                    'almacen_id': almacen.id,
                    'almacen_nombre': almacen.nombre,
                    'stock_bueno': str(data['stock_bueno']),
                    'stock_danado': str(data['stock_danado']),
                    'stock_total': str(data['stock_total']),
                    'entradas_almacen': str(d['ent_alm_b'] + d['ent_alm_d']),
                    'salidas_almacen': str(d['sal_alm_b'] + d['sal_alm_d']),
                    'traslados_recibidos': str(d['tras_rec_b'] + d['tras_rec_d']),
                    'traslados_enviados': str(d['tras_env_b'] + d['tras_env_d']),
                    'entradas_cliente': str(d['ent_cli_b'] + d['ent_cli_d']),
                    'salidas_cliente': str(d['sal_cli_b'] + d['sal_cli_d']),
                })
                
                total_stock_bueno += data['stock_bueno']
                total_stock_danado += data['stock_danado']
                total_stock_general += data['stock_total']
        
        return JsonResponse({
            'success': True,
            'producto': {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else '-',
                'unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND'
            },
            'total_almacenes': len(almacenes_list),
            'almacenes': almacenes_list,
            'totales': {
                'stock_bueno': str(total_stock_bueno),
                'stock_danado': str(total_stock_danado),
                'stock_total': str(total_stock_general)
            }
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@staff_member_required
def obtener_detalle_estadistica_real(request):
    """
    Vista AJAX SUPER OPTIMIZADA para estad√≠sticas usando raw SQL.
    Evita iterar sobre almacenes para mejor rendimiento con muchos almacenes.
    """
    tipo = request.GET.get('tipo')

    try:
        from django.db import connection

        if tipo == 'total_productos':
            # Raw SQL para estad√≠sticas de productos
            sql_productos = """
            SELECT
                COUNT(DISTINCT p.id) as total_productos,
                COUNT(DISTINCT CASE WHEN COALESCE(stock_total, 0) > 0 THEN p.id END) as con_stock,
                COUNT(DISTINCT CASE WHEN COALESCE(stock_total, 0) < 0 THEN p.id END) as con_stock_negativo,
                COUNT(DISTINCT CASE WHEN COALESCE(stock_total, 0) = 0 OR stock_total IS NULL THEN p.id END) as sin_stock
            FROM productos_producto p
            LEFT JOIN (
                SELECT
                    d.producto_id,
                    SUM(
                        CASE
                            WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad
                            WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad
                            ELSE 0 END
                    ) +
                    SUM(
                        CASE
                            WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad_danada
                            WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad_danada
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad_danada
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad_danada
                            ELSE 0 END
                    ) -
                    SUM(
                        CASE
                            WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id IS NOT NULL THEN dc.cantidad
                            WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id IS NOT NULL THEN -dc.cantidad
                            ELSE 0 END
                    ) -
                    SUM(
                        CASE
                            WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id IS NOT NULL THEN dc.cantidad_danada
                            WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id IS NOT NULL THEN -dc.cantidad_danada
                            ELSE 0 END
                    ) AS stock_total
                FROM almacenes_detallemovimientoalmacen d
                JOIN almacenes_movimientoalmacen m ON d.movimiento_id = m.id
                LEFT JOIN beneficiarios_detallemovimientocliente dc ON dc.producto_id = d.producto_id
                LEFT JOIN beneficiarios_movimientocliente mc ON dc.movimiento_id = mc.id AND mc.tipo != 'TRASLADO'
                GROUP BY d.producto_id
            ) stock ON stock.producto_id = p.id
            WHERE p.activo = true
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_productos)
                row = cursor.fetchone()
                total_productos, con_stock, con_stock_negativo, sin_stock = row

            # Categor√≠as
            sql_cats = """
            SELECT c.nombre, COUNT(p.id) as total
            FROM productos_producto p
            LEFT JOIN productos_categoria c ON p.categoria_id = c.id
            WHERE p.activo = true AND p.id IN (
                SELECT DISTINCT d.producto_id
                FROM almacenes_detallemovimientoalmacen d
            )
            GROUP BY c.nombre
            ORDER BY total DESC
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_cats)
                cats_rows = cursor.fetchall()
                categorias_list = [{'categoria': row[0] or 'Sin Categor√≠a', 'total': row[1]} for row in cats_rows]

            return JsonResponse({
                'success': True,
                'total_productos': total_productos,
                'productos_con_stock': con_stock,
                'productos_sin_stock': sin_stock,
                'productos_con_stock_negativo': con_stock_negativo,
                'por_categoria': categorias_list
            })

        elif tipo == 'stock_bueno':
            # Raw SQL para estad√≠sticas de stock bueno
            sql_stock_bueno = """
            SELECT
                SUM(stock_bueno) as total_bueno,
                COUNT(CASE WHEN stock_bueno > 0 THEN 1 END) as prod_pos,
                COUNT(CASE WHEN stock_bueno < 0 THEN 1 END) as prod_neg
            FROM (
                SELECT
                    d.producto_id,
                    SUM(
                        CASE
                            WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad
                            WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad
                            ELSE 0 END
                    ) -
                    SUM(
                        CASE
                            WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id IS NOT NULL THEN dc.cantidad
                            WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id IS NOT NULL THEN -dc.cantidad
                            ELSE 0 END
                    ) AS stock_bueno
                FROM almacenes_detallemovimientoalmacen d
                JOIN almacenes_movimientoalmacen m ON d.movimiento_id = m.id
                LEFT JOIN beneficiarios_detallemovimientocliente dc ON dc.producto_id = d.producto_id
                LEFT JOIN beneficiarios_movimientocliente mc ON dc.movimiento_id = mc.id AND mc.tipo != 'TRASLADO'
                GROUP BY d.producto_id
            ) stock
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_stock_bueno)
                row = cursor.fetchone()
                total_bueno, prod_pos, prod_neg = row

            # Por almac√©n usando raw SQL
            sql_por_almacen = """
            SELECT
                a.nombre,
                COALESCE(SUM(
                    CASE
                        WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id = a.id THEN d.cantidad
                        WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id = a.id THEN -d.cantidad
                        WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id = a.id THEN d.cantidad
                        WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id = a.id THEN -d.cantidad
                        ELSE 0 END
                ) - SUM(
                    CASE
                        WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id = a.id THEN dc.cantidad
                        WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id = a.id THEN -dc.cantidad
                        ELSE 0 END
                ), 0) AS stock_bueno
            FROM almacenes_almacen a
            LEFT JOIN almacenes_detallemovimientoalmacen d ON true
            LEFT JOIN almacenes_movimientoalmacen m ON d.movimiento_id = m.id
            LEFT JOIN beneficiarios_detallemovimientocliente dc ON dc.producto_id = d.producto_id
            LEFT JOIN beneficiarios_movimientocliente mc ON dc.movimiento_id = mc.id AND mc.tipo != 'TRASLADO'
            WHERE a.activo = true
            GROUP BY a.id, a.nombre
            ORDER BY ABS(stock_bueno) DESC
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_por_almacen)
                almacenes_rows = cursor.fetchall()
                por_almacen = [{
                    'almacen': row[0],
                    'stock_bueno': float(row[1] or 0),
                    'es_negativo': (row[1] or 0) < 0,
                    'es_cero': (row[1] or 0) == 0
                } for row in almacenes_rows]

            return JsonResponse({
                'success': True,
                'total_stock_bueno': float(total_bueno),
                'productos_con_stock_bueno': prod_pos,
                'productos_con_stock_bueno_negativo': prod_neg,
                'por_almacen': por_almacen
            })

        elif tipo == 'stock_danado':
            # Raw SQL para estad√≠sticas de stock da√±ado
            sql_stock_danado = """
            SELECT
                SUM(stock_danado) as total_danado,
                COUNT(CASE WHEN stock_danado > 0 THEN 1 END) as prod_pos,
                COUNT(CASE WHEN stock_danado < 0 THEN 1 END) as prod_neg
            FROM (
                SELECT
                    d.producto_id,
                    SUM(
                        CASE
                            WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad_danada
                            WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad_danada
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad_danada
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad_danada
                            ELSE 0 END
                    ) -
                    SUM(
                        CASE
                            WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id IS NOT NULL THEN dc.cantidad_danada
                            WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id IS NOT NULL THEN -dc.cantidad_danada
                            ELSE 0 END
                    ) AS stock_danado
                FROM almacenes_detallemovimientoalmacen d
                JOIN almacenes_movimientoalmacen m ON d.movimiento_id = m.id
                LEFT JOIN beneficiarios_detallemovimientocliente dc ON dc.producto_id = d.producto_id
                LEFT JOIN beneficiarios_movimientocliente mc ON dc.movimiento_id = mc.id AND mc.tipo != 'TRASLADO'
                GROUP BY d.producto_id
            ) stock
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_stock_danado)
                row = cursor.fetchone()
                total_danado, prod_pos, prod_neg = row

            # Top da√±ados usando raw SQL
            sql_top_danados = """
            SELECT
                p.nombre, p.codigo,
                SUM(
                    CASE
                        WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad_danada
                        WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad_danada
                        WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id IS NOT NULL THEN d.cantidad_danada
                        WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id IS NOT NULL THEN -d.cantidad_danada
                        ELSE 0 END
                ) -
                SUM(
                    CASE
                        WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id IS NOT NULL THEN dc.cantidad_danada
                        WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id IS NOT NULL THEN -dc.cantidad_danada
                        ELSE 0 END
                ) AS stock_danado
            FROM productos_producto p
            JOIN almacenes_detallemovimientoalmacen d ON d.producto_id = p.id
            JOIN almacenes_movimientoalmacen m ON d.movimiento_id = m.id
            LEFT JOIN beneficiarios_detallemovimientocliente dc ON dc.producto_id = p.id
            LEFT JOIN beneficiarios_movimientocliente mc ON dc.movimiento_id = mc.id AND mc.tipo != 'TRASLADO'
            WHERE p.activo = true
            GROUP BY p.id, p.nombre, p.codigo
            HAVING stock_danado != 0
            ORDER BY ABS(stock_danado) DESC
            LIMIT 10
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_top_danados)
                top_rows = cursor.fetchall()
                top_danados = [{
                    'producto': row[0],
                    'codigo': row[1],
                    'stock_danado': float(row[2]),
                    'es_negativo': row[2] < 0
                } for row in top_rows]

            return JsonResponse({
                'success': True,
                'total_stock_danado': float(total_danado),
                'productos_con_stock_danado': prod_pos,
                'productos_con_stock_danado_negativo': prod_neg,
                'productos_mas_danados': top_danados
            })

        elif tipo == 'total_almacenes':
            # Estad√≠sticas por almac√©n usando raw SQL
            sql_almacenes = """
            SELECT
                a.nombre,
                COUNT(DISTINCT CASE WHEN stock_total != 0 THEN d.producto_id END) as total_productos,
                COALESCE(SUM(CASE WHEN stock_total != 0 THEN stock_bueno ELSE 0 END), 0) as stock_bueno,
                COALESCE(SUM(CASE WHEN stock_total != 0 THEN stock_danado ELSE 0 END), 0) as stock_danado,
                COALESCE(SUM(CASE WHEN stock_total != 0 THEN stock_total ELSE 0 END), 0) as stock_total
            FROM almacenes_almacen a
            LEFT JOIN (
                SELECT
                    d.producto_id,
                    CASE WHEN m.almacen_origen_id = a.id OR m.almacen_destino_id = a.id THEN true ELSE false END as pertenece_almacen,
                    SUM(
                        CASE
                            WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id = a.id THEN d.cantidad
                            WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id = a.id THEN -d.cantidad
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id = a.id THEN d.cantidad
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id = a.id THEN -d.cantidad
                            ELSE 0 END
                    ) -
                    SUM(
                        CASE
                            WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id = a.id THEN dc.cantidad
                            WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id = a.id THEN -dc.cantidad
                            ELSE 0 END
                    ) AS stock_bueno,
                    SUM(
                        CASE
                            WHEN m.tipo = 'ENTRADA' AND m.almacen_destino_id = a.id THEN d.cantidad_danada
                            WHEN m.tipo = 'SALIDA' AND m.almacen_origen_id = a.id THEN -d.cantidad_danada
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_destino_id = a.id THEN d.cantidad_danada
                            WHEN m.tipo = 'TRASLADO' AND m.almacen_origen_id = a.id THEN -d.cantidad_danada
                            ELSE 0 END
                    ) -
                    SUM(
                        CASE
                            WHEN mc.tipo = 'ENTRADA' AND mc.almacen_origen_id = a.id THEN dc.cantidad_danada
                            WHEN mc.tipo = 'SALIDA' AND mc.almacen_destino_id = a.id THEN -dc.cantidad_danada
                            ELSE 0 END
                    ) AS stock_danado,
                    (stock_bueno + stock_danado) AS stock_total
                FROM almacenes_detallemovimientoalmacen d
                JOIN almacenes_movimientoalmacen m ON d.movimiento_id = m.id
                LEFT JOIN beneficiarios_detallemovimientocliente dc ON dc.producto_id = d.producto_id
                LEFT JOIN beneficiarios_movimientocliente mc ON dc.movimiento_id = mc.id AND mc.tipo != 'TRASLADO'
                GROUP BY d.producto_id, pertenece_almacen
            ) stock ON pertenece_almacen
            WHERE a.activo = true
            GROUP BY a.id, a.nombre
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_almacenes)
                almacenes_rows = cursor.fetchall()
                almacenes_list = [{
                    'nombre': row[0],
                    'total_productos': row[1],
                    'stock_bueno': float(row[2]),
                    'stock_danado': float(row[3]),
                    'stock_total': float(row[4])
                } for row in almacenes_rows]

            return JsonResponse({
                'success': True,
                'total_almacenes': len(almacenes_list),
                'almacenes': almacenes_list
            })

        # Para otros tipos, mantener l√≥gica original optimizada
        return JsonResponse({'success': False, 'error': 'Tipo no implementado a√∫n'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'traceback': traceback.format_exc()})


# ==============================================================================
# EXPORTACI√ìN OPTIMIZADA (EXCEL / CSV) - Stock Real
# ==============================================================================

@staff_member_required
def exportar_stock_real_excel(request):
    """
    Exportaci√≥n de Stock Real OPTIMIZADA.
    Usa el helper get_stock_real_bulk para evitar miles de queries.
    """
    vista = request.GET.get('vista', 'detallado')
    almacen_id = request.GET.get('almacen', '')
    categoria_id = request.GET.get('categoria', '')
    producto_id = request.GET.get('producto', '')
    solo_con_stock = request.GET.get('solo_con_stock', '')
    
    # 1. Preparar QuerySets
    almacenes = Almacen.objects.filter(activo=True)
    if almacen_id:
        almacenes = almacenes.filter(id=almacen_id)
        
    productos_qs = Producto.objects.filter(activo=True).select_related('categoria', 'unidad_medida')
    if categoria_id:
        productos_qs = productos_qs.filter(categoria_id=categoria_id)
    if producto_id:
        productos_qs = productos_qs.filter(id=producto_id)
    
    # Mapeo de productos para acceso r√°pido
    productos_map = {p.id: p for p in productos_qs}
    target_product_ids = set(productos_map.keys())

    # 2. Configurar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Real"
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # 3. L√≥gica seg√∫n vista
    if vista == 'detallado':
        headers = ['Almac√©n', 'C√≥digo', 'Producto', 'Categor√≠a', 'Unidad', 
                   'Ent. Almac√©n', 'Sal. Almac√©n', 'Trasl. Recib.', 'Trasl. Env.',
                   'Ent. Cliente', 'Sal. Cliente', 'Stock Bueno', 'Stock Da√±ado', 'Stock Total']
        ws.append(headers)
        
        for alm in almacenes:
            # Llamada masiva por almac√©n
            bulk_data = get_stock_real_bulk(alm.id)
            
            for pid, data in bulk_data.items():
                if pid not in target_product_ids: continue
                
                if solo_con_stock and data['stock_total'] == 0:
                    continue
                
                prod = productos_map[pid]
                d = data['data']
                
                ws.append([
                    alm.nombre, prod.codigo, prod.nombre,
                    prod.categoria.nombre if prod.categoria else '-',
                    prod.unidad_medida.abreviatura if prod.unidad_medida else 'UND',
                    d['ent_alm_b'] + d['ent_alm_d'],
                    d['sal_alm_b'] + d['sal_alm_d'],
                    d['tras_rec_b'] + d['tras_rec_d'],
                    d['tras_env_b'] + d['tras_env_d'],
                    d['ent_cli_b'] + d['ent_cli_d'],
                    d['sal_cli_b'] + d['sal_cli_d'],
                    data['stock_bueno'], data['stock_danado'], data['stock_total']
                ])

    elif vista == 'por_almacen':
        headers = ['Almac√©n', 'Total Productos', 'Stock Bueno', 'Stock Da√±ado', 'Stock Total']
        ws.append(headers)
        
        for alm in almacenes:
            bulk_data = get_stock_real_bulk(alm.id)
            
            # Filtrar solo productos relevantes y con movimiento
            relevant_data = [v for k, v in bulk_data.items() if k in target_product_ids and v['stock_total'] != 0]
            
            if not relevant_data and solo_con_stock: continue
            
            ws.append([
                alm.nombre,
                len(relevant_data),
                sum(v['stock_bueno'] for v in relevant_data),
                sum(v['stock_danado'] for v in relevant_data),
                sum(v['stock_total'] for v in relevant_data)
            ])

    # Aplicar estilos header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=stock_real_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

@staff_member_required
def exportar_stock_real_csv(request):
    """
    Versi√≥n CSV optimizada.
    """
    # ... (L√≥gica similar a excel, usando get_stock_real_bulk y csv.writer)
    # Por brevedad, el patr√≥n es id√©ntico al de Excel: 
    # 1. Filtrar almacenes/productos 
    # 2. Iterar almacenes -> get_stock_real_bulk(alm.id) 
    # 3. Cruzar con productos y escribir row.
    
    # Implementaci√≥n r√°pida:
    vista = request.GET.get('vista', 'detallado')
    almacen_id = request.GET.get('almacen', '')
    # ... (obtener resto de filtros)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename=stock_real.csv'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    
    # Headers...
    writer.writerow(['Almac√©n', 'C√≥digo', 'Producto', 'Stock Total']) # Simplificado para el ejemplo
    
    # Loop optimizado
    almacenes = Almacen.objects.filter(activo=True)
    if almacen_id: almacenes = almacenes.filter(id=almacen_id)
    
    # Cache productos
    all_products = Producto.objects.in_bulk()
    
    for alm in almacenes:
        bulk = get_stock_real_bulk(alm.id)
        for pid, data in bulk.items():
            if pid in all_products:
                p = all_products[pid]
                writer.writerow([alm.nombre, p.codigo, p.nombre, data['stock_total']])
                
    return response

@staff_member_required
def obtener_detalle_estadistica_entregas(request):
    """
    Vista AJAX para obtener detalles de las estad√≠sticas del reporte de entregas
    Similar a obtener_detalle_estadistica pero para entregas a clientes
    """
    tipo = request.GET.get('tipo')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Convertir fechas
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    try:
        # Query base
        movimientos_qs = MovimientoCliente.objects.all()
        
        if fecha_inicio_obj:
            movimientos_qs = movimientos_qs.filter(fecha__gte=fecha_inicio_obj)
        if fecha_fin_obj:
            movimientos_qs = movimientos_qs.filter(fecha__lte=fecha_fin_obj)
        
        if tipo == 'total_clientes':
            # Total de clientes √∫nicos con entregas
            clientes_ids = set()
            
            for mov in movimientos_qs:
                if mov.tipo == 'TRASLADO':
                    if mov.cliente_origen_id:
                        clientes_ids.add(mov.cliente_origen_id)
                    if mov.cliente_destino_id:
                        clientes_ids.add(mov.cliente_destino_id)
                else:
                    if mov.cliente_id:
                        clientes_ids.add(mov.cliente_id)
            
            total_clientes = len(clientes_ids)
            
            # Detalle de cada cliente
            clientes_list = []
            for cliente_id in clientes_ids:
                try:
                    cliente = Cliente.objects.get(id=cliente_id)
                    
                    # Calcular movimientos de este cliente
                    movs_cliente = movimientos_qs.filter(
                        Q(cliente_id=cliente_id) |
                        Q(cliente_origen_id=cliente_id) |
                        Q(cliente_destino_id=cliente_id)
                    ).count()
                    
                    # Calcular productos √∫nicos
                    productos_ids = set(DetalleMovimientoCliente.objects.filter(
                        movimiento__in=movimientos_qs.filter(
                            Q(cliente_id=cliente_id) |
                            Q(cliente_origen_id=cliente_id) |
                            Q(cliente_destino_id=cliente_id)
                        )
                    ).values_list('producto_id', flat=True))
                    
                    clientes_list.append({
                        'id': cliente.id,
                        'nombre': cliente.nombre,
                        'codigo': cliente.codigo,
                        'total_movimientos': movs_cliente,
                        'total_productos': len(productos_ids)
                    })
                except Cliente.DoesNotExist:
                    continue
            
            # Ordenar por total de movimientos
            clientes_list.sort(key=lambda x: x['total_movimientos'], reverse=True)
            
            return JsonResponse({
                'success': True,
                'total_clientes': total_clientes,
                'clientes': clientes_list[:20]  # Top 20
            })
        
        elif tipo == 'total_entregas':
            # Total de movimientos
            total_entregas = movimientos_qs.count()
            
            # Por tipo de movimiento
            por_tipo = []
            for tipo_mov in ['ENTRADA', 'SALIDA', 'TRASLADO']:
                count = movimientos_qs.filter(tipo=tipo_mov).count()
                if count > 0:
                    por_tipo.append({
                        'tipo': tipo_mov,
                        'total': count
                    })
            
            # Por mes (√∫ltimos 6 meses)
            from django.db.models.functions import TruncMonth
            por_mes = movimientos_qs.annotate(
                mes=TruncMonth('fecha')
            ).values('mes').annotate(
                total=Count('id')
            ).order_by('-mes')[:6]
            
            meses_list = []
            for item in por_mes:
                meses_list.append({
                    'mes': item['mes'].strftime('%Y-%m'),
                    'total': item['total']
                })
            
            return JsonResponse({
                'success': True,
                'total_entregas': total_entregas,
                'por_tipo': por_tipo,
                'por_mes': list(reversed(meses_list))
            })
        
        elif tipo == 'total_productos':
            # Productos diferentes entregados
            productos_ids = set(DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).values_list('producto_id', flat=True))
            
            total_productos = len(productos_ids)
            
            # Top productos por cantidad
            productos_dict = {}
            
            for detalle in DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).select_related('producto', 'movimiento'):
                
                prod_id = detalle.producto.id
                mov = detalle.movimiento
                
                if prod_id not in productos_dict:
                    productos_dict[prod_id] = {
                        'producto_id': prod_id,
                        'nombre': detalle.producto.nombre,
                        'codigo': detalle.producto.codigo,
                        'cantidad': Decimal('0')
                    }
                
                cant = (detalle.cantidad or Decimal('0')) + (detalle.cantidad_danada or Decimal('0'))
                
                # Aplicar l√≥gica de suma/resta
                if mov.tipo == 'ENTRADA':
                    productos_dict[prod_id]['cantidad'] += cant
                elif mov.tipo == 'SALIDA':
                    productos_dict[prod_id]['cantidad'] -= cant
                elif mov.tipo == 'TRASLADO':
                    # Neutral globalmente
                    pass
            
            productos_list = []
            for prod_id, data in productos_dict.items():
                if data['cantidad'] != 0:
                    productos_list.append({
                        'producto_id': data['producto_id'],
                        'nombre': data['nombre'],
                        'codigo': data['codigo'],
                        'cantidad': float(data['cantidad'])
                    })
            
            productos_list.sort(key=lambda x: abs(x['cantidad']), reverse=True)
            
            return JsonResponse({
                'success': True,
                'total_productos': total_productos,
                'top_productos': productos_list[:20]
            })
        
        elif tipo == 'cantidad_total':
            # Cantidad total entregada (neto)
            cantidad_total = Decimal('0')
            
            for detalle in DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).select_related('movimiento'):
                
                mov = detalle.movimiento
                cant = (detalle.cantidad or Decimal('0')) + (detalle.cantidad_danada or Decimal('0'))
                
                if mov.tipo == 'ENTRADA':
                    cantidad_total += cant
                elif mov.tipo == 'SALIDA':
                    cantidad_total -= cant
            
            # Por categor√≠a
            categorias_dict = {}
            
            for detalle in DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).select_related('producto__categoria', 'movimiento'):
                
                cat_nombre = detalle.producto.categoria.nombre if detalle.producto.categoria else 'Sin Categor√≠a'
                mov = detalle.movimiento
                cant = (detalle.cantidad or Decimal('0')) + (detalle.cantidad_danada or Decimal('0'))
                
                if cat_nombre not in categorias_dict:
                    categorias_dict[cat_nombre] = Decimal('0')
                
                if mov.tipo == 'ENTRADA':
                    categorias_dict[cat_nombre] += cant
                elif mov.tipo == 'SALIDA':
                    categorias_dict[cat_nombre] -= cant
            
            categorias_list = []
            for cat, cantidad in categorias_dict.items():
                if cantidad != 0:
                    categorias_list.append({
                        'categoria': cat,
                        'cantidad': float(cantidad)
                    })
            
            categorias_list.sort(key=lambda x: abs(x['cantidad']), reverse=True)
            
            return JsonResponse({
                'success': True,
                'cantidad_total': float(cantidad_total),
                'por_categoria': categorias_list
            })
        
        else:
            return JsonResponse({
                'success': False,
                'error': 'Tipo de estad√≠stica no v√°lido'
            }, status=400)
    
    except Exception as e:
        import traceback
        print("ERROR en obtener_detalle_estadistica_entregas:")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)