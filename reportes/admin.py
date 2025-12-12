from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse, JsonResponse
from django.urls import path
from django.shortcuts import render
from django.utils.html import format_html
from django.db.models import Sum, Count, Q, F
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime, timedelta
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal

from .models import ReporteMovimiento, ReporteEntregas, ReporteStock
from .models import ReporteStockReal
from almacenes.models import MovimientoAlmacen, DetalleMovimientoAlmacen, Almacen
from beneficiarios.models import MovimientoCliente, DetalleMovimientoCliente, Cliente
from productos.models import Producto, Categoria
from proveedores.models import Proveedor
from recepcionistas.models import Recepcionista
from . import views


# ==========================================
# REPORTE DE MOVIMIENTOS - ACTUALIZADO
# ==========================================
class ReporteMovimientoAdmin(admin.ModelAdmin):
    """
    Admin personalizado para reportes de movimientos con PAGINACIÓN
    """
    
    change_list_template = 'admin/reportes/reporte_movimientos_list.html'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False
    
    def has_change_permission(self, request, obj=None):
        return True
    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        
        # 1. Filtros
        vista = request.GET.get('vista', 'detallado')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        cliente_id = request.GET.get('cliente', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        
        # Paginación
        page = request.GET.get('page', 1)
        items_por_pagina = int(request.GET.get('items_por_pagina', '100'))
        
        # Conversión de fechas
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date() if fecha_inicio else None
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date() if fecha_fin else None
        
        cliente_id_int = int(cliente_id) if cliente_id else None
        categoria_id_int = int(categoria_id) if categoria_id else None
        producto_id_int = int(producto_id) if producto_id else None

        entregas = []
        estadisticas = {}
        productos_top = [] # Se pueden calcular aparte si es necesario
        resumen_clientes = []

        # =================================================
        # 🚀 LLAMADA A LA LÓGICA OPTIMIZADA
        # =================================================
        if vista == 'detallado':
            # Obtenemos la lista procesada (mucho más pequeña que los movimientos crudos)
            data_procesada = ReporteEntregas.obtener_entregas_optimizadas(
                fecha_inicio=fecha_inicio_obj,
                fecha_fin=fecha_fin_obj,
                cliente_id=cliente_id_int,
                categoria_id=categoria_id_int,
                producto_id=producto_id_int
            )
            
            # Ordenar en Python (rápido porque la lista ya está resumida)
            # Ordenamos por Cliente -> Producto
            entregas = sorted(data_procesada, key=lambda x: (x['cliente'].nombre, x['producto'].nombre))
            
            # Calcular estadísticas sobre la lista resumida
            cantidad_total_global = sum(item['stock_total'] for item in entregas)
            clientes_unicos = set(item['cliente'].id for item in entregas)
            productos_unicos = set(item['producto'].id for item in entregas)
            
            estadisticas = {
                'cantidad_total': cantidad_total_global,
                'total_clientes_unicos': len(clientes_unicos),
                'total_productos_unicos': len(productos_unicos),
            }

        elif vista == 'por_cliente':
             # Reutilizamos la lógica optimizada pero agrupamos un nivel más
            data_procesada = ReporteEntregas.obtener_entregas_optimizadas(
                fecha_inicio=fecha_inicio_obj, fecha_fin=fecha_fin_obj, 
                cliente_id=cliente_id_int, categoria_id=categoria_id_int, producto_id=producto_id_int
            )
            
            temp_cli = {}
            for item in data_procesada:
                cid = item['cliente'].id
                if cid not in temp_cli:
                    temp_cli[cid] = {
                        'cliente_id': cid,
                        'cliente_nombre': item['cliente'].nombre,
                        'cliente_codigo': item['cliente'].codigo,
                        'cliente_direccion': item['cliente'].direccion,
                        'cliente_telefono': item['cliente'].telefono,
                        'total_entregas': 0, # Esto es aproximado en vista agrupada
                        'total_productos': 0,
                        'cantidad_total': 0
                    }
                temp_cli[cid]['total_productos'] += 1
                # Nota: total_entregas es difícil de sumar exacto sin sets, pero es una aprox válida
                temp_cli[cid]['total_entregas'] += item['total_entregas'] 
                temp_cli[cid]['cantidad_total'] += item['stock_total']
            
            entregas = list(temp_cli.values())
            entregas.sort(key=lambda x: x['cliente_nombre'])

        # =================================================
        # PAGINACIÓN (Ahora sí es segura porque 'entregas' es una lista resumida)
        # =================================================
        # Si tienes 10,000 movimientos, 'entregas' tendrá quizás 500 filas (Combinaciones Cliente-Producto)
        # Esto es manejable en memoria.
        
        paginator = Paginator(entregas, items_por_pagina)
        try:
            entregas_paginadas = paginator.page(page)
        except PageNotAnInteger:
            entregas_paginadas = paginator.page(1)
        except EmptyPage:
            entregas_paginadas = paginator.page(paginator.num_pages)

        context = {
            **self.admin_site.each_context(request),
            'title': _('Reportes de Entregas (Optimizado)'),
            'entregas': entregas_paginadas,
            'estadisticas': estadisticas,
            'productos_top': [], # Puedes reactivarlo si haces una query dedicada pequeña
            'resumen_clientes': [], 
            'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
            'categorias': Categoria.objects.all(),
            'productos': Producto.objects.filter(activo=True).order_by('codigo'),
            'filtros': {
                'vista': vista, 'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin,
                'cliente': cliente_id, 'categoria': categoria_id, 'producto': producto_id,
            },
            'page_obj': entregas_paginadas,
            'paginator': paginator,
            'items_por_pagina': items_por_pagina,
            'opciones_items': [50, 100, 200, 500],
            'opts': self.model._meta,
        }
        return render(request, self.change_list_template, context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('api/numeros_movimiento/', 
                self.admin_site.admin_view(self.obtener_numeros_movimiento_ajax),
                name='reportes_reportemovimiento_numeros_json'),
            path('exportar-excel/', 
                self.admin_site.admin_view(views.exportar_movimientos_excel), 
                name='reportes_reportemovimiento_exportar_excel'),
            path('exportar-csv/', 
                self.admin_site.admin_view(views.exportar_movimientos_csv), 
                name='reportes_reportemovimiento_exportar_csv'),
            path('obtener-datos-graficos-movimientos/', # <-- ADICIÓN DE LA URL
                self.admin_site.admin_view(views.obtener_datos_graficos_movimientos),
                name='reportes_movimientos_datos_graficos'),
        ]
        return custom_urls + urls

    def obtener_numeros_movimiento_ajax(self, request):
        """Endpoint AJAX para obtener números de movimiento filtrados"""
        tipo_reporte = request.GET.get('tipo_reporte', 'almacen')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        tipo_movimiento = request.GET.get('tipo_movimiento', '')
        almacen_id = request.GET.get('almacen', '')
        cliente_id = request.GET.get('cliente', '')
        proveedor_id = request.GET.get('proveedor', '')
        recepcionista_id = request.GET.get('recepcionista', '')
        
        # Convertir fechas
        fecha_inicio_obj = None
        fecha_fin_obj = None
        
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Obtener números según tipo de reporte
        if tipo_reporte == 'almacen':
            movimientos_query = MovimientoAlmacen.objects.all()
            
            if fecha_inicio_obj:
                movimientos_query = movimientos_query.filter(fecha__gte=fecha_inicio_obj)
            if fecha_fin_obj:
                movimientos_query = movimientos_query.filter(fecha__lte=fecha_fin_obj)
            if almacen_id:
                movimientos_query = movimientos_query.filter(
                    Q(almacen_origen_id=almacen_id) | Q(almacen_destino_id=almacen_id)
                )
            if tipo_movimiento:
                movimientos_query = movimientos_query.filter(tipo=tipo_movimiento)
            if proveedor_id:
                movimientos_query = movimientos_query.filter(proveedor_id=proveedor_id)
            if recepcionista_id:
                movimientos_query = movimientos_query.filter(recepcionista_id=recepcionista_id)
                
            numeros = movimientos_query.values_list('numero_movimiento', flat=True).distinct().order_by('-numero_movimiento')[:200]
        else:
            movimientos_query = MovimientoCliente.objects.all()
            
            if fecha_inicio_obj:
                movimientos_query = movimientos_query.filter(fecha__gte=fecha_inicio_obj)
            if fecha_fin_obj:
                movimientos_query = movimientos_query.filter(fecha__lte=fecha_fin_obj)
            if cliente_id:
                movimientos_query = movimientos_query.filter(cliente_id=cliente_id)
            if tipo_movimiento:
                movimientos_query = movimientos_query.filter(tipo=tipo_movimiento)
            if proveedor_id:
                movimientos_query = movimientos_query.filter(proveedor_id=proveedor_id)
            if recepcionista_id:
                movimientos_query = movimientos_query.filter(recepcionista_id=recepcionista_id)
                
            numeros = movimientos_query.values_list('numero_movimiento', flat=True).distinct().order_by('-numero_movimiento')[:200]
        
        return JsonResponse({
            'numeros': list(numeros)
        })

# ==========================================
# REPORTE DE ENTREGAS - CORREGIDO CON TRASLADOS SEPARADOS
# ==========================================
class ReporteEntregasAdmin(admin.ModelAdmin):
    """
    Admin para reportes de entregas a clientes
    """
    
    change_list_template = 'admin/reportes/reporte_entregas_list.html'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False
    
    def has_change_permission(self, request, obj=None):
        return True
    
    def changelist_view(self, request, extra_context=None):
        """Vista personalizada con filtros para entregas a clientes"""
        extra_context = extra_context or {}
        
        # Obtener parámetros de filtro
        vista = request.GET.get('vista', 'detallado')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        cliente_id = request.GET.get('cliente', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        
        # ✅ PAGINACIÓN - Definir variables al inicio para evitar UnboundLocalError
        page = request.GET.get('page', 1)
        items_por_pagina = request.GET.get('items_por_pagina', '100')
        
        # Validar items por página
        try:
            items_por_pagina = int(items_por_pagina)
            if items_por_pagina not in [50, 100, 200, 500]:
                items_por_pagina = 100
        except (ValueError, TypeError):
            items_por_pagina = 100
        
        # Convertir fechas
        fecha_inicio_obj = None
        fecha_fin_obj = None
        
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Query base: TODOS los movimientos de clientes
        movimientos_qs = MovimientoCliente.objects.all()
        
        if fecha_inicio_obj:
            movimientos_qs = movimientos_qs.filter(fecha__gte=fecha_inicio_obj)
        
        if fecha_fin_obj:
            movimientos_qs = movimientos_qs.filter(fecha__lte=fecha_fin_obj)
        
        if cliente_id:
            # ✅ CORRECCIÓN: Filtrar también por cliente_origen y cliente_destino en traslados
            movimientos_qs = movimientos_qs.filter(
                Q(cliente_id=cliente_id) | 
                Q(cliente_origen_id=cliente_id) | 
                Q(cliente_destino_id=cliente_id)
            )

        # Inicializar variables
        entregas = []
        productos_top = []
        resumen_clientes = []
        estadisticas = {}
        
        if vista == 'detallado':
            # Verificar si se debe mostrar todos los productos
            mostrar_todos = request.GET.get('mostrar_todos', '') == '1'
            entregas_dict = {}

            if mostrar_todos:
                # Pre-llenar con todos los productos/clientes activos
                clientes_activos = Cliente.objects.filter(activo=True)
                if cliente_id:  # ✅ YA ESTÁ CORRECTO AQUÍ
                    clientes_activos = clientes_activos.filter(id=cliente_id)
                                    
                productos_activos = Producto.objects.filter(activo=True)
                if categoria_id:
                    productos_activos = productos_activos.filter(categoria_id=categoria_id)
                if producto_id:
                    productos_activos = productos_activos.filter(id=producto_id)
                
                for cliente in clientes_activos:
                    for producto in productos_activos:
                        key = f"{cliente.id}_{producto.id}"
                        entregas_dict[key] = {
                            'cliente_id': cliente.id,
                            'cliente_nombre': cliente.nombre,
                            'cliente_codigo': cliente.codigo,
                            'cliente_direccion': cliente.direccion or '-',
                            'producto_id': producto.id,
                            'producto_codigo': producto.codigo,
                            'producto_nombre': producto.nombre,
                            'producto_categoria': producto.categoria.nombre if producto.categoria else '-',
                            'producto_unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND',
                            'total_entregas': set(),
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),    # ✅ NUEVO
                            'cantidad_traslado_destino': Decimal('0'),   # ✅ NUEVO
                            'cantidad_buena': Decimal('0'),
                            'cantidad_danada': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
            
            # Obtener los detalles
            entregas_qs = DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).select_related(
                'movimiento', 'movimiento__cliente', 'producto',
                'producto__categoria', 'producto__unidad_medida'
            )
            
            if categoria_id:
                entregas_qs = entregas_qs.filter(producto__categoria_id=categoria_id)
            if producto_id:
                entregas_qs = entregas_qs.filter(producto_id=producto_id)
            
            # ✅ CORRECCIÓN: Procesar movimientos considerando traslados
            for detalle in entregas_qs:
                mov = detalle.movimiento
                cant_b = detalle.cantidad or Decimal('0')
                cant_d = detalle.cantidad_danada or Decimal('0')
                cant_total = cant_b + cant_d
                
                # ✅ PARA TRASLADOS: Crear DOS registros (origen y destino)
                if mov.tipo == 'TRASLADO' and mov.cliente_origen and mov.cliente_destino:
                    # 1. CLIENTE ORIGEN (RESTA)
                    key_origen = f"{mov.cliente_origen.id}_{detalle.producto.id}"
                    
                    if key_origen not in entregas_dict:
                        entregas_dict[key_origen] = {
                            'cliente_id': mov.cliente_origen.id,
                            'cliente_nombre': mov.cliente_origen.nombre,
                            'cliente_codigo': mov.cliente_origen.codigo,
                            'cliente_direccion': mov.cliente_origen.direccion or '-',
                            'producto_id': detalle.producto.id,
                            'producto_codigo': detalle.producto.codigo,
                            'producto_nombre': detalle.producto.nombre,
                            'producto_categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                            'producto_unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                            'total_entregas': set(),
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),    # ✅ NUEVO
                            'cantidad_traslado_destino': Decimal('0'),   # ✅ NUEVO
                            'cantidad_buena': Decimal('0'),
                            'cantidad_danada': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
                    
                    entregas_dict[key_origen]['total_entregas'].add(mov.id)
                    entregas_dict[key_origen]['cantidad_traslado_origen'] += cant_total  # ✅ NUEVO
                    entregas_dict[key_origen]['stock_bueno'] -= cant_b  # RESTA
                    entregas_dict[key_origen]['stock_danado'] -= cant_d  # RESTA
                    entregas_dict[key_origen]['cantidad_buena'] += cant_b
                    entregas_dict[key_origen]['cantidad_danada'] += cant_d
                    
                    # 2. CLIENTE DESTINO (SUMA)
                    key_destino = f"{mov.cliente_destino.id}_{detalle.producto.id}"
                    
                    if key_destino not in entregas_dict:
                        entregas_dict[key_destino] = {
                            'cliente_id': mov.cliente_destino.id,
                            'cliente_nombre': mov.cliente_destino.nombre,
                            'cliente_codigo': mov.cliente_destino.codigo,
                            'cliente_direccion': mov.cliente_destino.direccion or '-',
                            'producto_id': detalle.producto.id,
                            'producto_codigo': detalle.producto.codigo,
                            'producto_nombre': detalle.producto.nombre,
                            'producto_categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                            'producto_unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                            'total_entregas': set(),
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),    # ✅ NUEVO
                            'cantidad_traslado_destino': Decimal('0'),   # ✅ NUEVO
                            'cantidad_buena': Decimal('0'),
                            'cantidad_danada': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
                    
                    entregas_dict[key_destino]['total_entregas'].add(mov.id)
                    entregas_dict[key_destino]['cantidad_traslado_destino'] += cant_total  # ✅ NUEVO
                    entregas_dict[key_destino]['stock_bueno'] += cant_b  # SUMA
                    entregas_dict[key_destino]['stock_danado'] += cant_d  # SUMA
                    entregas_dict[key_destino]['cantidad_buena'] += cant_b
                    entregas_dict[key_destino]['cantidad_danada'] += cant_d
                    
                else:
                    # PARA ENTRADA Y SALIDA (lógica original)
                    cli_id = mov.cliente.id
                    key = f"{cli_id}_{detalle.producto.id}"
                    
                    if not mostrar_todos and key not in entregas_dict:
                        entregas_dict[key] = {
                            'cliente_id': cli_id,
                            'cliente_nombre': mov.cliente.nombre,
                            'cliente_codigo': mov.cliente.codigo,
                            'cliente_direccion': mov.cliente.direccion or '-',
                            'producto_id': detalle.producto.id,
                            'producto_codigo': detalle.producto.codigo,
                            'producto_nombre': detalle.producto.nombre,
                            'producto_categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                            'producto_unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                            'total_entregas': set(),
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),    # ✅ NUEVO
                            'cantidad_traslado_destino': Decimal('0'),   # ✅ NUEVO
                            'cantidad_buena': Decimal('0'),
                            'cantidad_danada': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
                    
                    if key in entregas_dict:
                        entregas_dict[key]['total_entregas'].add(mov.id)
                        
                        if mov.tipo == 'ENTRADA':
                            entregas_dict[key]['cantidad_entrada'] += cant_total
                            entregas_dict[key]['stock_bueno'] += cant_b
                            entregas_dict[key]['stock_danado'] += cant_d
                            
                        elif mov.tipo == 'SALIDA':
                            entregas_dict[key]['cantidad_salida'] += cant_total
                            entregas_dict[key]['stock_bueno'] -= cant_b
                            entregas_dict[key]['stock_danado'] -= cant_d

                        entregas_dict[key]['cantidad_buena'] += cant_b
                        entregas_dict[key]['cantidad_danada'] += cant_d
            
            # Convertir a lista y calcular totales
            cantidad_total_global = Decimal('0')
            clientes_unicos = set()
            productos_unicos = set()
            
            for key, item in entregas_dict.items():
                # ✅ CORRECCIÓN: Si hay filtro de cliente, mostrar SOLO ese cliente
                if cliente_id and str(item['cliente_id']) != str(cliente_id):
                    continue
                
                item['total_entregas'] = len(item['total_entregas'])
                item['stock_total'] = item['stock_bueno'] + item['stock_danado']
                cantidad_total_global += item['stock_total']
                
                clientes_unicos.add(item['cliente_id'])
                productos_unicos.add(item['producto_id'])
                entregas.append(item)
            
            entregas.sort(key=lambda x: (x['cliente_codigo'], x['producto_codigo']))

            estadisticas = {
                'cantidad_total': cantidad_total_global,
                'total_clientes_unicos': len(clientes_unicos),
                'total_productos_unicos': len(productos_unicos),
            }
        
        elif vista == 'por_cliente':
            entregas_qs = DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).select_related(
                'movimiento', 'movimiento__cliente', 'producto'
            )
            
            if categoria_id:
                entregas_qs = entregas_qs.filter(producto__categoria_id=categoria_id)
            if producto_id:
                entregas_qs = entregas_qs.filter(producto_id=producto_id)
            
            # Agrupar por cliente
            clientes_dict = {}
            for detalle in entregas_qs:
                mov = detalle.movimiento
                cant_b = detalle.cantidad or Decimal('0')
                cant_d = detalle.cantidad_danada or Decimal('0')
                
                # ✅ PARA TRASLADOS: Procesar ORIGEN Y DESTINO
                if mov.tipo == 'TRASLADO' and mov.cliente_origen and mov.cliente_destino:
                    # CLIENTE ORIGEN (RESTA)
                    cliente_id_origen = mov.cliente_origen.id
                    
                    if cliente_id_origen not in clientes_dict:
                        clientes_dict[cliente_id_origen] = {
                            'cliente_id': cliente_id_origen,
                            'cliente_nombre': mov.cliente_origen.nombre,
                            'cliente_codigo': mov.cliente_origen.codigo,
                            'cliente_direccion': mov.cliente_origen.direccion or '-',
                            'cliente_telefono': mov.cliente_origen.telefono or '-',
                            'movimientos': set(),
                            'productos': set(),
                            'cantidad_buena': Decimal('0'),
                            'cantidad_danada': Decimal('0'),
                        }
                    
                    clientes_dict[cliente_id_origen]['movimientos'].add(mov.id)
                    clientes_dict[cliente_id_origen]['productos'].add(detalle.producto.id)
                    clientes_dict[cliente_id_origen]['cantidad_buena'] -= cant_b
                    clientes_dict[cliente_id_origen]['cantidad_danada'] -= cant_d
                    
                    # CLIENTE DESTINO (SUMA)
                    cliente_id_destino = mov.cliente_destino.id
                    
                    if cliente_id_destino not in clientes_dict:
                        clientes_dict[cliente_id_destino] = {
                            'cliente_id': cliente_id_destino,
                            'cliente_nombre': mov.cliente_destino.nombre,
                            'cliente_codigo': mov.cliente_destino.codigo,
                            'cliente_direccion': mov.cliente_destino.direccion or '-',
                            'cliente_telefono': mov.cliente_destino.telefono or '-',
                            'movimientos': set(),
                            'productos': set(),
                            'cantidad_buena': Decimal('0'),
                            'cantidad_danada': Decimal('0'),
                        }
                    
                    clientes_dict[cliente_id_destino]['movimientos'].add(mov.id)
                    clientes_dict[cliente_id_destino]['productos'].add(detalle.producto.id)
                    clientes_dict[cliente_id_destino]['cantidad_buena'] += cant_b
                    clientes_dict[cliente_id_destino]['cantidad_danada'] += cant_d
                    
                else:
                    # ENTRADA Y SALIDA (lógica original)
                    cliente_id_key = mov.cliente.id
                    
                    if cliente_id_key not in clientes_dict:
                        clientes_dict[cliente_id_key] = {
                            'cliente_id': cliente_id_key,
                            'cliente_nombre': mov.cliente.nombre,
                            'cliente_codigo': mov.cliente.codigo,
                            'cliente_direccion': mov.cliente.direccion or '-',
                            'cliente_telefono': mov.cliente.telefono or '-',
                            'movimientos': set(),
                            'productos': set(),
                            'cantidad_buena': Decimal('0'),
                            'cantidad_danada': Decimal('0'),
                        }
                    
                    clientes_dict[cliente_id_key]['movimientos'].add(mov.id)
                    clientes_dict[cliente_id_key]['productos'].add(detalle.producto.id)
                    
                    if mov.tipo == 'ENTRADA':
                        clientes_dict[cliente_id_key]['cantidad_buena'] += cant_b
                        clientes_dict[cliente_id_key]['cantidad_danada'] += cant_d
                        
                    elif mov.tipo == 'SALIDA':
                        clientes_dict[cliente_id_key]['cantidad_buena'] -= cant_b
                        clientes_dict[cliente_id_key]['cantidad_danada'] -= cant_d

            # Convertir a lista
            total_movimientos = 0
            total_general = Decimal('0')
            total_productos_diferentes = set()
            
            for cliente_id_key, item in clientes_dict.items():
                # ✅ CORRECCIÓN: Si hay filtro de cliente, mostrar SOLO ese cliente
                if cliente_id and str(cliente_id_key) != str(cliente_id):
                    continue
                
                item['total_entregas'] = len(item['movimientos'])
                item['total_productos'] = len(item['productos'])
                item['cantidad_total'] = item['cantidad_buena'] + item['cantidad_danada']
                
                total_movimientos += item['total_entregas']
                total_general += item['cantidad_total']
                
                total_productos_diferentes.update(item['productos'])
                
                del item['movimientos']
                del item['productos']
                entregas.append(item)
            
            entregas.sort(key=lambda x: x['cliente_codigo'])
            
            estadisticas = {
                'total_movimientos': total_movimientos,
                'total_productos_diferentes': len(total_productos_diferentes),
                'cantidad_total': total_general,
            }

        # ========================================
        # CORRECCIÓN VISTA PRODUCTOS_TOP
        # ========================================

        else:  # productos_top
            entregas_qs = DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).select_related(
                'movimiento', 'movimiento__cliente', 'producto',
                'producto__categoria', 'producto__unidad_medida'
            )
            
            if categoria_id:
                entregas_qs = entregas_qs.filter(producto__categoria_id=categoria_id)
            if producto_id:
                entregas_qs = entregas_qs.filter(producto_id=producto_id)
            
            # Agrupar por producto
            productos_dict = {}
            clientes_unicos = set()
            
            for detalle in entregas_qs:
                mov = detalle.movimiento
                producto_id_key = detalle.producto.id
                cant_b = detalle.cantidad or Decimal('0')
                cant_d = detalle.cantidad_danada or Decimal('0')
                
                if producto_id_key not in productos_dict:
                    productos_dict[producto_id_key] = {
                        'producto_id': producto_id_key,
                        'producto_codigo': detalle.producto.codigo,
                        'producto_nombre': detalle.producto.nombre,
                        'producto_categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                        'producto_unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                        'clientes': set(),
                        'movimientos': set(),
                        'cantidad_buena': Decimal('0'),
                        'cantidad_danada': Decimal('0'),
                    }
                
                productos_dict[producto_id_key]['movimientos'].add(mov.id)
                
                # ✅ PARA TRASLADOS: Registrar ambos clientes pero stock NEUTRO global
                if mov.tipo == 'TRASLADO' and mov.cliente_origen and mov.cliente_destino:
                    productos_dict[producto_id_key]['clientes'].add(mov.cliente_origen.id)
                    productos_dict[producto_id_key]['clientes'].add(mov.cliente_destino.id)
                    clientes_unicos.add(mov.cliente_origen.id)
                    clientes_unicos.add(mov.cliente_destino.id)
                    # Stock neto CERO (sale de uno, entra a otro)
                    
                else:
                    productos_dict[producto_id_key]['clientes'].add(mov.cliente.id)
                    clientes_unicos.add(mov.cliente.id)
                    
                    if mov.tipo == 'ENTRADA':
                        productos_dict[producto_id_key]['cantidad_buena'] += cant_b
                        productos_dict[producto_id_key]['cantidad_danada'] += cant_d
                    elif mov.tipo == 'SALIDA':
                        productos_dict[producto_id_key]['cantidad_buena'] -= cant_b
                        productos_dict[producto_id_key]['cantidad_danada'] -= cant_d

            # Convertir a lista
            total_general = Decimal('0')
            total_entregas_global = 0
            
            for producto_id_key, item in productos_dict.items():
                item['total_clientes'] = len(item['clientes'])
                item['total_entregas'] = len(item['movimientos'])
                item['cantidad_total'] = item['cantidad_buena'] + item['cantidad_danada']
                
                total_general += item['cantidad_total']
                total_entregas_global += item['total_entregas']
                
                del item['clientes']
                del item['movimientos']
                entregas.append(item)
            
            entregas.sort(key=lambda x: x['producto_codigo'])
            
            estadisticas = {
                'total_clientes': len(clientes_unicos),
                'total_entregas': total_entregas_global,
                'cantidad_total': total_general,
            }
        
        # --- SIDEBAR (Top 10 Productos y Clientes) ---
        # Recalculamos usando la misma lógica para que coincida con la tabla principal
        
        # Top 10 Productos (Sidebar)
        productos_top_qs = DetalleMovimientoCliente.objects.filter(
            movimiento__in=movimientos_qs
        ).select_related('producto', 'producto__unidad_medida', 'movimiento__cliente')

        productos_top_dict = {}
        for detalle in productos_top_qs:
            mov = detalle.movimiento
            prod_id = detalle.producto.id
            cant_b = detalle.cantidad or Decimal('0')
            cant_d = detalle.cantidad_danada or Decimal('0')
            
            if prod_id not in productos_top_dict:
                productos_top_dict[prod_id] = {
                    'producto__codigo': detalle.producto.codigo,
                    'producto__nombre': detalle.producto.nombre,
                    'producto__unidad_medida__abreviatura': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                    'clientes': set(),
                    'cantidad_buena': Decimal('0'),
                    'cantidad_danada': Decimal('0'),
                }
            
            # ✅ PARA TRASLADOS: Registrar ambos clientes pero stock NEUTRO
            if mov.tipo == 'TRASLADO' and mov.cliente_origen and mov.cliente_destino:
                productos_top_dict[prod_id]['clientes'].add(mov.cliente_origen.id)
                productos_top_dict[prod_id]['clientes'].add(mov.cliente_destino.id)
                # Stock neto CERO
            else:
                productos_top_dict[prod_id]['clientes'].add(mov.cliente.id)
                
                if mov.tipo == 'ENTRADA':
                    productos_top_dict[prod_id]['cantidad_buena'] += cant_b
                    productos_top_dict[prod_id]['cantidad_danada'] += cant_d
                elif mov.tipo == 'SALIDA':
                    productos_top_dict[prod_id]['cantidad_buena'] -= cant_b
                    productos_top_dict[prod_id]['cantidad_danada'] -= cant_d

        for prod_id, item in productos_top_dict.items():
            item['total_clientes'] = len(item['clientes'])
            item['cantidad_total'] = item['cantidad_buena'] + item['cantidad_danada']
            del item['clientes']
            productos_top.append(item)

        productos_top.sort(key=lambda x: x['cantidad_total'], reverse=True)
        productos_top = productos_top[:10]

        # ========================================
        # CORRECCIÓN SIDEBAR - Top 10 Clientes
        # ========================================

        clientes_top_qs = DetalleMovimientoCliente.objects.filter(
            movimiento__in=movimientos_qs
        ).select_related('movimiento', 'movimiento__cliente', 'producto')

        clientes_top_dict = {}
        for detalle in clientes_top_qs:
            mov = detalle.movimiento
            cant_b = detalle.cantidad or Decimal('0')
            cant_d = detalle.cantidad_danada or Decimal('0')
            
            # ✅ PARA TRASLADOS: Procesar ORIGEN Y DESTINO
            if mov.tipo == 'TRASLADO' and mov.cliente_origen and mov.cliente_destino:
                # CLIENTE ORIGEN (RESTA)
                cli_id_origen = mov.cliente_origen.id
                if cli_id_origen not in clientes_top_dict:
                    clientes_top_dict[cli_id_origen] = {
                        'movimiento__cliente__nombre': mov.cliente_origen.nombre,
                        'movimientos': set(),
                        'productos': set(),
                        'cantidad_buena': Decimal('0'),
                        'cantidad_danada': Decimal('0'),
                    }
                clientes_top_dict[cli_id_origen]['movimientos'].add(mov.id)
                clientes_top_dict[cli_id_origen]['productos'].add(detalle.producto.id)
                clientes_top_dict[cli_id_origen]['cantidad_buena'] -= cant_b
                clientes_top_dict[cli_id_origen]['cantidad_danada'] -= cant_d
                
                # CLIENTE DESTINO (SUMA)
                cli_id_destino = mov.cliente_destino.id
                if cli_id_destino not in clientes_top_dict:
                    clientes_top_dict[cli_id_destino] = {
                        'movimiento__cliente__nombre': mov.cliente_destino.nombre,
                        'movimientos': set(),
                        'productos': set(),
                        'cantidad_buena': Decimal('0'),
                        'cantidad_danada': Decimal('0'),
                    }
                clientes_top_dict[cli_id_destino]['movimientos'].add(mov.id)
                clientes_top_dict[cli_id_destino]['productos'].add(detalle.producto.id)
                clientes_top_dict[cli_id_destino]['cantidad_buena'] += cant_b
                clientes_top_dict[cli_id_destino]['cantidad_danada'] += cant_d
                
            else:
                # ENTRADA Y SALIDA
                cli_id = mov.cliente.id
                if cli_id not in clientes_top_dict:
                    clientes_top_dict[cli_id] = {
                        'movimiento__cliente__nombre': mov.cliente.nombre,
                        'movimientos': set(),
                        'productos': set(),
                        'cantidad_buena': Decimal('0'),
                        'cantidad_danada': Decimal('0'),
                    }
                clientes_top_dict[cli_id]['movimientos'].add(mov.id)
                clientes_top_dict[cli_id]['productos'].add(detalle.producto.id)
                
                if mov.tipo == 'ENTRADA':
                    clientes_top_dict[cli_id]['cantidad_buena'] += cant_b
                    clientes_top_dict[cli_id]['cantidad_danada'] += cant_d
                elif mov.tipo == 'SALIDA':
                    clientes_top_dict[cli_id]['cantidad_buena'] -= cant_b
                    clientes_top_dict[cli_id]['cantidad_danada'] -= cant_d

        for cli_id, item in clientes_top_dict.items():
            item['total_entregas'] = len(item['movimientos'])
            item['total_productos'] = len(item['productos'])
            item['cantidad_total'] = item['cantidad_buena'] + item['cantidad_danada']
            del item['movimientos']
            del item['productos']
            resumen_clientes.append(item)

        resumen_clientes.sort(key=lambda x: x['cantidad_total'], reverse=True)
        resumen_clientes = resumen_clientes[:10]
        
        # ✅ APLICAR PAGINACIÓN A LA LISTA 'entregas'
        # Calcular total antes de recortar la lista
        total_entregas = len(entregas)
        
        paginator = Paginator(entregas, items_por_pagina)
        
        try:
            entregas_paginados = paginator.page(page)
        except PageNotAnInteger:
            entregas_paginados = paginator.page(1)
        except EmptyPage:
            entregas_paginados = paginator.page(paginator.num_pages)

        context = {
            **self.admin_site.each_context(request),
            'title': _('Reportes de ENTRADA, SALIDA Y TRASPASO de CLIENTES'),
            'entregas': entregas,
            'estadisticas': estadisticas,
            'productos_top': productos_top,
            'resumen_clientes': resumen_clientes,
            'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
            'categorias': Categoria.objects.all(),
            'productos': Producto.objects.filter(activo=True).order_by('codigo'),
            'filtros': {
                'vista': vista,
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'cliente': cliente_id,
                'categoria': categoria_id,
                'producto': producto_id,
                'mostrar_todos': request.GET.get('mostrar_todos', ''),
                'items_por_pagina': items_por_pagina,
            },
            'page_obj': entregas_paginados,
            'paginator': paginator,
            'items_por_pagina': items_por_pagina,
            'opciones_items': [50, 100, 200, 500],
            'opts': self.model._meta,
        }
        
        if extra_context:
            context.update(extra_context)
        
        return render(request, self.change_list_template, context)
    
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('obtener-detalle-entrega/',
                self.admin_site.admin_view(views.obtener_detalle_entrega_cliente),
                name='reportes_entregas_obtener_detalle'),
            path('obtener-productos-cliente/',
                self.admin_site.admin_view(views.obtener_productos_cliente),
                name='reportes_entregas_productos_cliente'),
            path('obtener-clientes-producto/',
                self.admin_site.admin_view(views.obtener_clientes_producto),
                name='reportes_entregas_clientes_producto'),
            path('obtener-detalle-estadistica-entregas/',
                self.admin_site.admin_view(views.obtener_detalle_estadistica_entregas),
                name='reportes_entregas_obtener_estadistica'),
            path('exportar-excel/', 
                self.admin_site.admin_view(self.exportar_excel), 
                name='reportes_entregas_exportar_excel'),
            path('exportar-csv/', 
                self.admin_site.admin_view(self.exportar_csv), 
                name='reportes_entregas_exportar_csv'),
            path('obtener-productos-cliente/',
                 self.admin_site.admin_view(views.obtener_productos_cliente),
                 name='reportes_entregas_obtener_productos'), 
        ]
        return custom_urls + urls
    
    def get_queryset(self, request):
        """Retorna queryset vacío"""
        return self.model.objects.none()
    
    def exportar_excel(self, request):
        """Exporta las entregas a Excel"""
        from django.http import HttpResponse
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        
        # Obtener parámetros de filtro
        vista = request.GET.get('vista', 'detallado')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        cliente_id = request.GET.get('cliente', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        mostrar_todos = request.GET.get('mostrar_todos', '') == '1'
        
        # Convertir fechas
        fecha_inicio_obj = None
        fecha_fin_obj = None
        
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Query base
        movimientos_qs = MovimientoCliente.objects.all()
        
        if fecha_inicio_obj:
            movimientos_qs = movimientos_qs.filter(fecha__gte=fecha_inicio_obj)
        
        if fecha_fin_obj:
            movimientos_qs = movimientos_qs.filter(fecha__lte=fecha_fin_obj)
        
        if cliente_id:
            movimientos_qs = movimientos_qs.filter(
                Q(cliente_id=cliente_id) | 
                Q(cliente_origen_id=cliente_id) | 
                Q(cliente_destino_id=cliente_id)
            )
        
        # Obtener entregas según vista
        entregas = []
        
        if vista == 'detallado':
            entregas_dict = {}
            
            if mostrar_todos:
                clientes_activos = Cliente.objects.filter(activo=True)
                if cliente_id:
                    clientes_activos = clientes_activos.filter(id=cliente_id)
                
                productos_activos = Producto.objects.filter(activo=True)
                if categoria_id:
                    productos_activos = productos_activos.filter(categoria_id=categoria_id)
                if producto_id:
                    productos_activos = productos_activos.filter(id=producto_id)
                
                for cliente in clientes_activos:
                    for producto in productos_activos:
                        key = f"{cliente.id}_{producto.id}"
                        entregas_dict[key] = {
                            'cliente_nombre': cliente.nombre,
                            'cliente_codigo': cliente.codigo,
                            'producto_codigo': producto.codigo,
                            'producto_nombre': producto.nombre,
                            'producto_categoria': producto.categoria.nombre if producto.categoria else '-',
                            'producto_unidad': producto.unidad_medida.abreviatura if producto.unidad_medida else 'UND',
                            'total_entregas': 0,
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),
                            'cantidad_traslado_destino': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
            
            entregas_qs = DetalleMovimientoCliente.objects.filter(
                movimiento__in=movimientos_qs
            ).select_related(
                'movimiento', 'movimiento__cliente', 'producto',
                'producto__categoria', 'producto__unidad_medida'
            )
            
            if categoria_id:
                entregas_qs = entregas_qs.filter(producto__categoria_id=categoria_id)
            if producto_id:
                entregas_qs = entregas_qs.filter(producto_id=producto_id)
            
            movimientos_procesados = {}
            
            for detalle in entregas_qs:
                mov = detalle.movimiento
                cant_b = detalle.cantidad or Decimal('0')
                cant_d = detalle.cantidad_danada or Decimal('0')
                cant_total = cant_b + cant_d
                
                if mov.tipo == 'TRASLADO' and mov.cliente_origen and mov.cliente_destino:
                    # Cliente origen
                    key_origen = f"{mov.cliente_origen.id}_{detalle.producto.id}"
                    
                    if key_origen not in entregas_dict:
                        entregas_dict[key_origen] = {
                            'cliente_nombre': mov.cliente_origen.nombre,
                            'cliente_codigo': mov.cliente_origen.codigo,
                            'producto_codigo': detalle.producto.codigo,
                            'producto_nombre': detalle.producto.nombre,
                            'producto_categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                            'producto_unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                            'total_entregas': 0,
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),
                            'cantidad_traslado_destino': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
                    
                    mov_key_origen = f"{key_origen}_{mov.id}"
                    if mov_key_origen not in movimientos_procesados:
                        entregas_dict[key_origen]['total_entregas'] += 1
                        movimientos_procesados[mov_key_origen] = True
                    
                    entregas_dict[key_origen]['cantidad_traslado_origen'] += cant_total
                    entregas_dict[key_origen]['stock_bueno'] -= cant_b
                    entregas_dict[key_origen]['stock_danado'] -= cant_d
                    
                    # Cliente destino
                    key_destino = f"{mov.cliente_destino.id}_{detalle.producto.id}"
                    
                    if key_destino not in entregas_dict:
                        entregas_dict[key_destino] = {
                            'cliente_nombre': mov.cliente_destino.nombre,
                            'cliente_codigo': mov.cliente_destino.codigo,
                            'producto_codigo': detalle.producto.codigo,
                            'producto_nombre': detalle.producto.nombre,
                            'producto_categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                            'producto_unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                            'total_entregas': 0,
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),
                            'cantidad_traslado_destino': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
                    
                    mov_key_destino = f"{key_destino}_{mov.id}"
                    if mov_key_destino not in movimientos_procesados:
                        entregas_dict[key_destino]['total_entregas'] += 1
                        movimientos_procesados[mov_key_destino] = True
                    
                    entregas_dict[key_destino]['cantidad_traslado_destino'] += cant_total
                    entregas_dict[key_destino]['stock_bueno'] += cant_b
                    entregas_dict[key_destino]['stock_danado'] += cant_d
                    
                else:
                    cli_id = mov.cliente.id
                    key = f"{cli_id}_{detalle.producto.id}"
                    
                    if not mostrar_todos and key not in entregas_dict:
                        entregas_dict[key] = {
                            'cliente_nombre': mov.cliente.nombre,
                            'cliente_codigo': mov.cliente.codigo,
                            'producto_codigo': detalle.producto.codigo,
                            'producto_nombre': detalle.producto.nombre,
                            'producto_categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else '-',
                            'producto_unidad': detalle.producto.unidad_medida.abreviatura if detalle.producto.unidad_medida else 'UND',
                            'total_entregas': 0,
                            'cantidad_entrada': Decimal('0'),
                            'cantidad_salida': Decimal('0'),
                            'cantidad_traslado_origen': Decimal('0'),
                            'cantidad_traslado_destino': Decimal('0'),
                            'stock_bueno': Decimal('0'),
                            'stock_danado': Decimal('0'),
                        }
                    
                    if key in entregas_dict:
                        mov_key = f"{key}_{mov.id}"
                        if mov_key not in movimientos_procesados:
                            entregas_dict[key]['total_entregas'] += 1
                            movimientos_procesados[mov_key] = True
                        
                        if mov.tipo == 'ENTRADA':
                            entregas_dict[key]['cantidad_entrada'] += cant_total
                            entregas_dict[key]['stock_bueno'] += cant_b
                            entregas_dict[key]['stock_danado'] += cant_d
                        elif mov.tipo == 'SALIDA':
                            entregas_dict[key]['cantidad_salida'] += cant_total
                            entregas_dict[key]['stock_bueno'] -= cant_b
                            entregas_dict[key]['stock_danado'] -= cant_d
            
            for key, item in entregas_dict.items():
                if cliente_id and str(item.get('cliente_id', '')) != str(cliente_id):
                    continue
                item['stock_total'] = item['stock_bueno'] + item['stock_danado']
                entregas.append(item)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Entregas"
        
        # Estilos
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        if vista == 'detallado':
            headers = ['Cliente', 'Cód. Cliente', 'Cód. Producto', 'Producto', 'Categoría', 
                    'Unidad', 'Entregas', 'Entrada', 'Salida', 'Trasl. Origen', 
                    'Trasl. Destino', 'Stock Bueno', 'Stock Dañado', 'Stock Total']
        elif vista == 'por_cliente':
            headers = ['Cliente', 'Código', 'Dirección', 'Teléfono', 'Entregas', 
                    'Productos', 'Cantidad Total']
        else:
            headers = ['#', 'Código', 'Producto', 'Categoría', 'Unidad', 'Clientes', 
                    'Entregas', 'Cantidad Total']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for row_idx, item in enumerate(entregas, start=2):
            if vista == 'detallado':
                row_data = [
                    item.get('cliente_nombre', ''),
                    item.get('cliente_codigo', ''),
                    item.get('producto_codigo', ''),
                    item.get('producto_nombre', ''),
                    item.get('producto_categoria', '-'),
                    item.get('producto_unidad', 'UND'),
                    item.get('total_entregas', 0),
                    float(item.get('cantidad_entrada', 0)),
                    float(item.get('cantidad_salida', 0)),
                    float(item.get('cantidad_traslado_origen', 0)),
                    float(item.get('cantidad_traslado_destino', 0)),
                    float(item.get('stock_bueno', 0)),
                    float(item.get('stock_danado', 0)),
                    float(item.get('stock_total', 0)),
                ]
            else:
                row_data = list(item.values())
            
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'entregas_{vista}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response


    def exportar_csv(self, request):
        """Exporta las entregas a CSV"""
        from django.http import HttpResponse
        from datetime import datetime
        import csv
        from decimal import Decimal
        
        # Obtener parámetros de filtro (mismo código que exportar_excel)
        vista = request.GET.get('vista', 'detallado')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        cliente_id = request.GET.get('cliente', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        mostrar_todos = request.GET.get('mostrar_todos', '') == '1'
        
        # [Aquí va el mismo código de procesamiento que en exportar_excel]
        # ... (copiar todo el código de procesamiento)
        
        # Respuesta HTTP
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'entregas_{vista}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        # BOM para Excel
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        
        # Encabezados (mismos que en Excel)
        if vista == 'detallado':
            headers = ['Cliente', 'Cód. Cliente', 'Cód. Producto', 'Producto', 'Categoría', 
                    'Unidad', 'Entregas', 'Entrada', 'Salida', 'Trasl. Origen', 
                    'Trasl. Destino', 'Stock Bueno', 'Stock Dañado', 'Stock Total']
        elif vista == 'por_cliente':
            headers = ['Cliente', 'Código', 'Dirección', 'Teléfono', 'Entregas', 
                    'Productos', 'Cantidad Total']
        else:
            headers = ['#', 'Código', 'Producto', 'Categoría', 'Unidad', 'Clientes', 
                    'Entregas', 'Cantidad Total']
        
        writer.writerow(headers)
        
        # Datos (mismo formato que Excel)
        for item in entregas:
            if vista == 'detallado':
                row_data = [
                    item.get('cliente_nombre', ''),
                    item.get('cliente_codigo', ''),
                    item.get('producto_codigo', ''),
                    item.get('producto_nombre', ''),
                    item.get('producto_categoria', '-'),
                    item.get('producto_unidad', 'UND'),
                    item.get('total_entregas', 0),
                    float(item.get('cantidad_entrada', 0)),
                    float(item.get('cantidad_salida', 0)),
                    float(item.get('cantidad_traslado_origen', 0)),
                    float(item.get('cantidad_traslado_destino', 0)),
                    float(item.get('stock_bueno', 0)),
                    float(item.get('stock_danado', 0)),
                    float(item.get('stock_total', 0)),
                ]
            else:
                row_data = list(item.values())
            
            writer.writerow(row_data)
        
        return response


# ==========================================
# REPORTE DE STOCK
# ==========================================
class ReporteStockAdmin(admin.ModelAdmin):
    """
    Admin para reportes de stock actual - ACTUALIZADO para usar cálculo dinámico
    """
    
    change_list_template = 'admin/reportes/reporte_stock_list.html'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False
    
    def has_change_permission(self, request, obj=None):
        return True
    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        
        # Obtener filtros
        vista = request.GET.get('vista', 'detallado')
        almacen_id = request.GET.get('almacen', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        stock_minimo = request.GET.get('stock_minimo', '')
        solo_con_stock = request.GET.get('solo_con_stock', '')
        
        # ✅ LÓGICA DE PAGINACIÓN UNIFICADA
        page = request.GET.get('page', 1)
        items_por_pagina = request.GET.get('items_por_pagina', '100')
        
        try:
            items_por_pagina = int(items_por_pagina)
            if items_por_pagina not in [50, 100, 200, 500]:
                items_por_pagina = 100
        except (ValueError, TypeError):
            items_por_pagina = 100
        
        # Obtener stocks usando el nuevo sistema
        stocks = []
        productos_bajo_stock = []
        resumen_almacenes = []
        
        # ⭐ CALCULAR ESTADÍSTICAS GLOBALES PRIMERO
        total_productos_sistema = Producto.objects.filter(activo=True).count()
        total_almacenes_activos = Almacen.objects.filter(activo=True).count()
        
        # Calcular stock total bueno y dañado
        stock_bueno_total = 0
        stock_danado_total = 0
        productos_bajo_minimo_count = 0
        
        # Obtener almacenes
        almacenes = Almacen.objects.filter(activo=True)
        if almacen_id:
            almacenes = almacenes.filter(id=almacen_id)
        
        # Obtener productos
        productos = Producto.objects.filter(activo=True)
        if categoria_id:
            productos = productos.filter(categoria_id=categoria_id)
        if producto_id:
            productos = productos.filter(id=producto_id)
        
        # ⭐ CALCULAR STOCKS GLOBALES (sin filtros de vista)
        for almacen_calc in Almacen.objects.filter(activo=True):
            for producto_calc in Producto.objects.filter(activo=True):
                stock_data = almacen_calc.get_stock_producto(producto_calc)
                stock_bueno_total += stock_data['stock_bueno']
                stock_danado_total += stock_data['stock_danado']
                
                # Verificar productos bajo mínimo
                if hasattr(producto_calc, 'stock_minimo') and producto_calc.stock_minimo and producto_calc.stock_minimo > 0:
                    if stock_data['stock_bueno'] <= producto_calc.stock_minimo:
                        productos_bajo_minimo_count += 1
        
        if vista == 'detallado':
            # Vista detallada: calcular stock para cada almacén y producto
            for almacen in almacenes:
                for producto in productos:
                    stock_data = almacen.get_stock_producto(producto)
                    
                    # Aplicar filtros
                    if solo_con_stock and stock_data['stock_total'] == 0:
                        continue
                    if stock_minimo and stock_data['stock_bueno'] > producto.stock_minimo:
                        continue
                    
                    stocks.append({
                        'almacen': almacen,
                        'producto': producto,
                        'stock_bueno': stock_data['stock_bueno'],
                        'stock_danado': stock_data['stock_danado'],
                        'stock_total': stock_data['stock_total'],
                        'entradas_total': stock_data['entradas_total'],
                        'salidas_total': stock_data['salidas_total'],
                        'traslados_netos': stock_data['traslados_netos_total']
                    })
            
        
        elif vista == 'por_almacen':
            for almacen in almacenes:
                stocks_almacen = almacen.get_todos_los_stocks()
                
                total_productos = len(stocks_almacen)
                stock_buena_total = sum(s['stock_bueno'] for s in stocks_almacen.values())
                stock_danada_total = sum(s['stock_danado'] for s in stocks_almacen.values())
                stock_total = stock_buena_total + stock_danada_total
                
                # ✅ Incluir stocks negativos
                if total_productos > 0 or stock_total != 0 or not solo_con_stock:
                    stocks.append({
                        'almacen': almacen,
                        'total_productos': total_productos,
                        'stock_buena_total': stock_buena_total,
                        'stock_danada_total': stock_danada_total,
                        'stock_total': stock_total
                    })
                
        else:  # por_producto
            for producto in productos:
                stock_buena_total = 0
                stock_danada_total = 0
                total_almacenes = 0
                
                for almacen in Almacen.objects.filter(activo=True):
                    stock_data = almacen.get_stock_producto(producto)
                    
                    # ✅ Contar almacenes con stock != 0 (incluye negativos)
                    if stock_data['stock_total'] != 0:
                        stock_buena_total += stock_data['stock_bueno']
                        stock_danada_total += stock_data['stock_danado']
                        total_almacenes += 1
                
                stock_total_producto = stock_buena_total + stock_danada_total
                if total_almacenes > 0 or stock_total_producto != 0 or not solo_con_stock:
                    stocks.append({
                        'producto': producto,
                        'total_almacenes': total_almacenes,
                        'stock_buena_total': stock_buena_total,
                        'stock_danada_total': stock_danada_total,
                        'stock_total': stock_total_producto
                    })
        
        # Productos bajo stock mínimo
        for almacen in Almacen.objects.filter(activo=True):
            for producto in Producto.objects.filter(activo=True, stock_minimo__gt=0):
                stock_data = almacen.get_stock_producto(producto)
                if stock_data['stock_bueno'] <= producto.stock_minimo:
                    productos_bajo_stock.append({
                        'almacen': almacen,
                        'producto': producto,
                        'stock_actual': stock_data['stock_bueno'],
                        'stock_minimo': producto.stock_minimo,
                        'diferencia': producto.stock_minimo - stock_data['stock_bueno']
                    })
        
        # Ordenar por diferencia
        productos_bajo_stock = sorted(productos_bajo_stock, key=lambda x: x['diferencia'], reverse=True)[:10]
        
        # Resumen por almacén para sidebar
        for almacen in Almacen.objects.filter(activo=True):
            stocks_almacen = almacen.get_todos_los_stocks()
            total_productos = len(stocks_almacen)
            stock_total = sum(s['stock_total'] for s in stocks_almacen.values())
            
            if total_productos > 0:
                resumen_almacenes.append({
                    'almacen': almacen.nombre,
                    'total_productos': total_productos,
                    'stock_total': stock_total
                })
        
        # Valoración total
        total_items = stock_bueno_total + stock_danado_total
        
        valoracion = {
            'total_productos': total_productos_sistema,
            'total_items': total_items,
            'valor_total': 0,  # Calcular si tienes precios
        }
        
        # ⭐ AGREGAR ESTADÍSTICAS AL CONTEXTO
        estadisticas = {
            'total_productos': total_productos_sistema,
            'stock_bueno_total': stock_bueno_total,
            'stock_danado_total': stock_danado_total,
            'total_almacenes': total_almacenes_activos,
            'productos_bajo_minimo': productos_bajo_minimo_count,
        }
        
        # Paginator funciona con listas también
        paginator = Paginator(stocks, items_por_pagina)
        
        try:
            stocks_paginados = paginator.page(page)
        except PageNotAnInteger:
            stocks_paginados = paginator.page(1)
        except EmptyPage:
            stocks_paginados = paginator.page(paginator.num_pages)

        context = {
            **self.admin_site.each_context(request),
            'title': _('Reportes de ENTRADA, SALIDA Y TRASPASO de ALMACENES'),
            'stocks': stocks_paginados,
            'estadisticas': estadisticas,  # ⭐ ESTA ES LA LÍNEA CLAVE
            'productos_bajo_minimo': productos_bajo_stock,
            'resumen_almacenes': resumen_almacenes,
            'valoracion': valoracion,
            'almacenes': Almacen.objects.filter(activo=True),
            'categorias': Categoria.objects.all(),
            'productos': Producto.objects.filter(activo=True).order_by('codigo'),
            'filtros': {
                'vista': vista,
                'almacen': almacen_id,
                'categoria': categoria_id,
                'producto': producto_id,
                'stock_minimo': stock_minimo,
                'solo_con_stock': solo_con_stock,
            },
            'page_obj': stocks_paginados,
            'paginator': paginator,
            'items_por_pagina': items_por_pagina,
            'opciones_items': [50, 100, 200, 500],
            'opts': self.model._meta,
        }
        
        if extra_context:
            context.update(extra_context)
        
        return render(request, self.change_list_template, context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            # Exportación
            path('exportar-excel/',
                self.admin_site.admin_view(self.exportar_excel),
                name='reportes_stock_exportar_excel'),
            path('exportar-csv/',
                self.admin_site.admin_view(self.exportar_csv),
                name='reportes_stock_exportar_csv'),

            # ✅ ESTAS SON LAS URLs NECESARIAS PARA LOS MODALES
            path('obtener-detalle-estadistica/',
                self.admin_site.admin_view(views.obtener_detalle_estadistica),
                name='reportes_reportestock_obtener_estadistica'),
            
            path('obtener-detalle-stock/',
                self.admin_site.admin_view(views.obtener_detalle_stock),
                name='reportes_reportestock_obtener_detalle'),
            
            path('obtener-detalle-almacen/',
                self.admin_site.admin_view(views.obtener_detalle_almacen),
                name='reportes_reportestock_obtener_almacen'),
            
            path('obtener-detalle-producto-almacenes/',
                self.admin_site.admin_view(views.obtener_detalle_producto_almacenes),
                name='reportes_reportestock_obtener_producto_almacenes'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        
        # Filtros
        vista = request.GET.get('vista', 'detallado')
        almacen_id = request.GET.get('almacen', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        stock_minimo = request.GET.get('stock_minimo', '') == 'on' # Checkbox devuelve 'on'
        solo_con_stock = request.GET.get('solo_con_stock', '') == 'on'
        
        # Paginación
        page = request.GET.get('page', 1)
        items_por_pagina = int(request.GET.get('items_por_pagina', '100'))

        # ========================================================
        # 🚀 OPTIMIZACIÓN: LLAMADA UNICA MASIVA
        # ========================================================
        almacen_id_int = int(almacen_id) if almacen_id else None
        categoria_id_int = int(categoria_id) if categoria_id else None
        producto_id_int = int(producto_id) if producto_id else None

        # Obtenemos TODOS los datos ya calculados en una lista de diccionarios
        dataset_completo = ReporteStock.obtener_data_stock_masivo(
            almacen_id=almacen_id_int,
            categoria_id=categoria_id_int,
            producto_id=producto_id_int,
            stock_minimo=stock_minimo,
            solo_con_stock=solo_con_stock
        )
        
        # Procesamiento de Vistas (Agrupación en Python, no en DB)
        stocks_display = []
        estadisticas = {
            'stock_bueno_total': Decimal(0),
            'stock_danado_total': Decimal(0),
            'productos_bajo_minimo': 0,
            'total_almacenes': 0 # Se calcula abajo
        }

        # Calcular totales globales rápidos recorriendo la lista una vez
        almacenes_unicos = set()
        
        for item in dataset_completo:
            estadisticas['stock_bueno_total'] += item['stock_bueno']
            estadisticas['stock_danado_total'] += item['stock_danado']
            almacenes_unicos.add(item['almacen'].id)
            
            p = item['producto']
            if p.stock_minimo and item['stock_bueno'] <= p.stock_minimo:
                estadisticas['productos_bajo_minimo'] += 1

        estadisticas['total_almacenes'] = len(almacenes_unicos)
        estadisticas['total_productos'] = len(dataset_completo) # Aprox

        if vista == 'detallado':
            stocks_display = dataset_completo # Ya viene en el formato correcto
            
        elif vista == 'por_almacen':
            # Agrupar en memoria
            temp_alm = {}
            for item in dataset_completo:
                aid = item['almacen'].id
                if aid not in temp_alm:
                    temp_alm[aid] = {
                        'almacen': item['almacen'],
                        'total_productos': 0,
                        'stock_buena_total': 0,
                        'stock_danada_total': 0,
                        'stock_total': 0
                    }
                temp_alm[aid]['total_productos'] += 1
                temp_alm[aid]['stock_buena_total'] += item['stock_bueno']
                temp_alm[aid]['stock_danada_total'] += item['stock_danado']
                temp_alm[aid]['stock_total'] += item['stock_total']
            stocks_display = list(temp_alm.values())
            
        else: # por_producto
            # Agrupar en memoria
            temp_prod = {}
            for item in dataset_completo:
                pid = item['producto'].id
                if pid not in temp_prod:
                    temp_prod[pid] = {
                        'producto': item['producto'],
                        'total_almacenes': 0,
                        'stock_buena_total': 0,
                        'stock_danada_total': 0,
                        'stock_total': 0
                    }
                temp_prod[pid]['total_almacenes'] += 1
                temp_prod[pid]['stock_buena_total'] += item['stock_bueno']
                temp_prod[pid]['stock_danada_total'] += item['stock_danado']
                temp_prod[pid]['stock_total'] += item['stock_total']
            stocks_display = list(temp_prod.values())

        # Paginación
        paginator = Paginator(stocks_display, items_por_pagina)
        try:
            stocks_paginados = paginator.page(page)
        except:
            stocks_paginados = paginator.page(1)

        # Contexto (El resto se mantiene igual)
        context = {
            **self.admin_site.each_context(request),
            'title': _('Reporte Stock Almacén (Optimizado)'),
            'stocks': stocks_paginados,
            'estadisticas': estadisticas,
            'almacenes': Almacen.objects.filter(activo=True),
            'categorias': Categoria.objects.all(),
            'productos': Producto.objects.filter(activo=True).order_by('codigo'),
            'filtros': {
                'vista': vista, 'almacen': almacen_id, 'categoria': categoria_id,
                'producto': producto_id, 'stock_minimo': request.GET.get('stock_minimo'),
                'solo_con_stock': request.GET.get('solo_con_stock')
            },
            'page_obj': stocks_paginados, 'paginator': paginator,
            'items_por_pagina': items_por_pagina,
            'opts': self.model._meta,
        }
        return render(request, self.change_list_template, context)


    def exportar_excel(self, request):
        """
        Exporta el reporte de stock a un archivo Excel (.xlsx) usando la data masiva optimizada.
        """
        # 1. Recuperar filtros (misma lógica que en changelist_view)
        almacen_id = request.GET.get('almacen', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        stock_minimo = request.GET.get('stock_minimo', '') == 'on'
        solo_con_stock = request.GET.get('solo_con_stock', '') == 'on'
        
        almacen_id_int = int(almacen_id) if almacen_id else None
        categoria_id_int = int(categoria_id) if categoria_id else None
        producto_id_int = int(producto_id) if producto_id else None

        # 2. 🚀 LLAMADA OPTIMIZADA: Obtener TODOS los datos en una lista de diccionarios
        try:
            from .models import ReporteStock
            dataset_completo = ReporteStock.obtener_data_stock_masivo(
                almacen_id=almacen_id_int,
                categoria_id=categoria_id_int,
                producto_id=producto_id_int,
                stock_minimo=stock_minimo,
                solo_con_stock=solo_con_stock
            )
        except Exception as e:
            # Manejar errores de consulta si es necesario
            return HttpResponse(f"Error al obtener datos: {e}", status=500)

        # 3. Preparar el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "ReporteStock"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="337AB7", end_color="337AB7", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        # Encabezados
        headers = [
            _("Almacén"), _("Código Producto"), _("Nombre Producto"), _("Categoría"), 
            _("U/M"), _("Stock Mínimo"), _("Entradas Totales"), _("Salidas Totales"), 
            _("Stock Bueno"), _("Stock Dañado"), _("Stock Total")
        ]
        ws.append(headers)

        # Aplicar estilos a la cabecera
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Ajustar anchos y formato de números
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['K'].width = 15 # Stock Total

        # 4. Iterar sobre la lista optimizada y escribir filas
        for row_num, data in enumerate(dataset_completo, 2):
            producto = data['producto']
            almacen = data['almacen']
            
            row = [
                almacen.nombre,
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.unidad_medida.nombre if producto.unidad_medida else '',
                producto.stock_minimo,
                data['entradas_total'],
                data['salidas_total'],
                data['stock_bueno'],
                data['stock_danado'],
                data['stock_total']
            ]
            ws.append(row)
            
            # Aplicar formato de números (columnas G a K) y alineación
            for col_idx in range(7, 12): 
                col_letter = get_column_letter(col_idx)
                cell = ws[f'{col_letter}{row_num}']
                cell.number_format = '#,##0.00'
                cell.alignment = right_alignment
                
            # Resaltar si está bajo stock mínimo
            if producto.stock_minimo and data['stock_bueno'] <= producto.stock_minimo:
                 for col_idx in range(1, 12):
                    col_letter = get_column_letter(col_idx)
                    ws[f'{col_letter}{row_num}'].fill = PatternFill(start_color="F5C3C2", end_color="F5C3C2", fill_type="solid") # Rojo claro

        # 5. Configurar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ReporteStock_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Guardar el libro de trabajo en la respuesta
        wb.save(response)
        return response

    def exportar_csv(self, request):
        """
        Exporta el reporte de stock a un archivo CSV.
        """
        # 1. Recuperar filtros (misma lógica que en changelist_view)
        almacen_id = request.GET.get('almacen', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        stock_minimo = request.GET.get('stock_minimo', '') == 'on'
        solo_con_stock = request.GET.get('solo_con_stock', '') == 'on'
        
        almacen_id_int = int(almacen_id) if almacen_id else None
        categoria_id_int = int(categoria_id) if categoria_id else None
        producto_id_int = int(producto_id) if producto_id else None

        # 2. 🚀 LLAMADA OPTIMIZADA: Obtener TODOS los datos en una lista de diccionarios
        try:
            from .models import ReporteStock
            dataset_completo = ReporteStock.obtener_data_stock_masivo(
                almacen_id=almacen_id_int,
                categoria_id=categoria_id_int,
                producto_id=producto_id_int,
                stock_minimo=stock_minimo,
                solo_con_stock=solo_con_stock
            )
        except Exception as e:
            return HttpResponse(f"Error al obtener datos: {e}", status=500)
        
        # 3. Preparar la respuesta HTTP para CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="ReporteStock_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        response.write(u'\ufeff'.encode('utf8')) # Escribir BOM para compatibilidad con Excel (UTF-8)
        
        writer = csv.writer(response)

        # 4. Escribir la cabecera
        headers = [
            _("Almacén"), _("Código Producto"), _("Nombre Producto"), _("Categoría"), 
            _("U/M"), _("Stock Mínimo"), _("Entradas Totales"), _("Salidas Totales"), 
            _("Stock Bueno"), _("Stock Dañado"), _("Stock Total")
        ]
        writer.writerow(headers)

        # 5. Iterar sobre la lista optimizada y escribir filas
        for data in dataset_completo:
            producto = data['producto']
            almacen = data['almacen']
            
            row = [
                almacen.nombre,
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.unidad_medida.nombre if producto.unidad_medida else '',
                data['producto'].stock_minimo,
                data['entradas_total'],
                data['salidas_total'],
                data['stock_bueno'],
                data['stock_danado'],
                data['stock_total']
            ]
            writer.writerow(row)

        return response

# ==============================================================================
# REPORTE DE STOCK REAL - ALMACENES (considera clientes)
# ==============================================================================
class ReporteStockRealAdmin(admin.ModelAdmin):
    """
    Admin para reportes de stock REAL de almacenes
    Considera movimientos de almacén Y de clientes
    """
    
    change_list_template = 'admin/reportes/reporte_stock_real_list.html'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False
    
    def has_change_permission(self, request, obj=None):
        return True
    
    def changelist_view(self, request, extra_context=None):
        from reportes.models import ReporteStockReal
        from decimal import Decimal
        
        extra_context = extra_context or {}
        
        # Obtener filtros
        vista = request.GET.get('vista', 'detallado')
        almacen_id = request.GET.get('almacen', '')
        categoria_id = request.GET.get('categoria', '')
        producto_id = request.GET.get('producto', '')
        stock_minimo = request.GET.get('stock_minimo', '')
        solo_con_stock = request.GET.get('solo_con_stock', '')
        
        # 2. Conversión de tipos (necesaria para la llamada)
        almacen_id_int = int(almacen_id) if almacen_id else None
        categoria_id_int = int(categoria_id) if categoria_id else None
        producto_id_int = int(producto_id) if producto_id else None
        
        # =================================================================
        # 🚀 OPTIMIZACIÓN: Inserción aquí
        # =================================================================

        dataset_completo = ReporteStockReal.obtener_data_masiva(
            almacen_id=almacen_id_int, # Usa el int(almacen_id)
            categoria_id=categoria_id_int,
            producto_id=producto_id_int,
            stock_minimo=(stock_minimo=='on'),
            solo_con_stock=(solo_con_stock=='on')
        )
        
        # ✅ LÓGICA DE PAGINACIÓN UNIFICADA
        page = request.GET.get('page', 1)
        items_por_pagina = request.GET.get('items_por_pagina', '100')
        
        try:
            items_por_pagina = int(items_por_pagina)
            if items_por_pagina not in [50, 100, 200, 500]:
                items_por_pagina = 100
        except (ValueError, TypeError):
            items_por_pagina = 100
        
        # Obtener stocks usando el cálculo real
        stocks = []
        productos_bajo_stock = []
        resumen_almacenes = []
        
        # Calcular estadísticas globales
        total_productos_sistema = Producto.objects.filter(activo=True).count()
        total_almacenes_activos = Almacen.objects.filter(activo=True).count()
        
        stock_bueno_total = Decimal('0')
        stock_danado_total = Decimal('0')
        productos_bajo_minimo_count = 0
        
        # Obtener almacenes
        almacenes = Almacen.objects.filter(activo=True)
        if almacen_id:
            almacenes = almacenes.filter(id=almacen_id)
        
        # Obtener productos
        productos = Producto.objects.filter(activo=True)
        if categoria_id:
            productos = productos.filter(categoria_id=categoria_id)
        if producto_id:
            productos = productos.filter(id=producto_id)
        
        # Calcular stocks globales (sin filtros de vista)
        for almacen_calc in Almacen.objects.filter(activo=True):
            for producto_calc in Producto.objects.filter(activo=True):
                stock_data = ReporteStockReal.calcular_stock_real_producto_almacen(
                    producto_calc, almacen_calc
                )
                stock_bueno_total += Decimal(str(stock_data['stock_bueno']))
                stock_danado_total += Decimal(str(stock_data['stock_danado']))
                
                # Verificar productos bajo mínimo
                if hasattr(producto_calc, 'stock_minimo') and producto_calc.stock_minimo and producto_calc.stock_minimo > 0:
                    if stock_data['stock_bueno'] <= producto_calc.stock_minimo:
                        productos_bajo_minimo_count += 1
        
        if vista == 'detallado':
            # Vista detallada: calcular stock para cada almacén y producto
            for almacen in almacenes:
                for producto in productos:
                    stock_data = ReporteStockReal.calcular_stock_real_producto_almacen(
                        producto, almacen
                    )
                    
                    # Aplicar filtros
                    if solo_con_stock and stock_data['stock_total'] == 0:
                        continue
                    if stock_minimo and stock_data['stock_bueno'] > producto.stock_minimo:
                        continue
                    
                    stocks.append({
                        'almacen': almacen,
                        'producto': producto,
                        'stock_bueno': stock_data['stock_bueno'],
                        'stock_danado': stock_data['stock_danado'],
                        'stock_total': stock_data['stock_total'],
                        'entradas_almacen': stock_data['entradas_almacen_total'],
                        'salidas_almacen': stock_data['salidas_almacen_total'],
                        'traslados_recibidos': stock_data['traslados_recibidos_total'],
                        'traslados_enviados': stock_data['traslados_enviados_total'],
                        'entradas_cliente': stock_data['entradas_cliente_total'],
                        'salidas_cliente': stock_data['salidas_cliente_total'],
                    })
                   
        elif vista == 'por_almacen':
            # Resumen por almacén
            for almacen in almacenes:
                total_productos = 0
                stock_buena_total = Decimal('0')
                stock_danada_total = Decimal('0')
                
                for producto in Producto.objects.filter(activo=True):
                    stock_data = ReporteStockReal.calcular_stock_real_producto_almacen(
                        producto, almacen
                    )
                    
                    # --- CORRECCIÓN AQUÍ ---
                    # Antes: if stock_data['stock_total'] > 0:
                    # Ahora: Verificamos que sea diferente de 0 para incluir negativos
                    if stock_data['stock_total'] != 0: 
                        total_productos += 1
                        stock_buena_total += Decimal(str(stock_data['stock_bueno']))
                        stock_danada_total += Decimal(str(stock_data['stock_danado']))
                
                stock_total = stock_buena_total + stock_danada_total
                
                # Mostrar si tiene productos o si no está activo el filtro "solo con stock"
                if total_productos > 0 or not solo_con_stock:
                    stocks.append({
                        'almacen': almacen,
                        'total_productos': total_productos,
                        'stock_buena_total': float(stock_buena_total),
                        'stock_danada_total': float(stock_danada_total),
                        'stock_total': float(stock_total)
                    })
        
        else:  # por_producto
            # Resumen por producto
            for producto in productos:
                stock_buena_total = Decimal('0')
                stock_danada_total = Decimal('0')
                total_almacenes = 0
                
                for almacen in Almacen.objects.filter(activo=True):
                    stock_data = ReporteStockReal.calcular_stock_real_producto_almacen(
                        producto, almacen
                    )
                    
                    # --- CORRECCIÓN AQUÍ ---
                    # Antes: if stock_data['stock_total'] > 0:
                    # Ahora: != 0 para incluir negativos
                    if stock_data['stock_total'] != 0:
                        stock_buena_total += Decimal(str(stock_data['stock_bueno']))
                        stock_danada_total += Decimal(str(stock_data['stock_danado']))
                        total_almacenes += 1
                
                # Mostrar si está en algún almacén o si no está activo el filtro
                if total_almacenes > 0 or not solo_con_stock:
                    stocks.append({
                        'producto': producto,
                        'total_almacenes': total_almacenes,
                        'stock_buena_total': float(stock_buena_total),
                        'stock_danada_total': float(stock_danada_total),
                        'stock_total': float(stock_buena_total + stock_danada_total)
                    })
        
        # Productos bajo stock mínimo
        for almacen in Almacen.objects.filter(activo=True):
            for producto in Producto.objects.filter(activo=True, stock_minimo__gt=0):
                stock_data = ReporteStockReal.calcular_stock_real_producto_almacen(
                    producto, almacen
                )
                
                if stock_data['stock_bueno'] <= producto.stock_minimo:
                    productos_bajo_stock.append({
                        'almacen': almacen,
                        'producto': producto,
                        'stock_actual': stock_data['stock_bueno'],
                        'stock_minimo': producto.stock_minimo,
                        'diferencia': producto.stock_minimo - stock_data['stock_bueno']
                    })
        
        # Ordenar por diferencia
        productos_bajo_stock = sorted(productos_bajo_stock, key=lambda x: x['diferencia'], reverse=True)[:10]
        
        # Resumen por almacén para sidebar
        for almacen in Almacen.objects.filter(activo=True):
            total_productos = 0
            stock_total = Decimal('0')
            
            for producto in Producto.objects.filter(activo=True):
                stock_data = ReporteStockReal.calcular_stock_real_producto_almacen(
                    producto, almacen
                )
                
                if stock_data['stock_total'] > 0:
                    total_productos += 1
                    stock_total += Decimal(str(stock_data['stock_total']))
            
            if total_productos > 0:
                resumen_almacenes.append({
                    'almacen': almacen.nombre,
                    'total_productos': total_productos,
                    'stock_total': float(stock_total)
                })
        
        # Valoración total
        total_items = stock_bueno_total + stock_danado_total
        
        valoracion = {
            'total_productos': total_productos_sistema,
            'total_items': float(total_items),
            'valor_total': 0,
        }
        
        estadisticas = {
            'total_productos': total_productos_sistema,
            'stock_bueno_total': float(stock_bueno_total),
            'stock_danado_total': float(stock_danado_total),
            'total_almacenes': total_almacenes_activos,
            'productos_bajo_minimo': productos_bajo_minimo_count,
        }

        paginator = Paginator(stocks, items_por_pagina)
        
        try:
            stocks_paginados = paginator.page(page)
        except PageNotAnInteger:
            stocks_paginados = paginator.page(1)
        except EmptyPage:
            stocks_paginados = paginator.page(paginator.num_pages)
        
        context = {
            **self.admin_site.each_context(request),
            'title': _('Stock Real de Almacenes (incluye movimientos de clientes)'),
            'stocks': stocks_paginados,
            'estadisticas': estadisticas,
            'productos_bajo_minimo': productos_bajo_stock,
            'resumen_almacenes': resumen_almacenes,
            'valoracion': valoracion,
            'almacenes': Almacen.objects.filter(activo=True),
            'categorias': Categoria.objects.all(),
            'productos': Producto.objects.filter(activo=True).order_by('codigo'),
            'filtros': {
                'vista': vista,
                'almacen': almacen_id,
                'categoria': categoria_id,
                'producto': producto_id,
                'stock_minimo': stock_minimo,
                'solo_con_stock': solo_con_stock,
            },
            'page_obj': stocks_paginados,
            'paginator': paginator,
            'items_por_pagina': items_por_pagina,
            'opciones_items': [50, 100, 200, 500],
            'opts': self.model._meta,
        }
        
        if extra_context:
            context.update(extra_context)
        
        return render(request, self.change_list_template, context)
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('obtener-detalle-stock-real/',
                self.admin_site.admin_view(views.obtener_detalle_stock_real),
                name='reportes_stockreal_obtener_detalle'),
            path('obtener-detalle-almacen-real/',
                self.admin_site.admin_view(views.obtener_detalle_almacen_real),
                name='reportes_stockreal_obtener_detalle_almacen'),
            path('obtener-detalle-producto-almacenes-real/',
                self.admin_site.admin_view(views.obtener_detalle_producto_almacenes_real),
                name='reportes_stockreal_obtener_detalle_producto'),
            path('obtener-detalle-estadistica-real/',
                self.admin_site.admin_view(views.obtener_detalle_estadistica_real),
                name='reportes_stockreal_obtener_estadistica'),
            path('exportar-excel/',
                self.admin_site.admin_view(views.exportar_stock_real_excel),
                name='reportes_stockreal_exportar_excel'),
            path('exportar-csv/',
                self.admin_site.admin_view(views.exportar_stock_real_csv),
                name='reportes_stockreal_exportar_csv'),
        ]
        return custom_urls + urls

# Registrar todos los reportes
admin.site.register(ReporteMovimiento, ReporteMovimientoAdmin)
admin.site.register(ReporteStock, ReporteStockAdmin)
admin.site.register(ReporteStockReal, ReporteStockRealAdmin)
admin.site.register(ReporteEntregas, ReporteEntregasAdmin)