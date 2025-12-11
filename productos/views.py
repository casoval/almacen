from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import Producto, Categoria, UnidadMedida
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
from io import BytesIO

def next_code(request):
    tipo = request.GET.get('tipo')
    if not tipo:
        return JsonResponse({'codigo': '-'})
    
    prefijos = {'INSUMOS':'I','EQUIPOS':'E','HERRAMIENTAS':'H','OTROS':'O'}
    prefijo = prefijos.get(tipo, 'P')
    
    ultimo = Producto.objects.filter(tipo=tipo).order_by('-codigo').first()
    if ultimo and ultimo.codigo:
        try:
            num = int(ultimo.codigo[1:]) + 1
        except:
            num = 1
    else:
        num = 1
    
    codigo = f"{prefijo}{num:04d}"
    return JsonResponse({'codigo': codigo})


def exportar_productos(request):
    """Exportar todos los productos a Excel"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Encabezados
    headers = ['Código', 'Categoría', 'Nombre', 'Unidad de Medida']
    ws.append(headers)
    
    # Estilos para encabezados
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Obtener productos
    productos = Producto.objects.all().select_related('categoria', 'unidad_medida')
    
    # Agregar datos
    for producto in productos:
        ws.append([
            producto.codigo,
            producto.categoria.nombre if producto.categoria else '',
            producto.nombre,
            producto.unidad_medida.nombre if producto.unidad_medida else ''
        ])
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    
    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos_exportados.xlsx"'
    
    return response


def importar_productos(request):
    """Vista para importar productos desde Excel - Paso a paso"""
    
    if request.method == 'POST':
        paso = request.POST.get('paso', '1')
        
        # PASO 1: Subir archivo
        if paso == '1':
            if 'archivo' not in request.FILES:
                messages.error(request, '❌ Debe seleccionar un archivo Excel')
                return render(request, 'admin/productos/importar_paso1.html')
            
            archivo = request.FILES['archivo']
            
            # Validar extensión
            if not archivo.name.endswith(('.xlsx', '.xls')):
                messages.error(request, '❌ El archivo debe ser formato Excel (.xlsx o .xls)')
                return render(request, 'admin/productos/importar_paso1.html')
            
            try:
                # Leer Excel
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                
                # Procesar datos (saltando fila 1 de encabezados)
                datos_excel = []
                filas_con_error = []
                
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Fila completamente vacía
                        continue
                    
                    codigo = str(row[0]).strip() if row[0] else ''
                    categoria = str(row[1]).strip() if row[1] else ''
                    nombre = str(row[2]).strip() if row[2] else ''
                    unidad = str(row[3]).strip() if row[3] else ''
                    
                    # Validar campos obligatorios
                    if not nombre or not unidad:
                        filas_con_error.append({
                            'fila': idx,
                            'datos': row,
                            'razon': 'Falta nombre o unidad de medida'
                        })
                        continue
                    
                    datos_excel.append({
                        'fila': idx,
                        'codigo': codigo,
                        'categoria': categoria if categoria else 'VARIOS',
                        'nombre': nombre,
                        'unidad': unidad
                    })
                
                if not datos_excel:
                    messages.error(request, '❌ No se encontraron datos válidos en el archivo')
                    return render(request, 'admin/productos/importar_paso1.html')
                
                # Guardar en sesión
                request.session['datos_importacion'] = datos_excel
                request.session['filas_error'] = filas_con_error
                
                return render(request, 'admin/productos/importar_paso2.html', {
                    'total_filas': len(datos_excel),
                    'filas_error': len(filas_con_error)
                })
                
            except Exception as e:
                messages.error(request, f'❌ Error al leer el archivo: {str(e)}')
                return render(request, 'admin/productos/importar_paso1.html')
        
        # PASO 2: Modo de importación
        elif paso == '2':
            modo = request.POST.get('modo')
            if not modo:
                messages.error(request, '❌ Debe seleccionar un modo de importación')
                return redirect('admin:productos_producto_importar')
            
            request.session['modo_importacion'] = modo
            
            return render(request, 'admin/productos/importar_paso3.html')
        
        # PASO 3: Método de códigos
        elif paso == '3':
            metodo_codigo = request.POST.get('metodo_codigo')
            tipo_producto = request.POST.get('tipo_producto', '')
            
            if not metodo_codigo:
                messages.error(request, '❌ Debe seleccionar un método para los códigos')
                return redirect('admin:productos_producto_importar')
            
            request.session['metodo_codigo'] = metodo_codigo
            request.session['tipo_producto'] = tipo_producto
            
            # Generar vista previa
            datos = request.session.get('datos_importacion', [])
            vista_previa = generar_vista_previa(
                datos[:20],  # Solo primeros 20
                metodo_codigo,
                tipo_producto,
                request.session.get('modo_importacion')
            )
            
            return render(request, 'admin/productos/importar_paso4.html', {
                'vista_previa': vista_previa,
                'total_registros': len(datos)
            })
        
        # PASO 4: Confirmar y procesar
        elif paso == '4':
            confirmar = request.POST.get('confirmar')
            if confirmar != 'si':
                messages.warning(request, '⚠️ Importación cancelada')
                return redirect('admin:productos_producto_changelist')
            
            # Procesar importación
            resultado = procesar_importacion(
                request.session.get('datos_importacion', []),
                request.session.get('modo_importacion'),
                request.session.get('metodo_codigo'),
                request.session.get('tipo_producto', '')
            )
            
            # Limpiar sesión
            for key in ['datos_importacion', 'filas_error', 'modo_importacion', 'metodo_codigo', 'tipo_producto']:
                if key in request.session:
                    del request.session[key]
            
            return render(request, 'admin/productos/importar_paso5.html', {
                'resultado': resultado
            })
    
    # GET - Mostrar paso 1
    return render(request, 'admin/productos/importar_paso1.html')


def generar_vista_previa(datos, metodo_codigo, tipo_producto, modo):
    """Genera vista previa de cómo quedarán los productos"""
    vista_previa = []
    
    for item in datos:
        # Determinar código que se usará
        if metodo_codigo == 'reasignar':
            prefijos = {'INSUMOS':'I','EQUIPOS':'E','HERRAMIENTAS':'H','OTROS':'O'}
            prefijo = prefijos.get(tipo_producto, 'P')
            codigo_nuevo = f"{prefijo}####"  # Placeholder
        else:
            codigo_excel = item['codigo']
            if codigo_excel:
                codigo_nuevo = validar_y_ajustar_codigo(codigo_excel)
            else:
                codigo_nuevo = "AUTO####"
        
        vista_previa.append({
            'codigo_original': item['codigo'] if item['codigo'] else '-',
            'codigo_nuevo': codigo_nuevo,
            'nombre': item['nombre'],
            'categoria': item['categoria'],
            'unidad': item['unidad']
        })
    
    return vista_previa


def validar_y_ajustar_codigo(codigo):
    """Valida y ajusta el formato del código"""
    if not codigo:
        return None
    
    # Extraer letra y números
    match = re.match(r'^([A-Za-z])(\d+)$', codigo)
    if not match:
        return None
    
    letra = match.group(1).upper()
    numero = match.group(2)
    
    # Ajustar a 4 dígitos
    if len(numero) < 4:
        numero = numero.zfill(4)
    elif len(numero) > 4:
        numero = '9999'
    
    return f"{letra}{numero}"


def procesar_importacion(datos, modo, metodo_codigo, tipo_producto):
    """Procesa la importación de productos"""
    resultado = {
        'exitosos': 0,
        'saltados': 0,
        'errores': [],
        'categorias_creadas': [],
        'unidades_creadas': [],
        'codigos_ajustados': [],
        'productos_por_tipo': {'I': 0, 'E': 0, 'H': 0, 'O': 0},
        'productos_protegidos': 0
    }
    
    try:
        with transaction.atomic():
            # Si modo es reemplazar, intentar borrar todos los productos
            if modo == 'reemplazar':
                # Verificar si hay productos en uso
                from almacenes.models import DetalleMovimientoAlmacen
                from beneficiarios.models import DetalleMovimientoCliente
                
                productos_en_uso = []
                
                for producto in Producto.objects.all():
                    en_uso = (
                        DetalleMovimientoAlmacen.objects.filter(producto=producto).exists() or
                        DetalleMovimientoCliente.objects.filter(producto=producto).exists()
                    )
                    if en_uso:
                        productos_en_uso.append(producto.codigo)
                
                if productos_en_uso:
                    # NO se puede hacer reemplazo total
                    resultado['error_general'] = (
                        f"❌ No se puede reemplazar la base de datos porque hay {len(productos_en_uso)} "
                        f"producto(s) en uso en movimientos. "
                        f"Productos protegidos: {', '.join(productos_en_uso[:10])}"
                        f"{'...' if len(productos_en_uso) > 10 else ''}. "
                        f"Por favor, use el modo 'Importar solo nuevos' en su lugar."
                    )
                    resultado['productos_protegidos'] = len(productos_en_uso)
                    return resultado
                
                # Si no hay productos en uso, proceder con eliminación
                Producto.objects.all().delete()
            
            for item in datos:
                try:
                    # Obtener o crear categoría
                    categoria_nombre = item['categoria']
                    categoria, created = Categoria.objects.get_or_create(
                        nombre=categoria_nombre,
                        defaults={'descripcion': 'Creada automáticamente durante importación'}
                    )
                    if created:
                        resultado['categorias_creadas'].append(categoria_nombre)
                    
                    # CORREGIDO: Buscar unidad por nombre O abreviatura
                    unidad_nombre = item['unidad']
                    unidad = None
                    
                    # Intentar encontrar por nombre exacto
                    try:
                        unidad = UnidadMedida.objects.get(nombre__iexact=unidad_nombre)
                    except UnidadMedida.DoesNotExist:
                        # Intentar encontrar por abreviatura
                        try:
                            unidad = UnidadMedida.objects.get(abreviatura__iexact=unidad_nombre)
                        except UnidadMedida.DoesNotExist:
                            # No existe, crear nueva
                            unidad = UnidadMedida.objects.create(
                                nombre=unidad_nombre,
                                abreviatura=unidad_nombre
                            )
                            resultado['unidades_creadas'].append(unidad_nombre)
                    
                    # Si modo es solo nuevos, verificar duplicados
                    if modo == 'solo_nuevos':
                        existe = Producto.objects.filter(
                            nombre__iexact=item['nombre'],
                            unidad_medida=unidad
                        ).exists()
                        
                        if existe:
                            resultado['saltados'] += 1
                            continue
                    
                    # Determinar código
                    if metodo_codigo == 'reasignar':
                        # Generar código según tipo seleccionado
                        prefijos = {'INSUMOS':'I','EQUIPOS':'E','HERRAMIENTAS':'H','OTROS':'O'}
                        prefijo = prefijos.get(tipo_producto, 'P')
                        tipo_final = tipo_producto
                        codigo = None  # Se generará automáticamente
                    else:
                        # Usar código del Excel
                        codigo_excel = item['codigo']
                        if codigo_excel:
                            codigo_ajustado = validar_y_ajustar_codigo(codigo_excel)
                            if codigo_ajustado:
                                # Verificar si existe
                                if Producto.objects.filter(codigo=codigo_ajustado).exists():
                                    # Re-enumerar
                                    letra = codigo_ajustado[0]
                                    codigo = generar_codigo_unico(letra)
                                    resultado['codigos_ajustados'].append({
                                        'original': codigo_excel,
                                        'nuevo': codigo
                                    })
                                else:
                                    codigo = codigo_ajustado
                                
                                # Determinar tipo según inicial
                                inicial_map = {'I':'INSUMOS','E':'EQUIPOS','H':'HERRAMIENTAS','O':'OTROS'}
                                tipo_final = inicial_map.get(codigo[0], 'OTROS')
                            else:
                                # Código inválido, generar automático
                                tipo_final = 'OTROS'
                                codigo = None
                        else:
                            # Sin código, generar automático
                            tipo_final = 'OTROS'
                            codigo = None
                    
                    # Crear producto
                    producto = Producto(
                        tipo=tipo_final,
                        nombre=item['nombre'],
                        categoria=categoria,
                        unidad_medida=unidad,
                        codigo=codigo if codigo else ''
                    )
                    producto.save()
                    
                    resultado['exitosos'] += 1
                    resultado['productos_por_tipo'][producto.codigo[0]] = resultado['productos_por_tipo'].get(producto.codigo[0], 0) + 1
                    
                except Exception as e:
                    resultado['errores'].append({
                        'fila': item['fila'],
                        'nombre': item['nombre'],
                        'error': str(e)
                    })
            
    except Exception as e:
        resultado['error_general'] = str(e)
    
    return resultado


def generar_codigo_unico(letra):
    """Genera un código único con la letra especificada"""
    tipo_map = {'I':'INSUMOS','E':'EQUIPOS','H':'HERRAMIENTAS','O':'OTROS'}
    tipo = tipo_map.get(letra, 'OTROS')
    
    ultimo = Producto.objects.filter(tipo=tipo).order_by('-codigo').first()
    if ultimo and ultimo.codigo:
        try:
            num = int(ultimo.codigo[1:]) + 1
        except:
            num = 1
    else:
        num = 1
    
    return f"{letra}{num:04d}"